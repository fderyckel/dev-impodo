"""Generate the sanitized Product and BoM end-to-end XLSX fixture set."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_DIR = (
    Path(__file__).resolve().parents[1]
    / "scenarios"
    / "fixtures"
    / "file-products-and-boms"
    / "v1"
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FONT = Font(color="0070C0")
THIN_FILL = PatternFill("solid", fgColor="D9EAF7")


ARTICLE_HEADERS = [
    "Numéro d'article",
    "Nom du produit",
    "Nom de recherche",
    "Type de produit",
    "Sous-type de produit",
    "Code Statut Produit",
    "UnitId",
]

ARTICLE_ROWS = [
    ["DEMO-CAP-BLUE", "Demo Blue 26 mm Cap", "DEMO BLUE CAP", "Article", "Produit", 30, "PCE"],
    ["DEMO-CAP-WHITE", "Demo White 26 mm Cap", "DEMO WHITE CAP", "Article", "Produit", 30, "PCE"],
    ["DEMO-RESIN-HDPE", "Demo HDPE Resin", "DEMO HDPE", "Article", "Produit", 30, "g"],
    ["DEMO-COLOR-BLUE", "Demo Blue Colorant", "DEMO BLUE COLOR", "Article", "Produit", 30, "g"],
    ["DEMO-LINER-A", "Demo Liner A", "DEMO LINER A", "Article", "Produit", 30, "PCE"],
    ["DEMO-LINER-B", "Demo Liner B", "DEMO LINER B", "Article", "Produit", 30, "PCE"],
]


BOM_VERSION_HEADERS = [
    "what",
    "RecId",
    "Active",
    "Approved",
    "Approver",
    "BOMId",
    "Construction",
    "FromDate",
    "FromQty",
    "InventDimId",
    "ItemId",
    "Name",
    "PmfBatchSize",
    "PmfBulkParent",
    "PmfCoByVarAllow",
    "PmfFormulaChangeDate",
    "PmfFormulaMultiple",
    "PmfFormulaVersionCalculation",
    "PmfTotalCostAllocation",
    "PmfTypeId",
    "PmfYieldPct",
    "ToDate",
    "modifiedDateTime",
    "dEL_ModifiedTime",
    "modifiedBy",
    "dataAreaId",
    "recVersion",
]


def bom_version_row(
    rec_id: int,
    bom_id: str,
    item_id: str,
    name: str,
    active: str,
    approved: str,
) -> list[object]:
    return [
        9_000_000_000,
        rec_id,
        active,
        approved,
        None,
        bom_id,
        "Non",
        datetime(2026, 1, 1),
        1,
        "DEMO-DIM",
        item_id,
        name,
        1,
        None,
        "Non",
        None,
        0,
        "Non",
        "Non",
        "Nomenclature",
        0,
        None,
        datetime(2026, 1, 1, 12, 0, 0),
        None,
        "fixture.generator",
        "demo",
        1,
    ]


BOM_VERSION_ROWS = [
    bom_version_row(
        9_001_000_001,
        "DEMO-BOM-BLUE",
        "DEMO-CAP-BLUE",
        "Demo Blue Cap - 1000 PCE",
        "Oui",
        "Oui",
    ),
    bom_version_row(
        9_001_000_002,
        "DEMO-BOM-WHITE",
        "DEMO-CAP-WHITE",
        "Demo White Cap - 1000 PCE",
        "Oui",
        "Non",
    ),
]


BOM_HEADERS = [
    "Partition",
    "RecId",
    "BOMConsump",
    "BOMId",
    "BOMQty",
    "BOMQtySerie",
    "BOMType",
    "Calculation",
    "Constant",
    "Density",
    "Depth",
    "EndSchedConsump",
    "Formula",
    "FromDate",
    "Height",
    "InventDimId",
    "ItemBOMId",
    "ItemId",
    "ItemPBAId",
    "ItemRouteId",
    "LineNum",
    "OprNum",
    "PDSBaseValue",
    "PDSIngredientType",
    "PDSInheritCoProductBatchAttrib",
    "PDSInheritCoProductShelfLife",
    "PDSInheritEndItemBatchAttrib",
    "PDSInheritEndItemShelfLife",
    "PmfFormulaPct",
    "PmfPctEnable",
    "PmfPlanGroupId",
    "PmfPlanGroupPriority",
    "PmfScalable",
    "Position",
    "ProdFlushingPrincip",
    "ProjSetSubProdToConsumed",
    "RoundUp",
    "RoundUpQty",
    "ScrapConst",
    "ScrapVar",
    "SILBOMCalcWeight",
    "SILEraseQtyError",
    "SILEraseQtyGood",
    "SILEraseScrapVar",
    "ToDate",
    "UCITDSVirtualItemId",
    "UCSDosePercentage",
    "UCSIsMixture",
    "UnitId",
    "VendId",
    "Width",
    "WrkCtrConsumption",
    "modifiedDateTime",
    "dataAreaId",
    "recVersion",
]


def bom_line_row(
    rec_id: int,
    bom_id: str,
    item_id: str,
    line_num: int,
    position: str,
    bom_qty: float,
    bom_qty_serie: float,
    unit_id: str,
) -> list[object]:
    values: dict[str, object] = {
        "Partition": 9_000_000_000,
        "RecId": rec_id,
        "BOMConsump": "Variable",
        "BOMId": bom_id,
        "BOMQty": bom_qty,
        "BOMQtySerie": bom_qty_serie,
        "BOMType": "Article",
        "Calculation": "Oui",
        "EndSchedConsump": "Non",
        "Formula": "Standard",
        "InventDimId": "DEMO-DIM",
        "ItemId": item_id,
        "LineNum": line_num,
        "OprNum": 20,
        "PDSIngredientType": "Aucun(e)",
        "PDSInheritCoProductBatchAttrib": "Non",
        "PDSInheritCoProductShelfLife": "Non",
        "PDSInheritEndItemBatchAttrib": "Non",
        "PDSInheritEndItemShelfLife": "Non",
        "PmfFormulaPct": 0,
        "PmfPctEnable": "Non",
        "PmfPlanGroupPriority": 0,
        "PmfScalable": "Non",
        "Position": position,
        "ProjSetSubProdToConsumed": "Non",
        "RoundUp": "Non",
        "SILBOMCalcWeight": "Oui",
        "SILEraseQtyError": " ",
        "SILEraseQtyGood": " ",
        "SILEraseScrapVar": "Oui",
        "UCSDosePercentage": 0,
        "UCSIsMixture": "Non",
        "UnitId": unit_id,
        "WrkCtrConsumption": "false",
        "modifiedDateTime": datetime(2026, 1, 1, 12, 0, 0),
        "dataAreaId": "demo",
        "recVersion": 1,
    }
    return [values.get(header) for header in BOM_HEADERS]


# Rows are deliberately out of Position order. LineNum must not become Odoo's
# sequence; Position is the reviewed business ordering field.
BOM_ROWS = [
    bom_line_row(9_002_000_001, "DEMO-BOM-BLUE", "DEMO-LINER-A", 3, "30", 1, 1, "PCE"),
    bom_line_row(9_002_000_002, "DEMO-BOM-BLUE", "DEMO-RESIN-HDPE", 1, "10", 2.95, 1, "G"),
    bom_line_row(9_002_000_003, "DEMO-BOM-BLUE", "DEMO-COLOR-BLUE", 2, "20", 0.05, 1, "g"),
    bom_line_row(9_002_000_004, "DEMO-BOM-WHITE", "DEMO-COLOR-BLUE", 2, "20", 40, 1000, "G"),
    bom_line_row(9_002_000_005, "DEMO-BOM-WHITE", "DEMO-LINER-B", 3, "30", 5000, 5000, "PCE"),
    bom_line_row(9_002_000_006, "DEMO-BOM-WHITE", "DEMO-RESIN-HDPE", 1, "10", 2900, 1000, "g"),
]


def create_workbook(
    file_name: str,
    sheet_name: str,
    headers: list[str],
    rows: list[list[object]],
    table_name: str,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = INPUT_FONT
            cell.alignment = Alignment(vertical="top")

    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(max(map(len, values)) + 2, 10), 32)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width

    worksheet.row_dimensions[1].height = 24
    worksheet.auto_filter.ref = worksheet.dimensions
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    # Mark source identity columns visually without adding non-source sheets.
    for header in ("Numéro d'article", "BOMId", "Position"):
        if header in headers:
            index = headers.index(header) + 1
            for cell in worksheet.iter_cols(
                min_col=index,
                max_col=index,
                min_row=1,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    if item.row > 1:
                        item.fill = THIN_FILL

    workbook.save(OUTPUT_DIR / file_name)


def read_records(file_name: str, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(OUTPUT_DIR / file_name, read_only=True, data_only=False)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    return [dict(zip(headers, row, strict=True)) for row in rows]


def validate_fixture() -> None:
    articles = read_records("DEMO_Article.xlsx", "PLW")
    versions = read_records("DEMO_BOMVersion.xlsx", "Sheet1")
    lines = read_records("DEMO_BOM.xlsx", "Sheet1")

    assert len(articles) == 6
    assert len(versions) == 2
    assert len(lines) == 6

    article_ids = {str(row["Numéro d'article"]) for row in articles}
    assert len(article_ids) == len(articles)

    bom_ids = {str(row["BOMId"]) for row in versions}
    assert len(bom_ids) == len(versions)
    assert {str(row["ItemId"]) for row in versions} <= article_ids
    assert {str(row["BOMId"]) for row in lines} <= bom_ids
    assert {str(row["ItemId"]) for row in lines} <= article_ids

    line_identities = {
        (str(row["BOMId"]), str(row["Position"])) for row in lines
    }
    assert len(line_identities) == len(lines)
    assert {str(row["Position"]) for row in lines} == {"10", "20", "30"}
    assert {int(row["LineNum"]) for row in lines} == {1, 2, 3}

    article_units = {
        str(row["Numéro d'article"]): str(row["UnitId"]) for row in articles
    }
    expected_quantities = {
        ("DEMO-BOM-BLUE", "10"): 2950.0,
        ("DEMO-BOM-BLUE", "20"): 50.0,
        ("DEMO-BOM-BLUE", "30"): 1000.0,
        ("DEMO-BOM-WHITE", "10"): 2900.0,
        ("DEMO-BOM-WHITE", "20"): 40.0,
        ("DEMO-BOM-WHITE", "30"): 1000.0,
    }
    for row in lines:
        identity = (str(row["BOMId"]), str(row["Position"]))
        prepared_quantity = (
            float(row["BOMQty"]) / float(row["BOMQtySerie"]) * 1000
        )
        assert abs(prepared_quantity - expected_quantities[identity]) < 1e-9
        line_unit = "g" if str(row["UnitId"]).lower() == "g" else str(row["UnitId"])
        assert line_unit == article_units[str(row["ItemId"])]


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    create_workbook(
        "DEMO_Article.xlsx",
        "PLW",
        ARTICLE_HEADERS,
        ARTICLE_ROWS,
        "DemoArticleTable",
    )
    create_workbook(
        "DEMO_BOMVersion.xlsx",
        "Sheet1",
        BOM_VERSION_HEADERS,
        BOM_VERSION_ROWS,
        "DemoBomVersionTable",
    )
    create_workbook(
        "DEMO_BOM.xlsx",
        "Sheet1",
        BOM_HEADERS,
        BOM_ROWS,
        "DemoBomLineTable",
    )
    validate_fixture()


if __name__ == "__main__":
    main()
