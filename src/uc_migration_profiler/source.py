"""Tabular source loading and prepared-record construction."""

from __future__ import annotations

import csv
from dataclasses import dataclass, replace
from hashlib import sha256
from pathlib import Path
from pathlib import PurePosixPath
import stat
from typing import Any, Iterable
import zipfile

from .canonical import ValueParseError, parse_field, parse_value
from .models import (
    Issue,
    LogicalReference,
    PreparedRecord,
    ScalarValue,
    Severity,
)
from .profile import (
    DatasetSpec,
    IdentityComponent,
    NormalizationSpec,
    ProfileDocument,
    RelationSpec,
)


@dataclass(frozen=True, slots=True)
class SourceTable:
    dataset: str
    path: Path
    headers: tuple[str, ...]
    rows: tuple["SourceRow", ...]
    content_hash: str


@dataclass(frozen=True, slots=True)
class SourceRow:
    number: int
    values: dict[str, Any]


@dataclass(frozen=True, slots=True)
class PreparedBundle:
    records: tuple[PreparedRecord, ...]
    issues: tuple[Issue, ...]
    source_hashes: dict[str, str]

    def by_dataset(self) -> dict[str, tuple[PreparedRecord, ...]]:
        return {
            dataset: tuple(record for record in self.records if record.dataset == dataset)
            for dataset in dict.fromkeys(record.dataset for record in self.records)
        }


class SourceLoadError(ValueError):
    """Raised when an input file violates the governed source contract."""


MAX_SOURCE_FILE_BYTES = 50 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 10_000
MAX_XLSX_EXPANDED_BYTES = 512 * 1024 * 1024
MAX_XLSX_MEMBER_COMPRESSION_RATIO = 1_000
MAX_XLSX_METADATA_BYTES = 5 * 1024 * 1024
MAX_XLSX_WORKSHEETS = 256
MAX_SOURCE_ROWS = 500_000
MAX_SOURCE_COLUMNS = 2_048
MAX_CELL_STRING_LENGTH = 1_000_000


def load_source_tables(
    profile: ProfileDocument,
    input_directory: str | Path,
) -> tuple[SourceTable, ...]:
    root = Path(input_directory).resolve()
    if not root.is_dir():
        raise SourceLoadError(f"source input directory does not exist: {root}")

    tables: list[SourceTable] = []
    for dataset in profile.datasets:
        path = _contained_source_path(root, dataset.source.file)
        file_size = path.stat().st_size
        if file_size > MAX_SOURCE_FILE_BYTES:
            raise SourceLoadError(
                f"source file exceeds {MAX_SOURCE_FILE_BYTES} bytes: "
                f"{dataset.source.file}"
            )
        data = path.read_bytes()

        try:
            if path.suffix.casefold() == ".csv":
                headers, rows = _load_csv(
                    path,
                    dataset.source.encoding,
                    dataset.source.delimiter,
                )
            else:
                headers, rows = _load_xlsx(
                    path,
                    sheet=dataset.source.sheet or "",
                    header_row=dataset.source.header_row,
                )
        except SourceLoadError:
            raise
        except (csv.Error, LookupError, UnicodeError) as exc:
            raise SourceLoadError(
                f"cannot parse source file {dataset.source.file}: {exc}"
            ) from exc
        tables.append(
            SourceTable(
                dataset=dataset.name,
                path=path,
                headers=headers,
                rows=rows,
                content_hash="sha256:" + sha256(data).hexdigest(),
            )
        )
    return tuple(tables)


def _contained_source_path(root: Path, relative_name: str) -> Path:
    candidate = root / relative_name
    if candidate.is_symlink():
        raise SourceLoadError(f"source file must not be a symlink: {relative_name}")
    path = candidate.resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise SourceLoadError(
            f"source file escapes the input directory: {relative_name}"
        ) from exc
    if not path.is_file():
        raise SourceLoadError(f"source file does not exist: {relative_name}")
    return path


def _load_csv(
    path: Path,
    encoding: str,
    delimiter: str,
) -> tuple[tuple[str, ...], tuple[SourceRow, ...]]:
    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            raise SourceLoadError(f"CSV source has no header row: {path.name}") from exc
        headers = _validate_headers(raw_headers, path.name)
        rows: list[SourceRow] = []
        for values in reader:
            row_number = reader.line_num
            if not values:
                continue
            if len(rows) >= MAX_SOURCE_ROWS:
                raise SourceLoadError(
                    f"source exceeds {MAX_SOURCE_ROWS} data rows: {path.name}"
                )
            if len(values) > len(headers):
                raise SourceLoadError(
                    f"row {row_number} has {len(values)} cells but the header has "
                    f"{len(headers)}: {path.name}"
                )
            padded = [*values, *([None] * (len(headers) - len(values)))]
            _validate_cell_lengths(padded, path.name, row_number)
            rows.append(
                SourceRow(
                    number=row_number,
                    values=dict(zip(headers, padded, strict=True)),
                )
            )
    return headers, tuple(rows)


def _load_xlsx(
    path: Path,
    *,
    sheet: str,
    header_row: int,
) -> tuple[tuple[str, ...], tuple[SourceRow, ...]]:
    _validate_xlsx_container(path)

    try:
        from openpyxl import load_workbook
        from openpyxl.xml import DEFUSEDXML
    except ImportError as exc:
        raise SourceLoadError(
            "XLSX support requires openpyxl and defusedxml"
        ) from exc
    if not DEFUSEDXML:
        raise SourceLoadError("XLSX parsing requires active defusedxml protection")

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise SourceLoadError(f"cannot parse XLSX source: {path.name}") from exc
    try:
        if len(workbook.sheetnames) > MAX_XLSX_WORKSHEETS:
            raise SourceLoadError(
                f"workbook exceeds {MAX_XLSX_WORKSHEETS} worksheets: {path.name}"
            )
        if sheet not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise SourceLoadError(
                f"worksheet {sheet!r} does not exist in {path.name}; "
                f"available sheets: {available}"
            )
        worksheet = workbook[sheet]
        if worksheet.max_column > MAX_SOURCE_COLUMNS:
            raise SourceLoadError(
                f"worksheet {sheet!r} exceeds {MAX_SOURCE_COLUMNS} columns"
            )
        if worksheet.max_row - header_row > MAX_SOURCE_ROWS:
            raise SourceLoadError(
                f"worksheet {sheet!r} exceeds {MAX_SOURCE_ROWS} possible data rows"
            )

        iterator = worksheet.iter_rows(min_row=header_row)
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise SourceLoadError(
                f"worksheet {sheet!r} has no header row {header_row}"
            ) from exc
        _reject_unsafe_cells(header_cells, path.name, header_row)
        headers = _validate_headers(
            [cell.value for cell in header_cells],
            f"{path.name}#{sheet}",
        )

        rows: list[SourceRow] = []
        for cells in iterator:
            row_number = cells[0].row if cells else header_row + len(rows) + 1
            _reject_unsafe_cells(cells, path.name, row_number)
            values = [cell.value for cell in cells[: len(headers)]]
            if len(cells) > len(headers) and any(
                cell.value is not None for cell in cells[len(headers) :]
            ):
                raise SourceLoadError(
                    f"row {row_number} has data beyond the declared headers: "
                    f"{path.name}#{sheet}"
                )
            if not any(value is not None for value in values):
                continue
            if len(rows) >= MAX_SOURCE_ROWS:
                raise SourceLoadError(
                    f"source exceeds {MAX_SOURCE_ROWS} data rows: "
                    f"{path.name}#{sheet}"
                )
            padded = [*values, *([None] * (len(headers) - len(values)))]
            _validate_cell_lengths(padded, path.name, row_number)
            rows.append(
                SourceRow(
                    number=row_number,
                    values=dict(zip(headers, padded, strict=True)),
                )
            )
        return headers, tuple(rows)
    finally:
        workbook.close()


def _validate_headers(raw_headers: Iterable[Any], label: str) -> tuple[str, ...]:
    values = list(raw_headers)
    if not values:
        raise SourceLoadError(f"source has no columns: {label}")
    if len(values) > MAX_SOURCE_COLUMNS:
        raise SourceLoadError(
            f"source exceeds {MAX_SOURCE_COLUMNS} columns: {label}"
        )

    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        if value is None or str(value).strip() == "":
            raise SourceLoadError(f"column {index} has an empty header: {label}")
        header = str(value)
        if len(header) > MAX_CELL_STRING_LENGTH:
            raise SourceLoadError(f"column {index} header is too long: {label}")
        headers.append(header)

    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise SourceLoadError(f"duplicate headers {duplicates!r}: {label}")
    return tuple(headers)


def _validate_cell_lengths(values: Iterable[Any], label: str, row_number: int) -> None:
    for column_number, value in enumerate(values, start=1):
        if isinstance(value, str) and len(value) > MAX_CELL_STRING_LENGTH:
            raise SourceLoadError(
                f"cell at row {row_number}, column {column_number} exceeds "
                f"{MAX_CELL_STRING_LENGTH} characters: {label}"
            )


def _reject_unsafe_cells(cells: Iterable[Any], label: str, row_number: int) -> None:
    for column_number, cell in enumerate(cells, start=1):
        if cell.data_type == "f":
            raise SourceLoadError(
                f"formula cell rejected at row {row_number}, column "
                f"{column_number}: {label}"
            )
        if cell.data_type == "e":
            raise SourceLoadError(
                f"Excel error cell rejected at row {row_number}, column "
                f"{column_number}: {label}"
            )


def _validate_xlsx_container(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise SourceLoadError(
                    f"XLSX archive exceeds {MAX_XLSX_ARCHIVE_ENTRIES} entries: "
                    f"{path.name}"
                )
            names = {member.filename for member in members}
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                raise SourceLoadError(f"file is not a valid XLSX container: {path.name}")

            expanded_bytes = 0
            for member in members:
                member_path = PurePosixPath(member.filename)
                if member_path.is_absolute() or ".." in member_path.parts:
                    raise SourceLoadError(
                        f"unsafe XLSX archive member {member.filename!r}: {path.name}"
                    )
                if member.flag_bits & 0x1:
                    raise SourceLoadError(
                        f"encrypted XLSX archive member rejected: {path.name}"
                    )
                if stat.S_ISLNK(member.external_attr >> 16):
                    raise SourceLoadError(
                        f"symlink XLSX archive member rejected: {path.name}"
                    )
                expanded_bytes += member.file_size
                if expanded_bytes > MAX_XLSX_EXPANDED_BYTES:
                    raise SourceLoadError(
                        f"XLSX expands beyond {MAX_XLSX_EXPANDED_BYTES} bytes: "
                        f"{path.name}"
                    )
                if (
                    member.file_size > 0
                    and (
                        member.compress_size == 0
                        or member.file_size / member.compress_size
                        > MAX_XLSX_MEMBER_COMPRESSION_RATIO
                    )
                ):
                    raise SourceLoadError(
                        f"suspicious XLSX compression ratio in "
                        f"{member.filename!r}: {path.name}"
                    )

            content_types_info = archive.getinfo("[Content_Types].xml")
            if content_types_info.file_size > MAX_XLSX_METADATA_BYTES:
                raise SourceLoadError(
                    f"XLSX content-type metadata is too large: {path.name}"
                )
            content_types = archive.read(content_types_info)
            if b"macroEnabled" in content_types or b"vbaProject" in content_types:
                raise SourceLoadError(
                    f"macro-enabled XLSX content rejected: {path.name}"
                )

            prohibited_prefixes = (
                "xl/externalLinks/",
                "xl/embeddings/",
            )
            prohibited_names = {
                "xl/vbaProject.bin",
                "xl/connections.xml",
            }
            unsafe = sorted(
                name
                for name in names
                if name in prohibited_names
                or any(name.startswith(prefix) for prefix in prohibited_prefixes)
            )
            if unsafe:
                raise SourceLoadError(
                    f"XLSX contains prohibited active or external content "
                    f"{unsafe!r}: {path.name}"
                )
    except zipfile.BadZipFile as exc:
        raise SourceLoadError(
            f"file is not a readable, unencrypted XLSX container: {path.name}"
        ) from exc


def prepare_sources(
    profile: ProfileDocument,
    input_directory: str | Path,
) -> PreparedBundle:
    tables = load_source_tables(profile, input_directory)
    records: list[PreparedRecord] = []
    issues: list[Issue] = []

    for dataset in profile.datasets:
        table = next(item for item in tables if item.dataset == dataset.name)
        missing_headers = sorted(_required_headers(dataset) - set(table.headers))
        if missing_headers:
            issues.append(
                Issue(
                    code="SOURCE_FIELD_MISSING",
                    message=f"missing source headers: {', '.join(missing_headers)}",
                    dataset=dataset.name,
                )
            )
        for row in table.rows:
            records.append(
                _prepare_row(dataset, row.values, row.number, missing_headers)
            )

    records = _mark_duplicate_source_identities(records)
    issues.extend(
        issue
        for record in records
        for issue in record.issues
        if issue.code == "SOURCE_IDENTITY_DUPLICATE"
    )

    return PreparedBundle(
        records=tuple(records),
        issues=tuple(issues),
        source_hashes=dict(
            (
                table.path.relative_to(Path(input_directory).resolve()).as_posix(),
                table.content_hash,
            )
            for table in sorted(tables, key=lambda item: item.dataset)
        ),
    )


def _required_headers(dataset: DatasetSpec) -> set[str]:
    headers = set(dataset.source_identity.fields)
    for component in (
        *dataset.target_identity.components,
        *dataset.target_identity.scope,
    ):
        headers.update(component.source_fields)
    headers.update(field.source for field in dataset.fields.values())
    for relation in dataset.relations.values():
        headers.update(relation.source_fields)
    return headers


def _prepare_row(
    dataset: DatasetSpec,
    row: dict[str, Any],
    row_index: int,
    missing_headers: Iterable[str],
) -> PreparedRecord:
    row_issues: list[Issue] = []
    for header in missing_headers:
        row_issues.append(
            Issue(
                code="SOURCE_FIELD_MISSING",
                message=f"source field {header!r} is not present",
                dataset=dataset.name,
                row=row_index,
                field=header,
            )
        )

    source_identity: list[ScalarValue] = []
    identity_policy = NormalizationSpec(trim=True, empty_as_null=True)
    for field_name in dataset.source_identity.fields:
        try:
            value = parse_value(
                row.get(field_name),
                "string",
                identity_policy,
                required=True,
            )
        except ValueParseError as exc:
            value = None
            row_issues.append(
                Issue(
                    code="SOURCE_IDENTITY_INVALID",
                    message=str(exc),
                    dataset=dataset.name,
                    row=row_index,
                    field=field_name,
                )
            )
        source_identity.append(value)

    scalar_values: dict[str, ScalarValue] = {}
    for target_field, spec in dataset.fields.items():
        try:
            scalar_values[target_field] = parse_field(row.get(spec.source), spec)
        except ValueParseError as exc:
            scalar_values[target_field] = None
            row_issues.append(
                Issue(
                    code=(
                        "SOURCE_REQUIRED_VALUE_MISSING"
                        if "required value" in str(exc)
                        else "SOURCE_TYPE_INVALID"
                    ),
                    message=str(exc),
                    dataset=dataset.name,
                    row=row_index,
                    field=spec.source,
                )
            )

    target_identity = tuple(
        value
        for component in dataset.target_identity.components
        for value in _prepare_identity_component(
            component, row, dataset.name, row_index, row_issues
        )
    )
    target_scope = tuple(
        value
        for component in dataset.target_identity.scope
        for value in _prepare_identity_component(
            component, row, dataset.name, row_index, row_issues
        )
    )

    references: dict[str, Any] = {}
    for target_field, relation in dataset.relations.items():
        references[target_field] = _prepare_relation(
            relation, row, dataset.name, row_index, target_field, row_issues
        )

    return PreparedRecord(
        dataset=dataset.name,
        source_row=row_index,
        target_model=dataset.target.model,
        source_identity=tuple(source_identity),
        target_identity=target_identity,
        target_scope=target_scope,
        scalar_values=scalar_values,
        references=references,
        issues=tuple(row_issues),
    )


def _prepare_identity_component(
    component: IdentityComponent,
    row: dict[str, Any],
    dataset_name: str,
    row_index: int,
    issues: list[Issue],
) -> tuple[Any, ...]:
    if component.resolve is not None:
        key = _parse_reference_key(
            component.source_fields,
            row,
            dataset_name,
            row_index,
            component.target_fields[0],
            issues,
            required=True,
        )
        return (
            LogicalReference(
                origin=component.resolve.origin,
                key=key,
                dataset=component.resolve.dataset,
                model=component.resolve.target_model,
                target_fields=component.resolve.target_fields,
            ),
        )

    values: list[ScalarValue] = []
    for source_field in component.source_fields:
        try:
            value = parse_value(
                row.get(source_field),
                component.type,
                component.normalize,
                required=True,
            )
        except ValueParseError as exc:
            value = None
            issues.append(
                Issue(
                    code="SOURCE_IDENTITY_INVALID",
                    message=str(exc),
                    dataset=dataset_name,
                    row=row_index,
                    field=source_field,
                )
            )
        values.append(value)
    return tuple(values)


def _prepare_relation(
    relation: RelationSpec,
    row: dict[str, Any],
    dataset_name: str,
    row_index: int,
    target_field: str,
    issues: list[Issue],
) -> Any:
    if relation.kind == "many2one":
        key = _parse_reference_key(
            relation.source_fields,
            row,
            dataset_name,
            row_index,
            target_field,
            issues,
            required=relation.required,
        )
        if not key or all(value is None for value in key):
            return None
        return LogicalReference(
            origin=relation.resolve.origin,
            key=key,
            dataset=relation.resolve.dataset,
            model=relation.resolve.target_model,
            target_fields=relation.resolve.target_fields,
        )

    raw = row.get(relation.source_fields[0])
    if raw is None or str(raw).strip() == "":
        if relation.required:
            issues.append(
                Issue(
                    code="SOURCE_REQUIRED_VALUE_MISSING",
                    message="required many2many value is empty",
                    dataset=dataset_name,
                    row=row_index,
                    field=target_field,
                )
            )
        return ()
    keys = [item.strip() for item in str(raw).split(relation.separator)]
    if any(item == "" for item in keys):
        issues.append(
            Issue(
                code="SOURCE_TYPE_INVALID",
                message="many2many value contains an empty item",
                dataset=dataset_name,
                row=row_index,
                field=target_field,
            )
        )
        keys = [item for item in keys if item]
    if len(set(keys)) != len(keys):
        issues.append(
            Issue(
                code="SOURCE_REFERENCE_DUPLICATE",
                message="many2many value contains duplicate business keys",
                dataset=dataset_name,
                row=row_index,
                field=target_field,
            )
        )
    return tuple(
        LogicalReference(
            origin=relation.resolve.origin,
            key=(key,),
            dataset=relation.resolve.dataset,
            model=relation.resolve.target_model,
            target_fields=relation.resolve.target_fields,
        )
        for key in dict.fromkeys(keys)
    )


def _parse_reference_key(
    source_fields: Iterable[str],
    row: dict[str, Any],
    dataset_name: str,
    row_index: int,
    target_field: str,
    issues: list[Issue],
    *,
    required: bool,
) -> tuple[ScalarValue, ...]:
    values: list[ScalarValue] = []
    policy = NormalizationSpec(trim=True, empty_as_null=True)
    for source_field in source_fields:
        try:
            value = parse_value(
                row.get(source_field),
                "string",
                policy,
                required=required,
            )
        except ValueParseError as exc:
            value = None
            issues.append(
                Issue(
                    code="SOURCE_REQUIRED_VALUE_MISSING",
                    message=str(exc),
                    dataset=dataset_name,
                    row=row_index,
                    field=target_field,
                )
            )
        values.append(value)
    return tuple(values)


def _mark_duplicate_source_identities(
    records: list[PreparedRecord],
) -> list[PreparedRecord]:
    indexes: dict[tuple[str, tuple[ScalarValue, ...]], list[int]] = {}
    for index, record in enumerate(records):
        indexes.setdefault((record.dataset, record.source_identity), []).append(index)

    result = list(records)
    for (dataset, identity), record_indexes in indexes.items():
        if len(record_indexes) < 2:
            continue
        for index in record_indexes:
            record = result[index]
            issue = Issue(
                code="SOURCE_IDENTITY_DUPLICATE",
                message=f"source identity {identity!r} occurs {len(record_indexes)} times",
                dataset=dataset,
                row=record.source_row,
                affected_count=len(record_indexes),
            )
            result[index] = replace(record, issues=(*record.issues, issue))
    return result
