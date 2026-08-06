"""Write portable Stage-H evidence and its Excel review projection.

``models.PreflightResult`` is the canonical decision source.  The JSON
manifest is written first and the workbook is then generated locally with the
same controlled Python runtime that Impodo already uses for XLSX intake.  The
workbook is a projection and never feeds conclusions back into the engine.

The manifest contains business identities and portable differences only.
Protected target snapshots—with environment-local numeric Odoo IDs—are stored
separately by the browser ``PreflightRepository`` and never enter the workbook.
The CLI writes no repository state and builds both files from the same result.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import PreflightResult, canonical_json_bytes


MANIFEST_NAME = "impodo_preflight_manifest.json"
WORKBOOK_NAME = "impodo_preflight_report.xlsx"

_COLORS = {
    "brand": "E8473F",
    "brand_dark": "C93C35",
    "charcoal": "494D46",
    "charcoal_dark": "292C28",
    "gray": "868981",
    "paper": "F4F4F1",
    "surface": "FFFFFF",
    "soft": "EEEEEA",
    "line": "D6D7D2",
    "ready": "4D7C5B",
    "ready_text": "315B3B",
    "ready_bg": "EDF7EF",
    "warning": "7D4F00",
    "warning_bg": "FFF5DF",
    "danger": "9F2F2F",
    "danger_bg": "FCE8E7",
    "white": "FFFFFF",
}
_THIN_BORDER = Border(
    bottom=Side(style="thin", color=_COLORS["line"]),
)
_BODY_FONT = Font(color=_COLORS["charcoal"])
_SURFACE_FILL = PatternFill("solid", fgColor=_COLORS["surface"])
_PAPER_FILL = PatternFill("solid", fgColor=_COLORS["paper"])
_EXCEL_MAX_ROW = 1_048_576


class ReportGenerationError(RuntimeError):
    """Raised when canonical report artifacts cannot be generated safely."""


def write_preflight_outputs(
    result: PreflightResult,
    output_directory: str | Path,
    *,
    preview_directory: str | Path | None = None,
) -> tuple[Path, Path]:
    """Write the canonical manifest and its Excel review workbook.

    The manifest and workbook use same-directory temporary files followed by
    atomic replacement.  No Node.js runtime, network request, or browser
    extension is involved.
    """

    output = Path(output_directory)
    output.mkdir(parents=True, exist_ok=True)
    manifest_path = output / MANIFEST_NAME
    workbook_path = output / WORKBOOK_NAME
    manifest = result.to_portable_dict()
    temporary_manifest = manifest_path.with_suffix(".json.partial")
    temporary_manifest.write_bytes(canonical_json_bytes(manifest) + b"\n")
    temporary_manifest.replace(manifest_path)

    _build_workbook(
        manifest_path,
        workbook_path,
        preview_directory=preview_directory,
    )
    return manifest_path, workbook_path


def write_review_workbook(
    manifest_path: str | Path,
    workbook_path: str | Path,
    *,
    preview_directory: str | Path | None = None,
) -> Path:
    """Build the Excel review projection from an existing manifest."""

    manifest = Path(manifest_path)
    workbook = Path(workbook_path)
    if not manifest.is_file():
        raise ReportGenerationError("The readiness manifest does not exist")
    workbook.parent.mkdir(parents=True, exist_ok=True)
    _build_workbook(
        manifest,
        workbook,
        preview_directory=preview_directory,
    )
    return workbook


def _build_workbook(
    manifest_path: Path,
    workbook_path: Path,
    *,
    preview_directory: str | Path | None,
) -> None:
    """Create the business-facing workbook with the bundled Python runtime."""

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as error:
        raise ReportGenerationError("The readiness manifest is invalid") from error

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    sheets = {
        name: workbook.create_sheet(name)
        for name in (
            "Target",
            "Dataset Summary",
            "Proposed Creates",
            "Proposed Updates",
            "Field Differences",
            "Unchanged",
            "Ambiguous Matches",
            "Blocked Records",
            "Reference Resolution",
            "Source Issues",
            "Metadata Coverage",
        )
    }

    decisions = tuple(manifest.get("decisions", ()))
    dataset_names = tuple(dict.fromkeys(item.get("dataset", "") for item in decisions))
    dataset_summary_rows = []
    for dataset in dataset_names:
        rows = tuple(item for item in decisions if item.get("dataset") == dataset)
        dataset_summary_rows.append(
            [
                dataset,
                len(rows),
                _classification_count(rows, "CREATE"),
                _classification_count(rows, "UPDATE"),
                _classification_count(rows, "UNCHANGED"),
                _classification_count(rows, "AMBIGUOUS"),
                _classification_count(rows, "BLOCKED"),
            ]
        )
    _write_data_sheet(
        sheets["Dataset Summary"],
        ["Dataset", "Candidates", "CREATE", "UPDATE", "UNCHANGED", "AMBIGUOUS", "BLOCKED"],
        dataset_summary_rows,
        _COLORS["brand"],
    )

    decision_headers = [
        "Dataset",
        "Source Row",
        "Business Identity",
        "Business Scope",
        "Target Matches",
        "Issues",
    ]
    for sheet_name, classification, accent in (
        ("Proposed Creates", "CREATE", _COLORS["ready"]),
        ("Proposed Updates", "UPDATE", _COLORS["warning"]),
        ("Unchanged", "UNCHANGED", _COLORS["gray"]),
        ("Ambiguous Matches", "AMBIGUOUS", _COLORS["danger"]),
        ("Blocked Records", "BLOCKED", _COLORS["danger"]),
    ):
        _write_data_sheet(
            sheets[sheet_name],
            decision_headers,
            _decision_rows(decisions, classification),
            accent,
        )

    difference_rows = [
        [
            decision.get("dataset"),
            decision.get("source_row"),
            _json_cell(decision.get("business_identity", ())),
            _json_cell(decision.get("business_scope", ())),
            difference.get("field"),
            difference.get("existing"),
            difference.get("proposed"),
            difference.get("comparison_rule"),
            difference.get("material"),
        ]
        for decision in decisions
        for difference in decision.get("differences", ())
    ]
    _write_data_sheet(
        sheets["Field Differences"],
        [
            "Dataset",
            "Source Row",
            "Business Identity",
            "Business Scope",
            "Field",
            "Existing Target",
            "Proposed Source",
            "Comparison Rule",
            "Material",
        ],
        difference_rows,
        _COLORS["warning"],
    )

    reference_rows = [
        [
            item.get("dataset"),
            item.get("field"),
            item.get("reference"),
            item.get("status"),
            item.get("match_count"),
            item.get("affected_count"),
        ]
        for item in manifest.get("reference_resolutions", ())
    ]
    _write_data_sheet(
        sheets["Reference Resolution"],
        ["Dataset", "Field", "Business Reference", "Status", "Matches", "Affected Rows"],
        reference_rows,
        _COLORS["brand"],
    )

    issue_rows = [
        [
            item.get("severity"),
            item.get("code"),
            item.get("dataset"),
            item.get("row"),
            item.get("field"),
            item.get("message"),
            item.get("affected_count"),
        ]
        for item in manifest.get("source_issues", ())
    ]
    _write_data_sheet(
        sheets["Source Issues"],
        ["Severity", "Code", "Dataset", "Source Row", "Field", "Message", "Affected Rows"],
        issue_rows,
        _COLORS["danger"],
    )

    coverage_rows = [
        [
            item.get("dataset"),
            item.get("model"),
            item.get("status"),
            item.get("requested_fields"),
            item.get("available_fields"),
        ]
        for item in manifest.get("metadata_coverage", ())
    ]
    _write_data_sheet(
        sheets["Metadata Coverage"],
        ["Dataset", "Model", "Status", "Requested Fields", "Available Fields"],
        coverage_rows,
        _COLORS["brand"],
    )

    target = manifest.get("target", {})
    _write_data_sheet(
        sheets["Target"],
        ["Attribute", "Value"],
        [
            ["Connection mode", target.get("connection_mode")],
            ["Database", target.get("database")],
            ["Odoo Version", target.get("odoo_version")],
            ["Snapshot Timestamp", target.get("snapshot_timestamp")],
            ["Profile ID", manifest.get("profile", {}).get("id")],
            ["Semantic Hash", manifest.get("semantic_hash")],
            ["Metadata Snapshot Hash", manifest.get("snapshot_hashes", {}).get("metadata")],
            ["Record Snapshot Hash", manifest.get("snapshot_hashes", {}).get("records")],
            ["Module Versions", target.get("module_versions")],
            ["Source Hashes", manifest.get("source_hashes")],
        ],
        _COLORS["brand"],
    )

    _write_dashboard(dashboard, manifest)

    temporary = workbook_path.with_name(f".{workbook_path.name}.partial")
    try:
        workbook.save(temporary)
        temporary.replace(workbook_path)
    except (OSError, ValueError) as error:
        raise ReportGenerationError("Excel review workbook generation failed") from error
    finally:
        temporary.unlink(missing_ok=True)

    if preview_directory is not None:
        _write_csv_previews(workbook, Path(preview_directory))


def _write_dashboard(sheet, manifest: dict[str, Any]) -> None:
    _title_band(
        sheet,
        "Impodo migration review",
        "Read-only comparison evidence - this workbook cannot write to Odoo",
        8,
    )
    sheet.append([])
    decisions = tuple(manifest.get("decisions", ()))
    rows = [
        ["Classification", "Count"],
        ["CREATE", _classification_count(decisions, "CREATE")],
        ["UPDATE", _classification_count(decisions, "UPDATE")],
        ["UNCHANGED", _classification_count(decisions, "UNCHANGED")],
        ["AMBIGUOUS", _classification_count(decisions, "AMBIGUOUS")],
        ["BLOCKED", _classification_count(decisions, "BLOCKED")],
    ]
    for row_index, values in enumerate(rows, start=4):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    _style_header(sheet, 4, 1, 2, _COLORS["brand"])
    classification_styles = (
        (_COLORS["ready_bg"], _COLORS["ready_text"]),
        (_COLORS["warning_bg"], _COLORS["warning"]),
        (_COLORS["soft"], _COLORS["charcoal"]),
        (_COLORS["danger_bg"], _COLORS["danger"]),
        (_COLORS["danger_bg"], _COLORS["danger"]),
    )
    for row_index in range(5, 10):
        fill_color, font_color = classification_styles[row_index - 5]
        for column_index in range(1, 3):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(
                bold=column_index == 1,
                color=font_color,
            )
        sheet.cell(row=row_index, column=2).number_format = "#,##0"

    target = manifest.get("target", {})
    assurance_rows = [
        ["Run assurance", ""],
        ["Connector capability", "Read only"],
        ["Portable IDs", "Numeric Odoo IDs excluded"],
        ["Profile", manifest.get("profile", {}).get("id")],
        ["Target", f"{target.get('connection_mode', '')} / {target.get('database', '')}"],
        ["Semantic hash", manifest.get("semantic_hash")],
    ]
    for row_index, values in enumerate(assurance_rows, start=4):
        sheet.cell(row=row_index, column=4, value=values[0])
        sheet.cell(row=row_index, column=5, value=_safe_cell(values[1]))
        sheet.merge_cells(start_row=row_index, start_column=5, end_row=row_index, end_column=8)
    _style_header(sheet, 4, 4, 8, _COLORS["charcoal_dark"])
    for row_index in range(5, 10):
        sheet.cell(row=row_index, column=4).fill = PatternFill(
            "solid", fgColor=_COLORS["soft"]
        )
        sheet.cell(row=row_index, column=4).font = Font(
            bold=True, color=_COLORS["charcoal_dark"]
        )

    sheet.merge_cells("A11:H11")
    sheet["A11"] = "Start here"
    sheet["A11"].fill = PatternFill("solid", fgColor=_COLORS["brand"])
    sheet["A11"].font = Font(bold=True, color=_COLORS["white"])
    sheet["A11"].alignment = Alignment(vertical="center")
    sheet.merge_cells("A12:H12")
    sheet["A12"] = (
        "Review Proposed Creates and Proposed Updates first. "
        "Use Field Differences to inspect every changed value."
    )
    sheet["A12"].fill = PatternFill("solid", fgColor=_COLORS["paper"])
    sheet["A12"].font = Font(color=_COLORS["charcoal"])
    sheet["A12"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[11].height = 24
    sheet.row_dimensions[12].height = 34

    chart = BarChart()
    chart.title = "Preflight classifications"
    chart.legend = None
    chart.add_data(Reference(sheet, min_col=2, min_row=4, max_row=9), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=5, max_row=9))
    chart.height = 8
    chart.width = 16
    chart.series[0].graphicalProperties.solidFill = _COLORS["brand"]
    chart.series[0].graphicalProperties.line.solidFill = _COLORS["brand_dark"]
    sheet.add_chart(chart, "A14")
    sheet.freeze_panes = "A3"
    for column in range(1, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.column_dimensions["D"].width = 24
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = "A1:H29"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = _COLORS["brand"]


def _write_data_sheet(
    sheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    accent: str,
) -> None:
    if len(rows) > _EXCEL_MAX_ROW - 3:
        raise ReportGenerationError(
            "Excel review workbook exceeds the supported rows on "
            f"the {sheet.title} sheet"
        )
    _title_band(
        sheet,
        sheet.title,
        f"{len(rows):,} review row{'s' if len(rows) != 1 else ''}",
        len(headers),
    )
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column_index, value=header)
    _style_header(sheet, 3, 1, len(headers), accent)

    for row_index, values in enumerate(rows, start=4):
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(
                row=row_index,
                column=column_index,
                value=_safe_cell(value),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
            cell.font = _BODY_FONT
            cell.fill = _SURFACE_FILL if row_index % 2 == 0 else _PAPER_FILL

    if rows:
        sheet.auto_filter.ref = (
            f"A3:{get_column_letter(len(headers))}{len(rows) + 3}"
        )

    for column_index, header in enumerate(headers, start=1):
        lengths = [len(header)]
        lengths.extend(
            len(str(_safe_cell(row[column_index - 1])))
            for row in rows[:100]
            if column_index <= len(row)
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            48,
            max(13, max(lengths) + 3),
        )
    sheet.freeze_panes = "A4"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape" if len(headers) > 5 else "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:3"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = accent


def _title_band(sheet, title: str, subtitle: str, columns: int) -> None:
    last_column = max(1, columns)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(row=1, column=1, value=title)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sheet.cell(row=2, column=1, value=subtitle)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=_COLORS["charcoal_dark"])
        cell.font = Font(bold=True, color=_COLORS["white"], size=16)
    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor=_COLORS["soft"])
        cell.font = Font(italic=True, color=_COLORS["charcoal"])
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 24


def _style_header(sheet, row: int, start_column: int, end_column: int, color: str) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=_COLORS["white"])
        cell.border = Border(
            left=Side(style="thin", color=_COLORS["line"]),
            right=Side(style="thin", color=_COLORS["line"]),
            top=Side(style="thin", color=_COLORS["line"]),
            bottom=Side(style="thin", color=_COLORS["line"]),
        )
    sheet.row_dimensions[row].height = 30


def _classification_count(rows: Iterable[dict[str, Any]], classification: str) -> int:
    return sum(item.get("classification") == classification for item in rows)


def _decision_rows(
    decisions: Sequence[dict[str, Any]],
    classification: str,
) -> list[list[Any]]:
    return [
        [
            decision.get("dataset"),
            decision.get("source_row"),
            _json_cell(decision.get("business_identity", ())),
            _json_cell(decision.get("business_scope", ())),
            decision.get("target_match_count"),
            "; ".join(
                str(issue.get("code", ""))
                for issue in decision.get("issues", ())
            ),
        ]
        for decision in decisions
        if decision.get("classification") == classification
    ]


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        value = _json_cell(value)
    if isinstance(value, str) and (
        value.lstrip("\t\r\n").startswith(("=", "+", "-", "@"))
        or _looks_like_structured_identifier(value)
    ):
        return f"'{value}"
    return value


def _json_cell(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)


def _looks_like_structured_identifier(value: str) -> bool:
    parts = value.split(".")
    return len(parts) > 1 and all(part.isdigit() for part in parts)


def _write_csv_previews(workbook: Workbook, directory: Path) -> None:
    """Write dependency-free CSV previews for developer verification.

    PNG rendering was the only part of the former report path that required a
    JavaScript runtime.  CSV previews keep this optional verification output
    inspectable without adding a second executable to the workstation.
    """

    directory.mkdir(parents=True, exist_ok=True)
    for sheet in workbook.worksheets:
        safe_name = "-".join(
            token for token in sheet.title.casefold().replace("_", "-").split() if token
        )
        target = directory / f"{safe_name}.csv"
        with target.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)


def read_manifest(path: str | Path) -> dict[str, Any]:
    """Load a previously generated canonical JSON manifest."""

    return json.loads(Path(path).read_text(encoding="utf-8"))
