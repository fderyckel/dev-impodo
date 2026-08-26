"""Build bounded, hash-bound catalogs for Stage B source discovery.

Layer: source-inspection domain plus application service.

``SourceInspectionService`` materializes registered artifacts and calls the
isolated worker used by the browser. ``inspect_source_file`` contains the
profile-free CSV/XLSX inspection logic. It verifies immutable bytes, never
changes the source file, and never requires a mapping profile or Odoo access.
Values exposed to the browser are deliberately sampled and truncated, while
statistics are accumulated in one streaming pass.

See ``docs/architecture/python-code-map.md`` and
``tests/test_inspection.py``.
"""

from __future__ import annotations

import csv
from dataclasses import asdict, dataclass, fields as dataclass_fields, replace
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from hashlib import sha256
import json
from pathlib import Path
import posixpath
import re
from typing import Any, Iterable, Protocol
import zipfile

from defusedxml import ElementTree as SafeElementTree

from .access import Actor, Capability
from .artifacts import DataVersionSourceArtifactStore, ArtifactStoreError
from .workspace_access import WorkspaceAccessService
from .workspace_state import WorkspaceStateError, WorkspaceStatus, SourceFile
from .source import (
    MAX_CELL_STRING_LENGTH,
    MAX_SOURCE_COLUMNS,
    MAX_SOURCE_ROWS,
    MAX_XLSX_METADATA_BYTES,
    MAX_XLSX_WORKSHEETS,
    SourceLoadError,
    validated_xlsx_table_bounds,
    validate_source_file,
)


CATALOG_CONTRACT_VERSION = 2
PREVIEW_ROW_LIMIT = 20
HEADER_SCAN_ROW_LIMIT = 25
DISTINCT_VALUE_LIMIT = 10_000
DISTINCT_TABLE_VALUE_LIMIT = 100_000
DISPLAY_VALUE_LIMIT = 200
MAX_CATALOG_HEADER_LENGTH = 1_000
CSV_SAMPLE_BYTES = 64 * 1024
SUPPORTED_CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252")
SUPPORTED_CSV_DELIMITERS = (",", ";", "\t", "|")

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOCUMENT_REL_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_INTEGER_PATTERN = re.compile(r"^[+-]?\d+$")


class SourceInspectionError(WorkspaceStateError):
    """Raised when a governed source file cannot be cataloged safely."""


@dataclass(frozen=True, slots=True)
class SourceInspectionOptions:
    """User-selected parsing settings used to regenerate a source catalog."""

    encoding: str | None = None
    delimiter: str | None = None
    csv_header_row: int = 1
    worksheet_header_rows: tuple[tuple[str, int], ...] = ()

    def header_row_for(self, table_key: str) -> int | None:
        """Return a governed worksheet header override when one was supplied."""

        return dict(self.worksheet_header_rows).get(table_key)


@dataclass(frozen=True, slots=True)
class NamedTableCatalog:
    """One named Excel table discovered in a worksheet."""

    name: str
    display_name: str
    cell_range: str
    disposition: str = "DISTINCT"
    message: str | None = None

    def __post_init__(self) -> None:
        if self.disposition not in {"DISTINCT", "EQUIVALENT", "INVALID"}:
            raise ValueError("Excel-table disposition is invalid")


@dataclass(frozen=True, slots=True)
class SourceColumnProfile:
    """Bounded statistics and a non-binding candidate type for one column."""

    ordinal: int
    name: str
    candidate_type: str
    null_count: int
    non_null_count: int
    distinct_count: int
    distinct_count_is_exact: bool
    duplicate_count: int | None
    minimum: str | None
    maximum: str | None
    minimum_length: int | None
    maximum_length: int | None


@dataclass(frozen=True, slots=True)
class SourceTableCatalog:
    """Inventory, preview, and statistics for one CSV table or worksheet."""

    table_key: str
    name: str
    kind: str
    hidden: bool
    header_row: int | None
    row_count: int
    column_count: int
    columns: tuple[SourceColumnProfile, ...]
    preview_rows: tuple[tuple[str | None, ...], ...]
    named_tables: tuple[NamedTableCatalog, ...] = ()
    formula_cell_count: int = 0
    error_cell_count: int = 0
    first_formula_cell: str | None = None
    first_formula_column: str | None = None
    first_error_cell: str | None = None
    first_error_column: str | None = None
    merged_range_count: int = 0
    warnings: tuple[str, ...] = ()

    @property
    def worksheet_name(self) -> str:
        """Return the physical worksheet containing this selectable table."""

        if self.kind == "WORKSHEET":
            return self.name
        if self.kind == "NAMED_TABLE" and self.table_key.startswith("table:"):
            return self.table_key.removeprefix("table:").rsplit(":", 1)[0]
        return ""


@dataclass(frozen=True, slots=True)
class SourceFileCatalog:
    """Hash-bound source catalog for one immutable project source file."""

    contract_version: int
    file_id: str
    display_name: str
    source_sha256: str
    source_size_bytes: int
    format: str
    inspected_at: datetime
    encoding: str | None
    delimiter: str | None
    tables: tuple[SourceTableCatalog, ...]
    warnings: tuple[str, ...] = ()

    def __post_init__(self) -> None:
        if self.contract_version != CATALOG_CONTRACT_VERSION:
            raise SourceInspectionError("Source catalog contract version is unsupported")

    @property
    def content_hash(self) -> str:
        """Bind confirmations to the exact generated catalog."""

        return "sha256:" + sha256(self.to_json().encode("utf-8")).hexdigest()

    def to_json(self) -> str:
        """Return deterministic JSON suitable for DuckDB and portable evidence."""

        payload = asdict(self)
        payload["inspected_at"] = self.inspected_at.isoformat()
        return json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        )

    @classmethod
    def from_json(cls, value: str) -> "SourceFileCatalog":
        """Rehydrate a catalog stored by :meth:`to_json`."""

        try:
            payload = json.loads(value)
            _require_dataclass_fields(payload, SourceFileCatalog, "source catalog")
            return cls(
                contract_version=int(payload["contract_version"]),
                file_id=str(payload["file_id"]),
                display_name=str(payload["display_name"]),
                source_sha256=str(payload["source_sha256"]),
                source_size_bytes=int(payload["source_size_bytes"]),
                format=str(payload["format"]),
                inspected_at=datetime.fromisoformat(str(payload["inspected_at"])),
                encoding=(
                    str(payload["encoding"])
                    if payload.get("encoding") is not None
                    else None
                ),
                delimiter=(
                    str(payload["delimiter"])
                    if payload.get("delimiter") is not None
                    else None
                ),
                tables=tuple(
                    _table_from_payload(item) for item in payload.get("tables", ())
                ),
                warnings=tuple(str(item) for item in payload.get("warnings", ())),
            )
        except (KeyError, TypeError, ValueError, json.JSONDecodeError) as error:
            if isinstance(error, SourceInspectionError):
                raise
            raise SourceInspectionError("Stored source catalog is invalid") from error


class WorkspaceCatalogReader(Protocol):
    """Read registered workspace files and lifecycle for source inspection."""

    def get(self, workspace_id: str):
        """Return the project whose immutable artifacts will be inspected."""
        ...


class SourceCatalogRepository(Protocol):
    """Persist complete or per-file hash-bound inspection catalogs."""
    """Structural protocol implemented by the local DuckDB repository."""

    def get_source_catalogs(
        self,
        workspace_id: str,
    ) -> tuple[SourceFileCatalog, ...]:
        """Return current catalogs in registered source-file order."""
        ...

    def save_source_catalogs(
        self,
        workspace_id: str,
        catalogs: Iterable[SourceFileCatalog],
        *,
        actor: Actor,
    ) -> None:
        """Atomically replace the complete catalog set and invalidate dependents."""
        ...

    def save_source_catalog(
        self,
        workspace_id: str,
        catalog: SourceFileCatalog,
        *,
        actor: Actor,
    ) -> None:
        """Replace one exact file catalog and invalidate its confirmations."""
        ...


class SourceInspectionService:
    """Inspect registered files and publish their hash-bound catalogs.

    The service owns authorization and project lifecycle checks. Artifact
    materialization and isolated parsing are boundary operations; the source
    repository owns the atomic replacement and downstream invalidation.
    """

    def __init__(
        self,
        workspace_states: WorkspaceCatalogReader,
        sources: SourceCatalogRepository,
        artifacts: DataVersionSourceArtifactStore,
        authorization: WorkspaceAccessService,
    ) -> None:
        self.workspace_states = workspace_states
        self.sources = sources
        self.artifacts = artifacts
        self.authorization = authorization

    def inspect_project(
        self,
        workspace_id: str,
        *,
        actor: Actor,
    ) -> tuple[SourceFileCatalog, ...]:
        """Reinspect every registered file and replace the complete catalog set."""

        context = self.authorization.require(
            actor,
            Capability.SOURCE_INSPECT,
            workspace_id=workspace_id,
        )
        workspace_state = self.workspace_states.get(workspace_id)
        if workspace_state.status is not WorkspaceStatus.REGISTERED:
            raise SourceInspectionError(
                "Register the migration project before inspecting its sources"
            )

        # Import here to keep multiprocessing bootstrapping independent of the
        # domain module import path.
        from .source_worker import inspect_source_file_isolated

        catalogs: list[SourceFileCatalog] = []
        for source_file in workspace_state.source_files:
            try:
                with self.artifacts.materialize_source(
                    context.data_version_id,
                    source_file.stored_name,
                ) as path:
                    catalogs.append(
                        inspect_source_file_isolated(
                            path,
                            source_file=source_file,
                            options=None,
                            inspector=inspect_source_file,
                            catalog_from_json=SourceFileCatalog.from_json,
                            inspection_error=SourceInspectionError,
                        )
                    )
            except ArtifactStoreError as error:
                raise SourceInspectionError(str(error)) from error
        self.sources.save_source_catalogs(
            workspace_id,
            catalogs,
            actor=actor,
        )
        return tuple(catalogs)

    def inspect_file(
        self,
        workspace_id: str,
        file_id: str,
        *,
        options: SourceInspectionOptions,
        actor: Actor,
    ) -> SourceFileCatalog:
        """Regenerate one catalog with governed user-selected settings."""

        context = self.authorization.require(
            actor,
            Capability.SOURCE_INSPECT,
            workspace_id=workspace_id,
        )
        workspace_state = self.workspace_states.get(workspace_id)
        if workspace_state.status is not WorkspaceStatus.REGISTERED:
            raise SourceInspectionError(
                "Register the migration project before configuring its sources"
            )
        try:
            source_file = next(
                item for item in workspace_state.source_files if item.file_id == file_id
            )
        except StopIteration as error:
            raise SourceInspectionError("Registered source file was not found") from error

        from .source_worker import inspect_source_file_isolated

        try:
            with self.artifacts.materialize_source(
                context.data_version_id,
                source_file.stored_name,
            ) as path:
                catalog = inspect_source_file_isolated(
                    path,
                    source_file=source_file,
                    options=options,
                    inspector=inspect_source_file,
                    catalog_from_json=SourceFileCatalog.from_json,
                    inspection_error=SourceInspectionError,
                )
        except ArtifactStoreError as error:
            raise SourceInspectionError(str(error)) from error
        self.sources.save_source_catalog(
            workspace_id,
            catalog,
            actor=actor,
        )
        return catalog


def inspect_source_file(
    path: str | Path,
    *,
    source_file: SourceFile,
    options: SourceInspectionOptions | None = None,
) -> SourceFileCatalog:
    """Inspect one immutable source file without a profile or Odoo access."""

    source_path = Path(path)
    try:
        actual_size, actual_sha256 = _hash_file(source_path)
        if actual_size != source_file.size_bytes or actual_sha256 != source_file.sha256:
            raise SourceInspectionError(
                "The stored source file no longer matches its registered evidence"
            )
        validate_source_file(source_path)
        if source_path.suffix.casefold() == ".csv":
            encoding, delimiter, table, warnings = _inspect_csv(
                source_path,
                options=options,
            )
            format_name = "CSV"
            tables = (table,)
        elif source_path.suffix.casefold() == ".xlsx":
            tables, warnings = _inspect_xlsx(source_path, options=options)
            encoding = None
            delimiter = None
            format_name = "XLSX"
        else:
            raise SourceInspectionError(
                "Only registered CSV and XLSX files can be inspected"
            )
    except SourceInspectionError:
        raise
    except (SourceLoadError, csv.Error, LookupError, UnicodeError) as error:
        raise SourceInspectionError(str(error)) from error
    except OSError as error:
        raise SourceInspectionError(
            f"Could not read the registered source file {source_file.display_name}"
        ) from error

    return SourceFileCatalog(
        contract_version=CATALOG_CONTRACT_VERSION,
        file_id=source_file.file_id,
        display_name=source_file.display_name,
        source_sha256=source_file.sha256,
        source_size_bytes=source_file.size_bytes,
        format=format_name,
        inspected_at=datetime.now(timezone.utc),
        encoding=encoding,
        delimiter=delimiter,
        tables=tables,
        warnings=warnings,
    )


def _inspect_csv(
    path: Path,
    *,
    options: SourceInspectionOptions | None,
) -> tuple[str, str, SourceTableCatalog, tuple[str, ...]]:
    sample = path.read_bytes()[:CSV_SAMPLE_BYTES]
    detected_encoding, encoding_warning = _detect_csv_encoding(sample)
    encoding = options.encoding if options and options.encoding else detected_encoding
    if encoding not in SUPPORTED_CSV_ENCODINGS:
        raise SourceInspectionError("Choose UTF-8, UTF-8 with BOM, or Windows-1252")
    decoded = sample.decode(encoding)
    detected_delimiter, delimiter_warning = _detect_csv_delimiter(decoded)
    delimiter = (
        options.delimiter
        if options and options.delimiter is not None
        else detected_delimiter
    )
    if delimiter not in SUPPORTED_CSV_DELIMITERS:
        raise SourceInspectionError("Choose comma, semicolon, tab, or pipe delimiter")
    header_row = options.csv_header_row if options else 1
    if header_row < 1 or header_row > HEADER_SCAN_ROW_LIMIT:
        raise SourceInspectionError(
            f"CSV header row must be between 1 and {HEADER_SCAN_ROW_LIMIT}"
        )
    warnings = [
        warning
        for warning in (encoding_warning, delimiter_warning)
        if warning is not None
    ]

    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            for _row_number in range(1, header_row):
                next(reader)
            raw_headers = next(reader)
        except StopIteration as error:
            raise SourceInspectionError(
                f"CSV source has no header row {header_row}"
            ) from error
        if not raw_headers:
            raise SourceInspectionError("CSV source has no columns")
        if len(raw_headers) > MAX_SOURCE_COLUMNS:
            raise SourceInspectionError(
                f"CSV source exceeds {MAX_SOURCE_COLUMNS} columns"
            )

        headers, header_warnings = _catalog_headers(raw_headers)
        warnings.extend(header_warnings)
        distinct_budget = _DistinctBudget()
        accumulators = [
            _ColumnAccumulator(distinct_budget) for _header in headers
        ]
        preview_rows: list[tuple[str | None, ...]] = []
        row_count = 0
        short_rows = 0
        long_rows = 0
        for values in reader:
            if not values or not any(value != "" for value in values):
                continue
            row_count += 1
            if row_count > MAX_SOURCE_ROWS:
                raise SourceInspectionError(
                    f"CSV source exceeds {MAX_SOURCE_ROWS} data rows"
                )
            if len(values) < len(headers):
                short_rows += 1
            if len(values) > len(headers):
                long_rows += 1
            row_values: list[Any] = [
                *values[: len(headers)],
                *([None] * max(0, len(headers) - len(values))),
            ]
            for value in row_values:
                if isinstance(value, str) and len(value) > MAX_CELL_STRING_LENGTH:
                    raise SourceInspectionError(
                        f"CSV cell at row {reader.line_num} exceeds "
                        f"{MAX_CELL_STRING_LENGTH} characters"
                    )
            for accumulator, value in zip(accumulators, row_values, strict=True):
                accumulator.observe(value)
            if len(preview_rows) < PREVIEW_ROW_LIMIT:
                preview_rows.append(
                    tuple(_display_value(value) for value in row_values)
                )

    if short_rows:
        warnings.append(f"{short_rows} data row(s) contain fewer cells than the header")
    if long_rows:
        warnings.append(
            f"{long_rows} data row(s) contain cells beyond the detected header"
        )
    return (
        encoding,
        delimiter,
        SourceTableCatalog(
            table_key="csv",
            name=path.stem,
            kind="CSV",
            hidden=False,
            header_row=header_row,
            row_count=row_count,
            column_count=len(headers),
            columns=tuple(
                accumulator.profile(index, header)
                for index, (header, accumulator) in enumerate(
                    zip(headers, accumulators, strict=True),
                    start=1,
                )
            ),
            preview_rows=tuple(preview_rows),
            warnings=tuple(header_warnings),
        ),
        tuple(warnings),
    )


def _inspect_xlsx(
    path: Path,
    *,
    options: SourceInspectionOptions | None,
) -> tuple[tuple[SourceTableCatalog, ...], tuple[str, ...]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.xml import DEFUSEDXML
    except ImportError as error:
        raise SourceInspectionError(
            "XLSX inspection requires openpyxl and defusedxml"
        ) from error
    if not DEFUSEDXML:
        raise SourceInspectionError(
            "XLSX parsing requires active defusedxml protection"
        )

    metadata = _xlsx_metadata(path)
    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise SourceInspectionError(f"Cannot parse XLSX source {path.name}") from error

    try:
        if len(workbook.sheetnames) > MAX_XLSX_WORKSHEETS:
            raise SourceInspectionError(
                f"Workbook exceeds {MAX_XLSX_WORKSHEETS} worksheets"
            )
        tables: list[SourceTableCatalog] = []
        workbook_warnings: list[str] = []
        for worksheet in workbook.worksheets:
            # Some exporters omit the optional worksheet dimension. In
            # read-only mode openpyxl then exposes None even though iter_rows
            # can stream the cells; the bounded scan below remains authoritative.
            if (
                worksheet.max_column is not None
                and worksheet.max_column > MAX_SOURCE_COLUMNS
            ):
                raise SourceInspectionError(
                    f"Worksheet {worksheet.title!r} exceeds "
                    f"{MAX_SOURCE_COLUMNS} columns"
                )
            if (
                worksheet.max_row is not None
                and worksheet.max_row > MAX_SOURCE_ROWS + HEADER_SCAN_ROW_LIMIT
            ):
                raise SourceInspectionError(
                    f"Worksheet {worksheet.title!r} exceeds "
                    f"{MAX_SOURCE_ROWS} possible data rows"
                )
            sheet_metadata = metadata.get(
                worksheet.title,
                _WorksheetMetadata(),
            )
            table = _inspect_worksheet(
                worksheet,
                hidden=sheet_metadata.hidden,
                named_tables=sheet_metadata.named_tables,
                merged_range_count=sheet_metadata.merged_range_count,
                formula_cell_count=sheet_metadata.formula_cell_count,
                header_row_override=(
                    options.header_row_for(f"sheet:{worksheet.title}")
                    if options
                    else None
                ),
            )
            classified_named_tables = tuple(
                _classify_named_table(
                    worksheet,
                    worksheet_table=table,
                    named_table=named_table,
                )
                for named_table in sheet_metadata.named_tables
            )
            table = replace(table, named_tables=classified_named_tables)
            tables.append(table)
            for named_table in classified_named_tables:
                if named_table.disposition != "DISTINCT":
                    continue
                tables.append(
                    _inspect_named_table(
                        worksheet,
                        named_table=named_table,
                        hidden=sheet_metadata.hidden,
                    )
                )
            if table.hidden:
                workbook_warnings.append(
                    f"Worksheet {worksheet.title!r} is hidden"
                )
        return tuple(tables), tuple(workbook_warnings)
    finally:
        workbook.close()


def _classify_named_table(
    worksheet: Any,
    *,
    worksheet_table: SourceTableCatalog,
    named_table: NamedTableCatalog,
) -> NamedTableCatalog:
    """Classify an Excel table without profiling redundant or unsafe ranges."""

    try:
        bounds = validated_xlsx_table_bounds(named_table.cell_range)
    except SourceLoadError as error:
        return replace(
            named_table,
            disposition="INVALID",
            message=(
                f"Excel table {named_table.display_name!r} was ignored because "
                f"its {error}."
            ),
        )
    minimum_column, header_row, maximum_column, maximum_row = bounds
    if (
        worksheet_table.header_row is not None
        and minimum_column == 1
        and header_row == worksheet_table.header_row
        and maximum_column == worksheet_table.column_count
        and maximum_row == worksheet.max_row
    ):
        return replace(
            named_table,
            disposition="EQUIVALENT",
            message=(
                f"Excel table {named_table.display_name!r} covers the same data "
                "and was combined with this worksheet."
            ),
        )
    return named_table


def _inspect_worksheet(
    worksheet: Any,
    *,
    hidden: bool,
    named_tables: tuple[NamedTableCatalog, ...],
    merged_range_count: int,
    formula_cell_count: int,
    header_row_override: int | None,
) -> SourceTableCatalog:
    first_rows = list(worksheet.iter_rows(max_row=HEADER_SCAN_ROW_LIMIT))
    header_row = header_row_override or _candidate_header_row(first_rows, named_tables)
    if header_row_override is not None and (
        header_row_override < 1
        or header_row_override > HEADER_SCAN_ROW_LIMIT
    ):
        raise SourceInspectionError(
            f"Worksheet {worksheet.title!r} has an invalid header row"
        )
    if header_row is None:
        warnings = ["No non-empty candidate header row was found"]
        if merged_range_count:
            warnings.append(f"{merged_range_count} merged range(s) detected")
        return SourceTableCatalog(
            table_key=f"sheet:{worksheet.title}",
            name=worksheet.title,
            kind="WORKSHEET",
            hidden=hidden,
            header_row=None,
            row_count=0,
            column_count=0,
            columns=(),
            preview_rows=(),
            named_tables=named_tables,
            formula_cell_count=formula_cell_count,
            merged_range_count=merged_range_count,
            warnings=tuple(warnings),
        )

    if header_row <= len(first_rows):
        header_cells = first_rows[header_row - 1]
    else:
        try:
            header_cells = next(
                worksheet.iter_rows(min_row=header_row, max_row=header_row)
            )
        except StopIteration as error:
            raise SourceInspectionError(
                f"Worksheet {worksheet.title!r} has no candidate header row "
                f"{header_row}"
            ) from error
    last_header_column = _last_non_empty_index(
        [cell.value for cell in header_cells]
    )
    if last_header_column > MAX_SOURCE_COLUMNS:
        raise SourceInspectionError(
            f"Worksheet {worksheet.title!r} exceeds {MAX_SOURCE_COLUMNS} columns"
        )
    raw_headers = [
        cell.value for cell in header_cells[:last_header_column]
    ]
    headers, header_warnings = _catalog_headers(raw_headers)
    distinct_budget = _DistinctBudget()
    accumulators = [
        _ColumnAccumulator(distinct_budget) for _header in headers
    ]
    preview_rows: list[tuple[str | None, ...]] = []
    row_count = 0
    selected_formula_cell_count = 0
    error_cell_count = 0
    first_formula_cell: str | None = None
    first_formula_column: str | None = None
    first_error_cell: str | None = None
    first_error_column: str | None = None
    data_beyond_headers = 0

    for index, cell in enumerate(header_cells, start=1):
        column_name = headers[index - 1] if index <= len(headers) else None
        if cell.data_type == "f":
            selected_formula_cell_count += 1
            if first_formula_cell is None:
                first_formula_cell = cell.coordinate
                first_formula_column = column_name
        elif cell.data_type == "e":
            error_cell_count += 1
            if first_error_cell is None:
                first_error_cell = cell.coordinate
                first_error_column = column_name

    for cells in worksheet.iter_rows(min_row=header_row + 1):
        values = [cell.value for cell in cells[: len(headers)]]
        overflow_cells = cells[len(headers) :]
        for index, cell in enumerate(cells, start=1):
            column_name = headers[index - 1] if index <= len(headers) else None
            if cell.data_type == "f":
                selected_formula_cell_count += 1
                if first_formula_cell is None:
                    first_formula_cell = cell.coordinate
                    first_formula_column = column_name
            elif cell.data_type == "e":
                error_cell_count += 1
                if first_error_cell is None:
                    first_error_cell = cell.coordinate
                    first_error_column = column_name
        if any(cell.value is not None for cell in overflow_cells):
            data_beyond_headers += 1
        if not any(value is not None and value != "" for value in values):
            continue
        row_count += 1
        if row_count > MAX_SOURCE_ROWS:
            raise SourceInspectionError(
                f"Worksheet {worksheet.title!r} exceeds "
                f"{MAX_SOURCE_ROWS} data rows"
            )
        padded = [*values, *([None] * (len(headers) - len(values)))]
        for accumulator, value in zip(accumulators, padded, strict=True):
            if isinstance(value, str) and len(value) > MAX_CELL_STRING_LENGTH:
                raise SourceInspectionError(
                    f"Worksheet {worksheet.title!r} contains a cell exceeding "
                    f"{MAX_CELL_STRING_LENGTH} characters"
                )
            accumulator.observe(value)
        if len(preview_rows) < PREVIEW_ROW_LIMIT:
            preview_rows.append(tuple(_display_value(value) for value in padded))

    warnings = list(header_warnings)
    if merged_range_count:
        warnings.append(f"{merged_range_count} merged range(s) detected")
    if data_beyond_headers:
        warnings.append(
            f"{data_beyond_headers} data row(s) contain cells beyond the "
            "candidate header"
        )
    return SourceTableCatalog(
        table_key=f"sheet:{worksheet.title}",
        name=worksheet.title,
        kind="WORKSHEET",
        hidden=hidden,
        header_row=header_row,
        row_count=row_count,
        column_count=len(headers),
        columns=tuple(
            accumulator.profile(index, header)
            for index, (header, accumulator) in enumerate(
                zip(headers, accumulators, strict=True),
                start=1,
            )
        ),
        preview_rows=tuple(preview_rows),
        named_tables=named_tables,
        formula_cell_count=selected_formula_cell_count,
        error_cell_count=error_cell_count,
        first_formula_cell=first_formula_cell,
        first_formula_column=first_formula_column,
        first_error_cell=first_error_cell,
        first_error_column=first_error_column,
        merged_range_count=merged_range_count,
        warnings=tuple(warnings),
    )


def _inspect_named_table(
    worksheet: Any,
    *,
    named_table: NamedTableCatalog,
    hidden: bool,
) -> SourceTableCatalog:
    """Profile an Excel named-table range as an independently selectable table."""

    try:
        minimum_column, header_row, maximum_column, maximum_row = (
            validated_xlsx_table_bounds(named_table.cell_range)
        )
    except SourceLoadError as error:
        raise SourceInspectionError(
            f"Named table {named_table.display_name!r} cannot be inspected: {error}"
        ) from error
    try:
        header_cells = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                min_col=minimum_column,
                max_col=maximum_column,
            )
        )
    except StopIteration as error:
        raise SourceInspectionError(
            f"Named table {named_table.display_name!r} has no header row"
        ) from error
    headers, header_warnings = _catalog_headers(
        cell.value for cell in header_cells
    )
    distinct_budget = _DistinctBudget()
    accumulators = [
        _ColumnAccumulator(distinct_budget) for _header in headers
    ]
    preview_rows: list[tuple[str | None, ...]] = []
    row_count = 0
    formula_count = 0
    error_count = 0
    first_formula_cell: str | None = None
    first_formula_column: str | None = None
    first_error_cell: str | None = None
    first_error_column: str | None = None
    for index, cell in enumerate(header_cells):
        if cell.data_type == "f":
            formula_count += 1
            if first_formula_cell is None:
                first_formula_cell = cell.coordinate
                first_formula_column = headers[index]
        elif cell.data_type == "e":
            error_count += 1
            if first_error_cell is None:
                first_error_cell = cell.coordinate
                first_error_column = headers[index]
    for cells in worksheet.iter_rows(
        min_row=header_row + 1,
        max_row=maximum_row,
        min_col=minimum_column,
        max_col=maximum_column,
    ):
        values = [cell.value for cell in cells]
        if not any(value is not None and value != "" for value in values):
            continue
        row_count += 1
        for index, cell in enumerate(cells):
            if cell.data_type == "f":
                formula_count += 1
                if first_formula_cell is None:
                    first_formula_cell = cell.coordinate
                    first_formula_column = headers[index]
            elif cell.data_type == "e":
                error_count += 1
                if first_error_cell is None:
                    first_error_cell = cell.coordinate
                    first_error_column = headers[index]
        for value in values:
            if isinstance(value, str) and len(value) > MAX_CELL_STRING_LENGTH:
                raise SourceInspectionError(
                    f"Named table {named_table.display_name!r} contains a cell "
                    f"exceeding {MAX_CELL_STRING_LENGTH} characters"
                )
        for accumulator, value in zip(accumulators, values, strict=True):
            accumulator.observe(value)
        if len(preview_rows) < PREVIEW_ROW_LIMIT:
            preview_rows.append(tuple(_display_value(value) for value in values))
    warnings = list(header_warnings)
    return SourceTableCatalog(
        table_key=(
            f"table:{worksheet.title}:{named_table.display_name}"
        ),
        name=named_table.display_name,
        kind="NAMED_TABLE",
        hidden=hidden,
        header_row=header_row,
        row_count=row_count,
        column_count=len(headers),
        columns=tuple(
            accumulator.profile(index, header)
            for index, (header, accumulator) in enumerate(
                zip(headers, accumulators, strict=True),
                start=1,
            )
        ),
        preview_rows=tuple(preview_rows),
        named_tables=(named_table,),
        formula_cell_count=formula_count,
        error_cell_count=error_count,
        first_formula_cell=first_formula_cell,
        first_formula_column=first_formula_column,
        first_error_cell=first_error_cell,
        first_error_column=first_error_column,
        warnings=tuple(warnings),
    )


@dataclass(frozen=True, slots=True)
class _WorksheetMetadata:
    hidden: bool = False
    named_tables: tuple[NamedTableCatalog, ...] = ()
    merged_range_count: int = 0
    formula_cell_count: int = 0


def _xlsx_metadata(path: Path) -> dict[str, _WorksheetMetadata]:
    """Read workbook inventory XML without loading worksheet cells into memory."""

    with zipfile.ZipFile(path) as archive:
        workbook_root = _read_xml(archive, "xl/workbook.xml")
        workbook_relationships = _relationship_targets(
            archive,
            "xl/_rels/workbook.xml.rels",
        )
        result: dict[str, _WorksheetMetadata] = {}
        for sheet in workbook_root.findall(f".//{{{_MAIN_NS}}}sheet"):
            name = str(sheet.attrib.get("name", ""))
            relationship_id = sheet.attrib.get(f"{{{_DOCUMENT_REL_NS}}}id", "")
            target = workbook_relationships.get(relationship_id)
            if not name or not target:
                continue
            sheet_path = _resolve_relationship_target("xl/workbook.xml", target)
            relationship_path = _relationship_part_path(sheet_path)
            sheet_relationships = _relationship_targets(
                archive,
                relationship_path,
                missing_ok=True,
            )
            named_tables: list[NamedTableCatalog] = []
            for relationship_target, relationship_type in sheet_relationships.values():
                if not relationship_type.endswith("/table"):
                    continue
                table_path = _resolve_relationship_target(
                    sheet_path,
                    relationship_target,
                )
                table_root = _read_xml(archive, table_path)
                named_tables.append(
                    NamedTableCatalog(
                        name=str(table_root.attrib.get("name", "")),
                        display_name=str(table_root.attrib.get("displayName", "")),
                        cell_range=str(table_root.attrib.get("ref", "")),
                    )
                )
            merged_ranges, formula_cells = _count_sheet_features(
                archive,
                sheet_path,
            )
            result[name] = _WorksheetMetadata(
                hidden=sheet.attrib.get("state", "visible") != "visible",
                named_tables=tuple(
                    sorted(named_tables, key=lambda item: item.display_name.casefold())
                ),
                merged_range_count=merged_ranges,
                formula_cell_count=formula_cells,
            )
        return result


def _relationship_targets(
    archive: zipfile.ZipFile,
    path: str,
    *,
    missing_ok: bool = False,
) -> dict[str, tuple[str, str]] | dict[str, str]:
    try:
        root = _read_xml(archive, path)
    except KeyError:
        if missing_ok:
            return {}
        raise
    relationships = {
        str(item.attrib.get("Id", "")): (
            str(item.attrib.get("Target", "")),
            str(item.attrib.get("Type", "")),
        )
        for item in root.findall(f".//{{{_PACKAGE_REL_NS}}}Relationship")
    }
    if path == "xl/_rels/workbook.xml.rels":
        return {
            relationship_id: target
            for relationship_id, (target, _relationship_type) in relationships.items()
        }
    return relationships


def _read_xml(archive: zipfile.ZipFile, path: str):
    information = archive.getinfo(path)
    if information.file_size > MAX_XLSX_METADATA_BYTES:
        raise SourceInspectionError(f"XLSX metadata part is too large: {path}")
    return SafeElementTree.fromstring(archive.read(information))


def _count_sheet_features(
    archive: zipfile.ZipFile,
    sheet_path: str,
) -> tuple[int, int]:
    merged_ranges = 0
    formula_cells = 0
    with archive.open(sheet_path) as stream:
        for _event, element in SafeElementTree.iterparse(stream, events=("end",)):
            if element.tag == f"{{{_MAIN_NS}}}mergeCell":
                merged_ranges += 1
            elif element.tag == f"{{{_MAIN_NS}}}f":
                formula_cells += 1
            element.clear()
    return merged_ranges, formula_cells


def _resolve_relationship_target(base_path: str, target: str) -> str:
    if target.startswith("/"):
        resolved = target.lstrip("/")
    else:
        resolved = posixpath.normpath(
            posixpath.join(posixpath.dirname(base_path), target)
        )
    if resolved.startswith("../") or resolved == "..":
        raise SourceInspectionError("XLSX relationship escapes the workbook")
    return resolved


def _relationship_part_path(part_path: str) -> str:
    directory, filename = posixpath.split(part_path)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _candidate_header_row(
    rows: list[tuple[Any, ...]],
    named_tables: tuple[NamedTableCatalog, ...],
) -> int | None:
    named_header_rows = [
        _range_start_row(table.cell_range)
        for table in named_tables
        if table.cell_range
    ]
    if named_header_rows:
        first_named_header = min(named_header_rows)
        if first_named_header >= 1:
            return first_named_header

    best: tuple[int, int] | None = None
    for row_number, cells in enumerate(rows, start=1):
        values = [cell.value for cell in cells]
        populated = [
            str(value).strip()
            for value in values
            if value is not None and str(value).strip()
        ]
        if not populated:
            continue
        text_count = sum(
            1
            for cell in cells
            if isinstance(cell.value, str) and cell.value.strip()
        )
        unique_count = len(set(populated))
        score = (text_count * 4) + (unique_count * 2) + len(populated)
        candidate = (score, -row_number)
        if best is None or candidate > best:
            best = candidate
    return -best[1] if best is not None else None


def _range_start_row(cell_range: str) -> int:
    first_cell = cell_range.split(":", 1)[0].replace("$", "")
    matched = re.search(r"(\d+)$", first_cell)
    return int(matched.group(1)) if matched else 1


def _last_non_empty_index(values: Iterable[Any]) -> int:
    last = 0
    for index, value in enumerate(values, start=1):
        if value is not None and str(value).strip():
            last = index
    return last


def _catalog_headers(values: Iterable[Any]) -> tuple[tuple[str, ...], tuple[str, ...]]:
    headers: list[str] = []
    warnings: list[str] = []
    for index, value in enumerate(values, start=1):
        cleaned = "" if value is None else str(value).strip()
        if not cleaned:
            cleaned = f"Column {index}"
            warnings.append(f"Column {index} has an empty candidate header")
        if len(cleaned) > MAX_CATALOG_HEADER_LENGTH:
            raise SourceInspectionError(
                f"Column {index} candidate header exceeds "
                f"{MAX_CATALOG_HEADER_LENGTH} characters"
            )
        headers.append(cleaned)
    duplicates = sorted(
        {
            header
            for header in headers
            if sum(item.casefold() == header.casefold() for item in headers) > 1
        },
        key=str.casefold,
    )
    if duplicates:
        warnings.append(
            "Duplicate candidate headers require correction before mapping: "
            + ", ".join(duplicates)
        )
    return tuple(headers), tuple(warnings)


class _DistinctBudget:
    def __init__(self) -> None:
        self.remaining = DISTINCT_TABLE_VALUE_LIMIT

    def claim(self) -> bool:
        """Reserve one bounded distinct-value slot across the complete table."""

        if self.remaining <= 0:
            return False
        self.remaining -= 1
        return True


class _ColumnAccumulator:
    def __init__(self, distinct_budget: _DistinctBudget) -> None:
        self.distinct_budget = distinct_budget
        self.null_count = 0
        self.non_null_count = 0
        self.distinct_values: set[str] = set()
        self.distinct_is_exact = True
        self.duplicate_count = 0
        self.minimum_length: int | None = None
        self.maximum_length: int | None = None
        self.kinds: set[str] = set()
        self.minimum_by_kind: dict[str, tuple[Any, str]] = {}
        self.maximum_by_kind: dict[str, tuple[Any, str]] = {}

    def observe(self, value: Any) -> None:
        """Accumulate one cell into bounded type, range, and distinct statistics."""

        if value is None or (isinstance(value, str) and not value.strip()):
            self.null_count += 1
            return
        rendered = _render_value(value)
        if rendered is None:
            self.null_count += 1
            return
        display = _bounded_display(rendered)
        self.non_null_count += 1
        length = len(rendered)
        self.minimum_length = (
            length if self.minimum_length is None else min(self.minimum_length, length)
        )
        self.maximum_length = (
            length if self.maximum_length is None else max(self.maximum_length, length)
        )
        if self.distinct_is_exact:
            distinct_key = sha256(rendered.encode("utf-8")).hexdigest()
            if distinct_key in self.distinct_values:
                self.duplicate_count += 1
            elif (
                len(self.distinct_values) < DISTINCT_VALUE_LIMIT
                and self.distinct_budget.claim()
            ):
                self.distinct_values.add(distinct_key)
            else:
                self.distinct_is_exact = False

        kind, comparison_key = _value_kind(value)
        self.kinds.add(kind)
        minimum = self.minimum_by_kind.get(kind)
        maximum = self.maximum_by_kind.get(kind)
        if minimum is None or comparison_key < minimum[0]:
            self.minimum_by_kind[kind] = (comparison_key, display)
        if maximum is None or comparison_key > maximum[0]:
            self.maximum_by_kind[kind] = (comparison_key, display)

    def profile(self, ordinal: int, name: str) -> SourceColumnProfile:
        """Freeze accumulated statistics into the portable column profile."""

        candidate_type = _candidate_type(self.kinds)
        minimum, maximum = self._range(candidate_type)
        distinct_count = len(self.distinct_values) + (
            0 if self.distinct_is_exact else 1
        )
        return SourceColumnProfile(
            ordinal=ordinal,
            name=name,
            candidate_type=candidate_type,
            null_count=self.null_count,
            non_null_count=self.non_null_count,
            distinct_count=distinct_count,
            distinct_count_is_exact=self.distinct_is_exact,
            duplicate_count=(
                self.duplicate_count if self.distinct_is_exact else None
            ),
            minimum=minimum,
            maximum=maximum,
            minimum_length=self.minimum_length,
            maximum_length=self.maximum_length,
        )

    def _range(self, candidate_type: str) -> tuple[str | None, str | None]:
        if not self.kinds:
            return None, None
        if candidate_type == "decimal":
            entries = [
                self.minimum_by_kind[kind]
                for kind in ("integer", "decimal")
                if kind in self.minimum_by_kind
            ]
            maximum_entries = [
                self.maximum_by_kind[kind]
                for kind in ("integer", "decimal")
                if kind in self.maximum_by_kind
            ]
            return (
                min(entries, key=lambda item: Decimal(str(item[0])))[1],
                max(maximum_entries, key=lambda item: Decimal(str(item[0])))[1],
            )
        if len(self.kinds) == 1:
            kind = next(iter(self.kinds))
            return self.minimum_by_kind[kind][1], self.maximum_by_kind[kind][1]
        return None, None


def _candidate_type(kinds: set[str]) -> str:
    if not kinds:
        return "empty"
    if kinds == {"boolean"}:
        return "boolean"
    if kinds == {"integer"}:
        return "integer"
    if kinds <= {"integer", "decimal"}:
        return "decimal"
    if kinds == {"date"}:
        return "date"
    if kinds <= {"date", "datetime"}:
        return "datetime"
    if kinds == {"string"}:
        return "string"
    return "mixed"


def _value_kind(value: Any) -> tuple[str, Any]:
    if isinstance(value, bool):
        return "boolean", value
    if isinstance(value, datetime):
        normalized = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
        return "datetime", normalized.astimezone(timezone.utc)
    if isinstance(value, date):
        return "date", value
    if isinstance(value, int):
        return "integer", Decimal(value)
    if isinstance(value, (float, Decimal)):
        return "decimal", Decimal(str(value))

    cleaned = str(value).strip()
    lowered = cleaned.casefold()
    if lowered in {"true", "false"}:
        return "boolean", lowered == "true"
    if _INTEGER_PATTERN.fullmatch(cleaned):
        unsigned = cleaned.lstrip("+-")
        if len(unsigned) > 1 and unsigned.startswith("0"):
            return "string", cleaned.casefold()
        return "integer", Decimal(cleaned)
    try:
        decimal_value = Decimal(cleaned)
    except InvalidOperation:
        decimal_value = None
    if decimal_value is not None and any(
        marker in cleaned for marker in (".", "e", "E")
    ):
        return "decimal", decimal_value
    try:
        if "T" in cleaned or " " in cleaned:
            parsed_datetime = datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
            if parsed_datetime.tzinfo is None:
                parsed_datetime = parsed_datetime.replace(tzinfo=timezone.utc)
            return "datetime", parsed_datetime.astimezone(timezone.utc)
        return "date", date.fromisoformat(cleaned)
    except ValueError:
        return "string", cleaned.casefold()


def _display_value(value: Any) -> str | None:
    rendered = _render_value(value)
    return _bounded_display(rendered) if rendered is not None else None


def _render_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _bounded_display(rendered: str) -> str:
    if len(rendered) > DISPLAY_VALUE_LIMIT:
        return rendered[: DISPLAY_VALUE_LIMIT - 1] + "…"
    return rendered


def _detect_csv_encoding(sample: bytes) -> tuple[str, str | None]:
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig", None
    for encoding in SUPPORTED_CSV_ENCODINGS[1:]:
        try:
            sample.decode(encoding)
        except UnicodeDecodeError:
            continue
        warning = (
            "Encoding was detected as Windows-1252 and should be confirmed"
            if encoding == "cp1252"
            else None
        )
        return encoding, warning
    raise SourceInspectionError(
        "CSV encoding is not supported; use UTF-8 or Windows-1252"
    )


def _detect_csv_delimiter(sample: str) -> tuple[str, str | None]:
    try:
        dialect = csv.Sniffer().sniff(
            sample,
            delimiters="".join(SUPPORTED_CSV_DELIMITERS),
        )
        return dialect.delimiter, None
    except csv.Error:
        first_line = sample.splitlines()[0] if sample.splitlines() else ""
        counts = {
            delimiter: first_line.count(delimiter)
            for delimiter in SUPPORTED_CSV_DELIMITERS
        }
        delimiter = max(counts, key=counts.get)
        if counts[delimiter] == 0:
            raise SourceInspectionError(
                "CSV delimiter could not be detected"
            )
        return (
            delimiter,
            "CSV delimiter detection was uncertain and should be confirmed",
        )


def _hash_file(path: Path) -> tuple[int, str]:
    digest = sha256()
    size = 0
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            size += len(chunk)
            digest.update(chunk)
    return size, digest.hexdigest()


def _table_from_payload(payload: dict[str, Any]) -> SourceTableCatalog:
    _require_dataclass_fields(payload, SourceTableCatalog, "source table")
    for column in payload["columns"]:
        _require_dataclass_fields(column, SourceColumnProfile, "source column")
    for table in payload["named_tables"]:
        _require_dataclass_fields(table, NamedTableCatalog, "named table")
    return SourceTableCatalog(
        table_key=str(payload["table_key"]),
        name=str(payload["name"]),
        kind=str(payload["kind"]),
        hidden=bool(payload["hidden"]),
        header_row=(
            int(payload["header_row"])
            if payload.get("header_row") is not None
            else None
        ),
        row_count=int(payload["row_count"]),
        column_count=int(payload["column_count"]),
        columns=tuple(
            SourceColumnProfile(
                ordinal=int(column["ordinal"]),
                name=str(column["name"]),
                candidate_type=str(column["candidate_type"]),
                null_count=int(column["null_count"]),
                non_null_count=int(column["non_null_count"]),
                distinct_count=int(column["distinct_count"]),
                distinct_count_is_exact=bool(column["distinct_count_is_exact"]),
                duplicate_count=(
                    int(column["duplicate_count"])
                    if column.get("duplicate_count") is not None
                    else None
                ),
                minimum=(
                    str(column["minimum"])
                    if column.get("minimum") is not None
                    else None
                ),
                maximum=(
                    str(column["maximum"])
                    if column.get("maximum") is not None
                    else None
                ),
                minimum_length=(
                    int(column["minimum_length"])
                    if column.get("minimum_length") is not None
                    else None
                ),
                maximum_length=(
                    int(column["maximum_length"])
                    if column.get("maximum_length") is not None
                    else None
                ),
            )
            for column in payload.get("columns", ())
        ),
        preview_rows=tuple(
            tuple(str(value) if value is not None else None for value in row)
            for row in payload.get("preview_rows", ())
        ),
        named_tables=tuple(
            NamedTableCatalog(
                name=str(table["name"]),
                display_name=str(table["display_name"]),
                cell_range=str(table["cell_range"]),
                disposition=str(table["disposition"]),
                message=(
                    str(table["message"])
                    if table.get("message") is not None
                    else None
                ),
            )
            for table in payload.get("named_tables", ())
        ),
        formula_cell_count=int(payload.get("formula_cell_count", 0)),
        error_cell_count=int(payload.get("error_cell_count", 0)),
        first_formula_cell=(
            str(payload["first_formula_cell"])
            if payload.get("first_formula_cell") is not None
            else None
        ),
        first_formula_column=(
            str(payload["first_formula_column"])
            if payload.get("first_formula_column") is not None
            else None
        ),
        first_error_cell=(
            str(payload["first_error_cell"])
            if payload.get("first_error_cell") is not None
            else None
        ),
        first_error_column=(
            str(payload["first_error_column"])
            if payload.get("first_error_column") is not None
            else None
        ),
        merged_range_count=int(payload.get("merged_range_count", 0)),
        warnings=tuple(str(item) for item in payload.get("warnings", ())),
    )


def _require_dataclass_fields(
    payload: object,
    contract: type[object],
    label: str,
) -> None:
    expected = {item.name for item in dataclass_fields(contract)}
    if not isinstance(payload, dict) or set(payload) != expected:
        raise SourceInspectionError(
            f"Stored {label} does not match the current contract"
        )
