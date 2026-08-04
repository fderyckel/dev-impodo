"""Load governed source files and convert their rows into domain records.

This module is the boundary between user-provided CSV/XLSX files and the
profiler's typed model layer:

1. :func:`load_source_tables` safely reads every file declared by the profile.
2. :func:`prepare_sources` maps source columns to scalar values, identities,
   and unresolved :class:`~impodo.models.LogicalReference`
   objects.
3. The engine later resolves those references against the incoming datasets
   and the read-only Odoo snapshots.

No Odoo lookup occurs here.  Keeping source parsing separate from resolution
makes every conversion issue traceable to its original dataset and row.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, replace
from hashlib import sha256
from pathlib import Path
from pathlib import PurePosixPath
import stat
from typing import Any, Iterable
import zipfile

from .canonical import ValueParseError, parse_field, parse_value
from .models import (
    InvalidPreparedValue,
    Issue,
    LogicalReference,
    PreparedRecord,
    ScalarValue,
    Severity,
)
from .profile import (
    DatasetSpec,
    IdentityComponent,
    NormalizationSpec,
    ProfileDocument,
    RelationSpec,
)


@dataclass(frozen=True, slots=True)
class SourceTable:
    """One validated input table before profile mappings are applied.

    ``rows`` preserves physical row numbers for actionable error reporting,
    while ``content_hash`` allows the final report to identify the exact input
    bytes used for the run.
    """

    dataset: str
    path: Path
    headers: tuple[str, ...]
    rows: tuple["SourceRow", ...]
    content_hash: str


@dataclass(frozen=True, slots=True)
class SourceRow:
    """A source row paired with its one-based CSV or worksheet row number."""

    number: int
    values: dict[str, Any]


@dataclass(frozen=True, slots=True)
class PreparedBundle:
    """All prepared records, global source issues, and source-file hashes."""

    records: tuple[PreparedRecord, ...]
    issues: tuple[Issue, ...]
    source_hashes: dict[str, str]

    def by_dataset(self) -> dict[str, tuple[PreparedRecord, ...]]:
        """Group records by dataset while preserving profile/record order.

        The engine consumes these groups when it resolves dependencies and
        classifies each dataset independently.
        """

        return {
            dataset: tuple(record for record in self.records if record.dataset == dataset)
            for dataset in dict.fromkeys(record.dataset for record in self.records)
        }


class SourceLoadError(ValueError):
    """Raised when an input file violates the governed source contract."""


MAX_SOURCE_FILE_BYTES = 50 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 10_000
MAX_XLSX_EXPANDED_BYTES = 512 * 1024 * 1024
MAX_XLSX_MEMBER_COMPRESSION_RATIO = 1_000
MAX_XLSX_METADATA_BYTES = 5 * 1024 * 1024
MAX_XLSX_WORKSHEETS = 256
MAX_SOURCE_ROWS = 500_000
MAX_SOURCE_COLUMNS = 2_048
MAX_CELL_STRING_LENGTH = 1_000_000


def load_source_tables(
    profile: ProfileDocument,
    input_directory: str | Path,
) -> tuple[SourceTable, ...]:
    """Load and validate every source file declared by ``profile``.

    The resolved files must remain inside ``input_directory`` and satisfy the
    size and format limits defined in this module.  CSV and XLSX parsing is
    deliberately strict so malformed, active, encrypted, or suspicious
    workbook content cannot silently enter the preflight pipeline.

    Returns:
        Tables in the same order as ``profile.datasets``.

    Raises:
        SourceLoadError: If the input directory, a file, or its content
            violates the source contract.
    """

    root = Path(input_directory).resolve()
    if not root.is_dir():
        raise SourceLoadError(f"source input directory does not exist: {root}")

    tables: list[SourceTable] = []
    for dataset in profile.datasets:
        path = _contained_source_path(root, dataset.source.file)
        file_size = path.stat().st_size
        if file_size > MAX_SOURCE_FILE_BYTES:
            raise SourceLoadError(
                f"source file exceeds {MAX_SOURCE_FILE_BYTES} bytes: "
                f"{dataset.source.file}"
            )
        data = path.read_bytes()

        try:
            if path.suffix.casefold() == ".csv":
                headers, rows = _load_csv(
                    path,
                    dataset.source.encoding,
                    dataset.source.delimiter,
                )
            else:
                headers, rows = _load_xlsx(
                    path,
                    sheet=dataset.source.sheet or "",
                    header_row=dataset.source.header_row,
                )
        except SourceLoadError:
            raise
        except (csv.Error, LookupError, UnicodeError) as exc:
            raise SourceLoadError(
                f"cannot parse source file {dataset.source.file}: {exc}"
            ) from exc
        tables.append(
            SourceTable(
                dataset=dataset.name,
                path=path,
                headers=headers,
                rows=rows,
                content_hash="sha256:" + sha256(data).hexdigest(),
            )
        )
    return tuple(tables)


def load_selected_source_table(
    path: str | Path,
    *,
    dataset: str,
    table_key: str,
    encoding: str | None,
    delimiter: str | None,
    header_row: int,
    named_table_range: str | None = None,
) -> SourceTable:
    """Load one frozen browser dataset through the strict source reader.

    Browser selections identify worksheets and named tables independently of
    the profile-driven CLI. This adapter preserves that exact selection while
    reusing the same passive XLSX, row, column, and cell safety limits.
    """

    source_path = Path(path).resolve()
    if not source_path.is_file() or source_path.is_symlink():
        raise SourceLoadError("stored source artifact is unavailable")
    if source_path.stat().st_size > MAX_SOURCE_FILE_BYTES:
        raise SourceLoadError(
            f"source file exceeds {MAX_SOURCE_FILE_BYTES} bytes: {source_path.name}"
        )
    data = source_path.read_bytes()
    suffix = source_path.suffix.casefold()
    if suffix == ".csv":
        if table_key != "csv":
            raise SourceLoadError("CSV dataset selection is invalid")
        headers, rows = _load_csv(
            source_path,
            encoding or "utf-8-sig",
            delimiter or ",",
        )
    elif suffix == ".xlsx":
        if table_key.startswith("sheet:"):
            sheet = table_key.removeprefix("sheet:")
            cell_range = None
        elif table_key.startswith("table:") and named_table_range:
            selected = table_key.removeprefix("table:")
            try:
                sheet, _table_name = selected.rsplit(":", 1)
            except ValueError as error:
                raise SourceLoadError("Named-table selection is invalid") from error
            cell_range = named_table_range
        else:
            raise SourceLoadError("XLSX dataset selection is invalid")
        headers, rows = _load_xlsx(
            source_path,
            sheet=sheet,
            header_row=header_row,
            cell_range=cell_range,
        )
    else:
        raise SourceLoadError("Only CSV and XLSX source files are supported")
    return SourceTable(
        dataset=dataset,
        path=source_path,
        headers=headers,
        rows=rows,
        content_hash="sha256:" + sha256(data).hexdigest(),
    )


def _contained_source_path(root: Path, relative_name: str) -> Path:
    """Resolve one declared source path without permitting link/path escape."""

    candidate = root / relative_name
    if candidate.is_symlink():
        raise SourceLoadError(f"source file must not be a symlink: {relative_name}")
    path = candidate.resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise SourceLoadError(
            f"source file escapes the input directory: {relative_name}"
        ) from exc
    if not path.is_file():
        raise SourceLoadError(f"source file does not exist: {relative_name}")
    return path


def _load_csv(
    path: Path,
    encoding: str,
    delimiter: str,
) -> tuple[tuple[str, ...], tuple[SourceRow, ...]]:
    """Read one CSV file and enforce row, column, and cell-size limits."""

    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            raise SourceLoadError(f"CSV source has no header row: {path.name}") from exc
        headers = _validate_headers(raw_headers, path.name)
        rows: list[SourceRow] = []
        for values in reader:
            row_number = reader.line_num
            if not values:
                continue
            if len(rows) >= MAX_SOURCE_ROWS:
                raise SourceLoadError(
                    f"source exceeds {MAX_SOURCE_ROWS} data rows: {path.name}"
                )
            if len(values) > len(headers):
                raise SourceLoadError(
                    f"row {row_number} has {len(values)} cells but the header has "
                    f"{len(headers)}: {path.name}"
                )
            padded = [*values, *([None] * (len(headers) - len(values)))]
            _validate_cell_lengths(padded, path.name, row_number)
            rows.append(
                SourceRow(
                    number=row_number,
                    values=dict(zip(headers, padded, strict=True)),
                )
            )
    return headers, tuple(rows)


def _load_xlsx(
    path: Path,
    *,
    sheet: str,
    header_row: int,
    cell_range: str | None = None,
) -> tuple[tuple[str, ...], tuple[SourceRow, ...]]:
    """Read one worksheet from a passive, bounded XLSX container.

    Workbook loading is read-only and formulas are rejected instead of being
    calculated or trusted.  The original worksheet row numbers are retained.
    """

    _validate_xlsx_container(path)

    try:
        from openpyxl import load_workbook
        from openpyxl.xml import DEFUSEDXML
    except ImportError as exc:
        raise SourceLoadError(
            "XLSX support requires openpyxl and defusedxml"
        ) from exc
    if not DEFUSEDXML:
        raise SourceLoadError("XLSX parsing requires active defusedxml protection")

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise SourceLoadError(f"cannot parse XLSX source: {path.name}") from exc
    try:
        if len(workbook.sheetnames) > MAX_XLSX_WORKSHEETS:
            raise SourceLoadError(
                f"workbook exceeds {MAX_XLSX_WORKSHEETS} worksheets: {path.name}"
            )
        if sheet not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise SourceLoadError(
                f"worksheet {sheet!r} does not exist in {path.name}; "
                f"available sheets: {available}"
            )
        worksheet = workbook[sheet]
        minimum_column = 1
        maximum_column = worksheet.max_column
        maximum_row: int | None = None
        if cell_range is not None:
            try:
                from openpyxl.utils.cell import range_boundaries

                (
                    minimum_column,
                    selected_header_row,
                    maximum_column,
                    maximum_row,
                ) = range_boundaries(cell_range)
            except (TypeError, ValueError) as error:
                raise SourceLoadError(
                    f"named table has an invalid range: {path.name}#{sheet}"
                ) from error
            if selected_header_row != header_row:
                raise SourceLoadError(
                    f"named-table header changed since inspection: {path.name}#{sheet}"
                )
        if maximum_column - minimum_column + 1 > MAX_SOURCE_COLUMNS:
            raise SourceLoadError(
                f"worksheet {sheet!r} exceeds {MAX_SOURCE_COLUMNS} columns"
            )
        selected_maximum_row = maximum_row or worksheet.max_row
        if selected_maximum_row - header_row > MAX_SOURCE_ROWS:
            raise SourceLoadError(
                f"worksheet {sheet!r} exceeds {MAX_SOURCE_ROWS} possible data rows"
            )

        iterator = worksheet.iter_rows(
            min_row=header_row,
            max_row=maximum_row,
            min_col=minimum_column,
            max_col=maximum_column,
        )
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise SourceLoadError(
                f"worksheet {sheet!r} has no header row {header_row}"
            ) from exc
        _reject_unsafe_cells(header_cells, path.name, header_row)
        headers = _validate_headers(
            [cell.value for cell in header_cells],
            f"{path.name}#{sheet}",
        )

        rows: list[SourceRow] = []
        for cells in iterator:
            row_number = cells[0].row if cells else header_row + len(rows) + 1
            _reject_unsafe_cells(cells, path.name, row_number)
            values = [cell.value for cell in cells[: len(headers)]]
            if cell_range is None and len(cells) > len(headers) and any(
                cell.value is not None for cell in cells[len(headers) :]
            ):
                raise SourceLoadError(
                    f"row {row_number} has data beyond the declared headers: "
                    f"{path.name}#{sheet}"
                )
            if not any(value is not None for value in values):
                continue
            if len(rows) >= MAX_SOURCE_ROWS:
                raise SourceLoadError(
                    f"source exceeds {MAX_SOURCE_ROWS} data rows: "
                    f"{path.name}#{sheet}"
                )
            padded = [*values, *([None] * (len(headers) - len(values)))]
            _validate_cell_lengths(padded, path.name, row_number)
            rows.append(
                SourceRow(
                    number=row_number,
                    values=dict(zip(headers, padded, strict=True)),
                )
            )
        return headers, tuple(rows)
    finally:
        workbook.close()


def _validate_headers(raw_headers: Iterable[Any], label: str) -> tuple[str, ...]:
    """Return string headers after enforcing presence, uniqueness, and limits."""

    values = list(raw_headers)
    if not values:
        raise SourceLoadError(f"source has no columns: {label}")
    if len(values) > MAX_SOURCE_COLUMNS:
        raise SourceLoadError(
            f"source exceeds {MAX_SOURCE_COLUMNS} columns: {label}"
        )

    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        if value is None or str(value).strip() == "":
            raise SourceLoadError(f"column {index} has an empty header: {label}")
        header = str(value)
        if len(header) > MAX_CELL_STRING_LENGTH:
            raise SourceLoadError(f"column {index} header is too long: {label}")
        headers.append(header)

    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise SourceLoadError(f"duplicate headers {duplicates!r}: {label}")
    return tuple(headers)


def _validate_cell_lengths(values: Iterable[Any], label: str, row_number: int) -> None:
    """Reject string cells whose size exceeds the governed input limit."""

    for column_number, value in enumerate(values, start=1):
        if isinstance(value, str) and len(value) > MAX_CELL_STRING_LENGTH:
            raise SourceLoadError(
                f"cell at row {row_number}, column {column_number} exceeds "
                f"{MAX_CELL_STRING_LENGTH} characters: {label}"
            )


def _reject_unsafe_cells(cells: Iterable[Any], label: str, row_number: int) -> None:
    """Reject formula and Excel-error cells before extracting their values."""

    for column_number, cell in enumerate(cells, start=1):
        if cell.data_type == "f":
            raise SourceLoadError(
                f"formula cell rejected at row {row_number}, column "
                f"{column_number}: {label}"
            )
        if cell.data_type == "e":
            raise SourceLoadError(
                f"Excel error cell rejected at row {row_number}, column "
                f"{column_number}: {label}"
            )


def _validate_xlsx_container(path: Path) -> None:
    """Inspect the XLSX ZIP container for unsafe or resource-heavy content.

    This preflight rejects traversal entries, encryption, symlinks, zip-bomb
    patterns, macros, embedded objects, connections, and external links before
    ``openpyxl`` parses the workbook.
    """

    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise SourceLoadError(
                    f"XLSX archive exceeds {MAX_XLSX_ARCHIVE_ENTRIES} entries: "
                    f"{path.name}"
                )
            names = {member.filename for member in members}
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                raise SourceLoadError(f"file is not a valid XLSX container: {path.name}")

            expanded_bytes = 0
            for member in members:
                member_path = PurePosixPath(member.filename)
                if member_path.is_absolute() or ".." in member_path.parts:
                    raise SourceLoadError(
                        f"unsafe XLSX archive member {member.filename!r}: {path.name}"
                    )
                if member.flag_bits & 0x1:
                    raise SourceLoadError(
                        f"encrypted XLSX archive member rejected: {path.name}"
                    )
                if stat.S_ISLNK(member.external_attr >> 16):
                    raise SourceLoadError(
                        f"symlink XLSX archive member rejected: {path.name}"
                    )
                expanded_bytes += member.file_size
                if expanded_bytes > MAX_XLSX_EXPANDED_BYTES:
                    raise SourceLoadError(
                        f"XLSX expands beyond {MAX_XLSX_EXPANDED_BYTES} bytes: "
                        f"{path.name}"
                    )
                if (
                    member.file_size > 0
                    and (
                        member.compress_size == 0
                        or member.file_size / member.compress_size
                        > MAX_XLSX_MEMBER_COMPRESSION_RATIO
                    )
                ):
                    raise SourceLoadError(
                        f"suspicious XLSX compression ratio in "
                        f"{member.filename!r}: {path.name}"
                    )

            content_types_info = archive.getinfo("[Content_Types].xml")
            if content_types_info.file_size > MAX_XLSX_METADATA_BYTES:
                raise SourceLoadError(
                    f"XLSX content-type metadata is too large: {path.name}"
                )
            content_types = archive.read(content_types_info)
            if b"macroEnabled" in content_types or b"vbaProject" in content_types:
                raise SourceLoadError(
                    f"macro-enabled XLSX content rejected: {path.name}"
                )

            prohibited_prefixes = (
                "xl/externalLinks/",
                "xl/embeddings/",
            )
            prohibited_names = {
                "xl/vbaProject.bin",
                "xl/connections.xml",
            }
            unsafe = sorted(
                name
                for name in names
                if name in prohibited_names
                or any(name.startswith(prefix) for prefix in prohibited_prefixes)
            )
            if unsafe:
                raise SourceLoadError(
                    f"XLSX contains prohibited active or external content "
                    f"{unsafe!r}: {path.name}"
                )
    except zipfile.BadZipFile as exc:
        raise SourceLoadError(
            f"file is not a readable, unencrypted XLSX container: {path.name}"
        ) from exc


def validate_source_file(path: str | Path) -> None:
    """Validate an intake file before it is accepted into a project.

    Dataset-specific sheet, header, and type checks still happen during source
    inspection.  This public boundary performs format and container checks
    without requiring a mapping profile.
    """

    source_path = Path(path)
    extension = source_path.suffix.casefold()
    if extension == ".xlsx":
        _validate_xlsx_container(source_path)
        return
    if extension == ".csv":
        with source_path.open("rb") as stream:
            sample = stream.read(64 * 1024)
        if b"\x00" in sample:
            raise SourceLoadError(
                f"CSV contains binary null bytes: {source_path.name}"
            )
        return
    raise SourceLoadError(
        f"only .csv and .xlsx source files are accepted: {source_path.name}"
    )


def prepare_sources(
    profile: ProfileDocument,
    input_directory: str | Path,
) -> PreparedBundle:
    """Convert validated source tables into engine-ready records.

    Each row is parsed according to its :class:`DatasetSpec`.  Scalar fields
    become canonical Python values, while target identities and relations that
    require lookup remain symbolic ``LogicalReference`` objects.  Cross-row
    duplicate source identities are marked after all rows have been prepared.
    """

    tables = load_source_tables(profile, input_directory)
    root = Path(input_directory).resolve()
    return prepare_source_tables(
        profile,
        tables,
        source_hashes={
            table.path.relative_to(root).as_posix(): table.content_hash
            for table in sorted(tables, key=lambda item: item.dataset)
        },
    )


def prepare_source_tables(
    profile: ProfileDocument,
    tables: Iterable[SourceTable],
    *,
    source_hashes: dict[str, str],
) -> PreparedBundle:
    """Prepare already selected full-row tables for browser or CLI preflight."""

    selected_tables = tuple(tables)
    by_dataset = {table.dataset: table for table in selected_tables}
    if len(by_dataset) != len(selected_tables):
        raise SourceLoadError("Prepared source tables must be unique by dataset")
    if set(by_dataset) != {dataset.name for dataset in profile.datasets}:
        raise SourceLoadError("Prepared source tables do not match the profile")
    records: list[PreparedRecord] = []
    issues: list[Issue] = []

    for dataset in profile.datasets:
        table = by_dataset[dataset.name]
        missing_headers = sorted(_required_headers(dataset) - set(table.headers))
        if missing_headers:
            issues.append(
                Issue(
                    code="SOURCE_FIELD_MISSING",
                    message=f"missing source headers: {', '.join(missing_headers)}",
                    dataset=dataset.name,
                )
            )
        for row in table.rows:
            records.append(
                _prepare_row(dataset, row.values, row.number, missing_headers)
            )

    records = _mark_duplicate_source_identities(records)
    issues.extend(
        issue
        for record in records
        for issue in record.issues
        if issue.code == "SOURCE_IDENTITY_DUPLICATE"
    )

    return PreparedBundle(
        records=tuple(records),
        issues=tuple(issues),
        source_hashes=dict(sorted(source_hashes.items())),
    )


def _required_headers(dataset: DatasetSpec) -> set[str]:
    """Collect every source column used by a dataset's mapping contract."""

    headers = set(dataset.source_identity.fields)
    for component in (
        *dataset.target_identity.components,
        *dataset.target_identity.scope,
    ):
        headers.update(component.source_fields)
    headers.update(field.source for field in dataset.fields.values())
    for relation in dataset.relations.values():
        headers.update(relation.source_fields)
    return headers


def _prepare_row(
    dataset: DatasetSpec,
    row: dict[str, Any],
    row_index: int,
    missing_headers: Iterable[str],
) -> PreparedRecord:
    """Map one raw row to a ``PreparedRecord`` and attach row-local issues.

    This function coordinates scalar conversion, source/target identity
    construction, and relation preparation.  It records validation failures
    instead of aborting the complete dataset, allowing the final report to
    describe every affected row in one run.
    """

    row_issues: list[Issue] = []
    for header in missing_headers:
        row_issues.append(
            Issue(
                code="SOURCE_FIELD_MISSING",
                message=f"source field {header!r} is not present",
                dataset=dataset.name,
                row=row_index,
                field=header,
            )
        )

    source_identity: list[ScalarValue] = []
    identity_policy = NormalizationSpec(trim=True, empty_as_null=True)
    for field_name in dataset.source_identity.fields:
        try:
            value = parse_value(
                row.get(field_name),
                "string",
                identity_policy,
                required=True,
            )
        except ValueParseError as exc:
            value = None
            row_issues.append(
                Issue(
                    code="SOURCE_IDENTITY_INVALID",
                    message=str(exc),
                    dataset=dataset.name,
                    row=row_index,
                    field=field_name,
                )
            )
        source_identity.append(value)

    scalar_values: dict[str, ScalarValue] = {}
    for target_field, spec in dataset.fields.items():
        raw_value = row.get(spec.source)
        if isinstance(raw_value, InvalidPreparedValue):
            scalar_values[target_field] = None
            row_issues.append(
                Issue(
                    code=raw_value.code,
                    message=raw_value.message,
                    dataset=dataset.name,
                    row=row_index,
                    field=spec.source,
                )
            )
            continue
        try:
            scalar_values[target_field] = parse_field(raw_value, spec)
        except ValueParseError as exc:
            scalar_values[target_field] = None
            row_issues.append(
                Issue(
                    code=(
                        "SOURCE_REQUIRED_VALUE_MISSING"
                        if "required value" in str(exc)
                        else "SOURCE_TYPE_INVALID"
                    ),
                    message=str(exc),
                    dataset=dataset.name,
                    row=row_index,
                    field=spec.source,
                )
            )

    target_identity = tuple(
        value
        for component in dataset.target_identity.components
        for value in _prepare_identity_component(
            component, row, dataset.name, row_index, row_issues
        )
    )
    target_scope = tuple(
        value
        for component in dataset.target_identity.scope
        for value in _prepare_identity_component(
            component, row, dataset.name, row_index, row_issues
        )
    )

    references: dict[str, Any] = {}
    for target_field, relation in dataset.relations.items():
        references[target_field] = _prepare_relation(
            relation, row, dataset.name, row_index, target_field, row_issues
        )

    return PreparedRecord(
        dataset=dataset.name,
        source_row=row_index,
        target_model=dataset.target.model,
        source_identity=tuple(source_identity),
        target_identity=target_identity,
        target_scope=target_scope,
        scalar_values=scalar_values,
        references=references,
        issues=tuple(row_issues),
    )


def _prepare_identity_component(
    component: IdentityComponent,
    row: dict[str, Any],
    dataset_name: str,
    row_index: int,
    issues: list[Issue],
) -> tuple[Any, ...]:
    """Prepare one target-identity component as values or a logical reference.

    Direct components are parsed with their declared type and normalization.
    Resolved components are left symbolic so the engine can use either another
    incoming dataset or the Odoo target catalog.
    """

    if component.resolve is not None:
        key = _parse_reference_key(
            component.source_fields,
            row,
            dataset_name,
            row_index,
            component.target_fields[0],
            issues,
            required=True,
        )
        return (
            LogicalReference(
                origin=component.resolve.origin,
                key=key,
                dataset=component.resolve.dataset,
                model=component.resolve.target_model,
                target_fields=component.resolve.target_fields,
            ),
        )

    values: list[ScalarValue] = []
    for source_field in component.source_fields:
        try:
            value = parse_value(
                row.get(source_field),
                component.type,
                component.normalize,
                required=True,
            )
        except ValueParseError as exc:
            value = None
            issues.append(
                Issue(
                    code="SOURCE_IDENTITY_INVALID",
                    message=str(exc),
                    dataset=dataset_name,
                    row=row_index,
                    field=source_field,
                )
            )
        values.append(value)
    return tuple(values)


def _prepare_relation(
    relation: RelationSpec,
    row: dict[str, Any],
    dataset_name: str,
    row_index: int,
    target_field: str,
    issues: list[Issue],
) -> Any:
    """Build unresolved many2one or many2many references for one target field.

    Required values, empty list items, and duplicate many2many business keys
    become issues tied to the source row.  No numeric Odoo identifier crosses
    this source-preparation boundary.
    """

    if relation.kind == "many2one":
        key = _parse_reference_key(
            relation.source_fields,
            row,
            dataset_name,
            row_index,
            target_field,
            issues,
            required=relation.required,
        )
        if not key or all(value is None for value in key):
            return None
        return LogicalReference(
            origin=relation.resolve.origin,
            key=key,
            dataset=relation.resolve.dataset,
            model=relation.resolve.target_model,
            target_fields=relation.resolve.target_fields,
        )

    raw = row.get(relation.source_fields[0])
    if raw is None or str(raw).strip() == "":
        if relation.required:
            issues.append(
                Issue(
                    code="SOURCE_REQUIRED_VALUE_MISSING",
                    message="required many2many value is empty",
                    dataset=dataset_name,
                    row=row_index,
                    field=target_field,
                )
            )
        return ()
    keys = [item.strip() for item in str(raw).split(relation.separator)]
    if any(item == "" for item in keys):
        issues.append(
            Issue(
                code="SOURCE_TYPE_INVALID",
                message="many2many value contains an empty item",
                dataset=dataset_name,
                row=row_index,
                field=target_field,
            )
        )
        keys = [item for item in keys if item]
    if len(set(keys)) != len(keys):
        issues.append(
            Issue(
                code="SOURCE_REFERENCE_DUPLICATE",
                message="many2many value contains duplicate business keys",
                dataset=dataset_name,
                row=row_index,
                field=target_field,
            )
        )
    return tuple(
        LogicalReference(
            origin=relation.resolve.origin,
            key=(key,),
            dataset=relation.resolve.dataset,
            model=relation.resolve.target_model,
            target_fields=relation.resolve.target_fields,
        )
        for key in dict.fromkeys(keys)
    )


def _parse_reference_key(
    source_fields: Iterable[str],
    row: dict[str, Any],
    dataset_name: str,
    row_index: int,
    target_field: str,
    issues: list[Issue],
    *,
    required: bool,
) -> tuple[ScalarValue, ...]:
    """Parse the source columns forming one symbolic reference key."""

    values: list[ScalarValue] = []
    policy = NormalizationSpec(trim=True, empty_as_null=True)
    for source_field in source_fields:
        try:
            value = parse_value(
                row.get(source_field),
                "string",
                policy,
                required=required,
            )
        except ValueParseError as exc:
            value = None
            issues.append(
                Issue(
                    code="SOURCE_REQUIRED_VALUE_MISSING",
                    message=str(exc),
                    dataset=dataset_name,
                    row=row_index,
                    field=target_field,
                )
            )
        values.append(value)
    return tuple(values)


def _mark_duplicate_source_identities(
    records: list[PreparedRecord],
) -> list[PreparedRecord]:
    """Attach an issue to every row sharing an identity within one dataset.

    A single dictionary index performs the grouping in linear time, avoiding
    pairwise row comparisons for large source files.
    """

    indexes: dict[tuple[str, tuple[ScalarValue, ...]], list[int]] = {}
    for index, record in enumerate(records):
        indexes.setdefault((record.dataset, record.source_identity), []).append(index)

    result = list(records)
    for (dataset, identity), record_indexes in indexes.items():
        if len(record_indexes) < 2:
            continue
        for index in record_indexes:
            record = result[index]
            issue = Issue(
                code="SOURCE_IDENTITY_DUPLICATE",
                message=f"source identity {identity!r} occurs {len(record_indexes)} times",
                dataset=dataset,
                row=record.source_row,
                affected_count=len(record_indexes),
            )
            result[index] = replace(record, issues=(*record.issues, issue))
    return result
