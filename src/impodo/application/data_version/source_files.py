"""Governed CSV/XLSX readers and bounded source-file streams."""

from __future__ import annotations

import csv
from contextlib import contextmanager
from dataclasses import dataclass, replace
from hashlib import sha256
from pathlib import Path, PurePosixPath
import stat
from typing import Any, Iterable, Iterator
import zipfile

from impodo.domain.compiler.contracts import CompiledMigrationPlan
from impodo.domain.recipe.profile import DatasetSpec
from impodo.domain.preparation.source import (
    PreparedBundle,
    SourceLoadError,
    SourceRow,
    SourceTable,
    prepare_source_tables,
)


@dataclass(slots=True)
class SelectedSourceBatchStream:
    """One-shot bounded reader for a validated browser source selection.

    The owning :func:`open_selected_source_batches` context keeps the CSV file
    or read-only workbook open while batches are consumed and closes it even
    when evaluation stops early.
    """

    dataset: str
    path: Path
    headers: tuple[str, ...]
    content_hash: str
    batch_size: int
    _rows: Iterator[SourceRow]

    def __post_init__(self) -> None:
        if self.batch_size < 1:
            raise ValueError("Source batch size must be positive")

    def iter_batches(self) -> Iterator[tuple[SourceRow, ...]]:
        """Yield at most ``batch_size`` rows without retaining prior batches."""

        batch: list[SourceRow] = []
        for row in self._rows:
            batch.append(row)
            if len(batch) == self.batch_size:
                yield tuple(batch)
                batch.clear()
        if batch:
            yield tuple(batch)


MAX_SOURCE_FILE_BYTES = 50 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 10_000
MAX_XLSX_EXPANDED_BYTES = 512 * 1024 * 1024
MAX_XLSX_MEMBER_COMPRESSION_RATIO = 1_000
MAX_XLSX_METADATA_BYTES = 5 * 1024 * 1024
MAX_XLSX_WORKSHEETS = 256
MAX_SOURCE_ROWS = 500_000
MAX_SOURCE_COLUMNS = 2_048
MAX_CELL_STRING_LENGTH = 1_000_000


def validated_xlsx_table_bounds(cell_range: str) -> tuple[int, int, int, int]:
    """Return a bounded Excel-table range before any worksheet scan begins."""

    try:
        from openpyxl.utils.cell import range_boundaries

        minimum_column, header_row, maximum_column, maximum_row = range_boundaries(
            cell_range
        )
    except (TypeError, ValueError) as error:
        raise SourceLoadError("range is invalid") from error
    bounds = (minimum_column, header_row, maximum_column, maximum_row)
    if (
        not all(isinstance(value, int) for value in bounds)
        or minimum_column < 1
        or header_row < 1
        or maximum_column < minimum_column
        or maximum_row < header_row
    ):
        raise SourceLoadError("range is invalid")
    column_count = maximum_column - minimum_column + 1
    possible_data_rows = maximum_row - header_row
    if column_count > MAX_SOURCE_COLUMNS:
        raise SourceLoadError(
            f"range exceeds {MAX_SOURCE_COLUMNS} columns"
        )
    if possible_data_rows > MAX_SOURCE_ROWS:
        raise SourceLoadError(
            "range contains "
            f"{possible_data_rows:,} possible data rows; "
            f"the limit is {MAX_SOURCE_ROWS:,}"
        )
    return minimum_column, header_row, maximum_column, maximum_row


def load_source_tables(
    plan: CompiledMigrationPlan,
    input_directory: str | Path,
) -> tuple[SourceTable, ...]:
    """Load and validate every source file declared by ``plan``.

    The resolved files must remain inside ``input_directory`` and satisfy the
    size and format limits defined in this module.  CSV and XLSX parsing is
    deliberately strict so malformed, active, encrypted, or suspicious
    workbook content cannot silently enter the preflight pipeline.

    Returns:
        Tables in the same order as ``plan.datasets``.

    Raises:
        SourceLoadError: If the input directory, a file, or its content
            violates the source contract.
    """

    root = Path(input_directory).resolve()
    if not root.is_dir():
        raise SourceLoadError(f"source input directory does not exist: {root}")

    tables: list[SourceTable] = []
    for dataset in plan.datasets:
        path = _contained_source_path(root, dataset.source.file)
        file_size = path.stat().st_size
        if file_size > MAX_SOURCE_FILE_BYTES:
            raise SourceLoadError(
                f"source file exceeds {MAX_SOURCE_FILE_BYTES} bytes: "
                f"{dataset.source.file}"
            )
        content_hash = _source_content_hash(path)

        try:
            if path.suffix.casefold() == ".csv":
                headers, rows = _load_csv(
                    path,
                    dataset.source.encoding,
                    dataset.source.delimiter,
                )
            else:
                headers, rows = _load_xlsx(
                    path,
                    sheet=dataset.source.sheet or "",
                    header_row=dataset.source.header_row,
                )
        except SourceLoadError:
            raise
        except (csv.Error, LookupError, UnicodeError) as exc:
            raise SourceLoadError(
                f"cannot parse source file {dataset.source.file}: {exc}"
            ) from exc
        tables.append(
            SourceTable(
                dataset=dataset.name,
                path=path,
                headers=headers,
                rows=rows,
                content_hash=content_hash,
            )
        )
    return tuple(tables)


def load_selected_source_table(
    path: str | Path,
    *,
    dataset: str,
    table_key: str,
    encoding: str | None,
    delimiter: str | None,
    header_row: int,
    named_table_range: str | None = None,
    source_display_name: str | None = None,
) -> SourceTable:
    """Load one frozen browser dataset through the strict source reader.

    Browser selections identify worksheets and named tables independently of
    the profile-driven CLI. This adapter preserves that exact selection while
    reusing the same passive XLSX, row, column, and cell safety limits.
    """

    with open_selected_source_batches(
        path,
        dataset=dataset,
        table_key=table_key,
        encoding=encoding,
        delimiter=delimiter,
        header_row=header_row,
        named_table_range=named_table_range,
        source_display_name=source_display_name,
    ) as source:
        return SourceTable(
            dataset=source.dataset,
            path=source.path,
            headers=source.headers,
            rows=tuple(
                row
                for batch in source.iter_batches()
                for row in batch
            ),
            content_hash=source.content_hash,
        )


@contextmanager
def open_selected_source_batches(
    path: str | Path,
    *,
    dataset: str,
    table_key: str,
    encoding: str | None,
    delimiter: str | None,
    header_row: int,
    named_table_range: str | None = None,
    source_display_name: str | None = None,
    batch_size: int = 1_000,
) -> Iterator[SelectedSourceBatchStream]:
    """Open one frozen browser dataset and yield validated bounded batches.

    This is the streaming counterpart to :func:`load_selected_source_table`.
    It preserves the exact strict reader and source-hash behavior while keeping
    only the current input batch in Python memory.
    """

    source_path = Path(path).resolve()
    if not source_path.is_file() or source_path.is_symlink():
        raise SourceLoadError("stored source artifact is unavailable")
    if source_path.stat().st_size > MAX_SOURCE_FILE_BYTES:
        raise SourceLoadError(
            f"source file exceeds {MAX_SOURCE_FILE_BYTES} bytes: {source_path.name}"
        )
    if batch_size < 1:
        raise ValueError("Source batch size must be positive")

    content_hash = _source_content_hash(source_path)
    suffix = source_path.suffix.casefold()
    if suffix == ".csv":
        if table_key != "csv":
            raise SourceLoadError("CSV dataset selection is invalid")
        with _open_csv_rows(
            source_path,
            encoding or "utf-8-sig",
            delimiter or ",",
        ) as (headers, rows):
            yield SelectedSourceBatchStream(
                dataset=dataset,
                path=source_path,
                headers=headers,
                content_hash=content_hash,
                batch_size=batch_size,
                _rows=rows,
            )
        return

    if suffix == ".xlsx":
        if table_key.startswith("sheet:"):
            sheet = table_key.removeprefix("sheet:")
            cell_range = None
        elif table_key.startswith("table:") and named_table_range:
            selected = table_key.removeprefix("table:")
            try:
                sheet, _table_name = selected.rsplit(":", 1)
            except ValueError as error:
                raise SourceLoadError("Named-table selection is invalid") from error
            cell_range = named_table_range
        else:
            raise SourceLoadError("XLSX dataset selection is invalid")
        with _open_xlsx_rows(
            source_path,
            sheet=sheet,
            header_row=header_row,
            cell_range=cell_range,
            source_display_name=source_display_name,
        ) as (headers, rows):
            yield SelectedSourceBatchStream(
                dataset=dataset,
                path=source_path,
                headers=headers,
                content_hash=content_hash,
                batch_size=batch_size,
                _rows=rows,
            )
        return

    raise SourceLoadError("Only CSV and XLSX source files are supported")


def _source_content_hash(path: Path) -> str:
    """Hash a governed source without retaining another full-file copy."""

    digest = sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return "sha256:" + digest.hexdigest()


def _contained_source_path(root: Path, relative_name: str) -> Path:
    """Resolve one declared source path without permitting link/path escape."""

    candidate = root / relative_name
    if candidate.is_symlink():
        raise SourceLoadError(f"source file must not be a symlink: {relative_name}")
    path = candidate.resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise SourceLoadError(
            f"source file escapes the input directory: {relative_name}"
        ) from exc
    if not path.is_file():
        raise SourceLoadError(f"source file does not exist: {relative_name}")
    return path


def _load_csv(
    path: Path,
    encoding: str,
    delimiter: str,
) -> tuple[tuple[str, ...], tuple[SourceRow, ...]]:
    """Read one CSV file and enforce row, column, and cell-size limits."""

    with _open_csv_rows(path, encoding, delimiter) as (headers, rows):
        return headers, tuple(rows)


@contextmanager
def _open_csv_rows(
    path: Path,
    encoding: str,
    delimiter: str,
) -> Iterator[tuple[tuple[str, ...], Iterator[SourceRow]]]:
    """Open one CSV file and expose its validated rows as a one-shot stream."""

    with path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            raise SourceLoadError(f"CSV source has no header row: {path.name}") from exc
        headers = _validate_headers(raw_headers, path.name)

        def iter_rows() -> Iterator[SourceRow]:
            data_rows = 0
            for values in reader:
                row_number = reader.line_num
                if not values:
                    continue
                if data_rows >= MAX_SOURCE_ROWS:
                    raise SourceLoadError(
                        f"source exceeds {MAX_SOURCE_ROWS} data rows: {path.name}"
                    )
                if len(values) > len(headers):
                    raise SourceLoadError(
                        f"row {row_number} has {len(values)} cells but the header has "
                        f"{len(headers)}: {path.name}"
                    )
                padded = [*values, *([None] * (len(headers) - len(values)))]
                _validate_cell_lengths(padded, path.name, row_number)
                data_rows += 1
                yield SourceRow(
                    number=row_number,
                    values=dict(zip(headers, padded, strict=True)),
                )

        yield headers, iter_rows()


def _load_xlsx(
    path: Path,
    *,
    sheet: str,
    header_row: int,
    cell_range: str | None = None,
) -> tuple[tuple[str, ...], tuple[SourceRow, ...]]:
    """Read one worksheet from a passive, bounded XLSX container."""

    with _open_xlsx_rows(
        path,
        sheet=sheet,
        header_row=header_row,
        cell_range=cell_range,
    ) as (headers, rows):
        return headers, tuple(rows)


@contextmanager
def _open_xlsx_rows(
    path: Path,
    *,
    sheet: str,
    header_row: int,
    cell_range: str | None = None,
    source_display_name: str | None = None,
) -> Iterator[tuple[tuple[str, ...], Iterator[SourceRow]]]:
    """Open one worksheet and expose validated selected rows lazily.

    Workbook loading is read-only and formulas are rejected instead of being
    calculated or trusted.  The original worksheet row numbers are retained.
    """

    _validate_xlsx_container(path)

    try:
        from openpyxl import load_workbook
        from openpyxl.xml import DEFUSEDXML
    except ImportError as exc:
        raise SourceLoadError(
            "XLSX support requires openpyxl and defusedxml"
        ) from exc
    if not DEFUSEDXML:
        raise SourceLoadError("XLSX parsing requires active defusedxml protection")

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise SourceLoadError(f"cannot parse XLSX source: {path.name}") from exc
    try:
        if len(workbook.sheetnames) > MAX_XLSX_WORKSHEETS:
            raise SourceLoadError(
                f"workbook exceeds {MAX_XLSX_WORKSHEETS} worksheets: {path.name}"
            )
        if sheet not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise SourceLoadError(
                f"worksheet {sheet!r} does not exist in {path.name}; "
                f"available sheets: {available}"
            )
        worksheet = workbook[sheet]
        minimum_column = 1
        maximum_column = worksheet.max_column
        maximum_row: int | None = None
        if cell_range is not None:
            try:
                (
                    minimum_column,
                    selected_header_row,
                    maximum_column,
                    maximum_row,
                ) = validated_xlsx_table_bounds(cell_range)
            except SourceLoadError as error:
                raise SourceLoadError(
                    f"named table {error}: {path.name}#{sheet}"
                ) from error
            if selected_header_row != header_row:
                raise SourceLoadError(
                    f"named-table header changed since inspection: {path.name}#{sheet}"
                )
        # Some XLSX exporters omit the optional worksheet dimension. In that
        # case openpyxl exposes None while its read-only iterator can still
        # stream the cells; header and row limits below remain authoritative.
        if (
            maximum_column is not None
            and maximum_column - minimum_column + 1 > MAX_SOURCE_COLUMNS
        ):
            raise SourceLoadError(
                f"worksheet {sheet!r} exceeds {MAX_SOURCE_COLUMNS} columns"
            )
        selected_maximum_row = maximum_row or worksheet.max_row
        if (
            selected_maximum_row is not None
            and selected_maximum_row - header_row > MAX_SOURCE_ROWS
        ):
            raise SourceLoadError(
                f"worksheet {sheet!r} exceeds {MAX_SOURCE_ROWS} possible data rows"
            )

        iterator = worksheet.iter_rows(
            min_row=header_row,
            max_row=maximum_row,
            min_col=minimum_column,
            max_col=maximum_column,
        )
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise SourceLoadError(
                f"worksheet {sheet!r} has no header row {header_row}"
            ) from exc
        source_label = source_display_name or path.name
        _reject_unsafe_cells(
            header_cells,
            source_label,
            header_row,
            sheet=sheet,
        )
        headers = _validate_headers(
            [cell.value for cell in header_cells],
            f"{path.name}#{sheet}",
        )

        def iter_rows() -> Iterator[SourceRow]:
            data_rows = 0
            for cells in iterator:
                row_number = cells[0].row if cells else header_row + data_rows + 1
                _reject_unsafe_cells(
                    cells,
                    source_label,
                    row_number,
                    sheet=sheet,
                    headers=headers,
                )
                values = [cell.value for cell in cells[: len(headers)]]
                if cell_range is None and len(cells) > len(headers) and any(
                    cell.value is not None for cell in cells[len(headers) :]
                ):
                    raise SourceLoadError(
                        f"row {row_number} has data beyond the declared headers: "
                        f"{path.name}#{sheet}"
                    )
                if not any(value is not None for value in values):
                    continue
                if data_rows >= MAX_SOURCE_ROWS:
                    raise SourceLoadError(
                        f"source exceeds {MAX_SOURCE_ROWS} data rows: "
                        f"{path.name}#{sheet}"
                    )
                padded = [*values, *([None] * (len(headers) - len(values)))]
                _validate_cell_lengths(padded, path.name, row_number)
                data_rows += 1
                yield SourceRow(
                    number=row_number,
                    values=dict(zip(headers, padded, strict=True)),
                )

        yield headers, iter_rows()
    finally:
        workbook.close()


def _validate_headers(raw_headers: Iterable[Any], label: str) -> tuple[str, ...]:
    """Return string headers after enforcing presence, uniqueness, and limits."""

    values = list(raw_headers)
    if not values:
        raise SourceLoadError(f"source has no columns: {label}")
    if len(values) > MAX_SOURCE_COLUMNS:
        raise SourceLoadError(
            f"source exceeds {MAX_SOURCE_COLUMNS} columns: {label}"
        )

    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        if value is None or str(value).strip() == "":
            raise SourceLoadError(f"column {index} has an empty header: {label}")
        header = str(value)
        if len(header) > MAX_CELL_STRING_LENGTH:
            raise SourceLoadError(f"column {index} header is too long: {label}")
        headers.append(header)

    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise SourceLoadError(f"duplicate headers {duplicates!r}: {label}")
    return tuple(headers)


def _validate_cell_lengths(values: Iterable[Any], label: str, row_number: int) -> None:
    """Reject string cells whose size exceeds the governed input limit."""

    for column_number, value in enumerate(values, start=1):
        if isinstance(value, str) and len(value) > MAX_CELL_STRING_LENGTH:
            raise SourceLoadError(
                f"cell at row {row_number}, column {column_number} exceeds "
                f"{MAX_CELL_STRING_LENGTH} characters: {label}"
            )


def _reject_unsafe_cells(
    cells: Iterable[Any],
    label: str,
    row_number: int,
    *,
    sheet: str | None = None,
    headers: tuple[str, ...] | None = None,
) -> None:
    """Reject formula and Excel-error cells before extracting their values."""

    for column_number, cell in enumerate(cells, start=1):
        column_name = (
            headers[column_number - 1]
            if headers is not None and column_number <= len(headers)
            else None
        )
        column_detail = f' in "{column_name}"' if column_name else ""
        coordinate = getattr(cell, "coordinate", None)
        location = (
            f" at {sheet}!{coordinate}"
            if sheet and coordinate
            else f" at row {row_number}, column {column_number}"
        )
        if cell.data_type == "f":
            raise SourceLoadError(
                f"Excel formula found{column_detail}{location} in {label}. "
                "Remove the formula or replace it with a fixed value before "
                "using this source."
            )
        if cell.data_type == "e":
            raise SourceLoadError(
                f"Excel error found{column_detail}{location} in {label}. "
                "Correct the cell before using this source."
            )


def _validate_xlsx_container(path: Path) -> None:
    """Inspect the XLSX ZIP container for unsafe or resource-heavy content.

    This preflight rejects traversal entries, encryption, symlinks, zip-bomb
    patterns, macros, embedded objects, connections, and external links before
    ``openpyxl`` parses the workbook.
    """

    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise SourceLoadError(
                    f"XLSX archive exceeds {MAX_XLSX_ARCHIVE_ENTRIES} entries: "
                    f"{path.name}"
                )
            names = {member.filename for member in members}
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            if not required.issubset(names):
                raise SourceLoadError(f"file is not a valid XLSX container: {path.name}")

            expanded_bytes = 0
            for member in members:
                member_path = PurePosixPath(member.filename)
                if member_path.is_absolute() or ".." in member_path.parts:
                    raise SourceLoadError(
                        f"unsafe XLSX archive member {member.filename!r}: {path.name}"
                    )
                if member.flag_bits & 0x1:
                    raise SourceLoadError(
                        f"encrypted XLSX archive member rejected: {path.name}"
                    )
                if stat.S_ISLNK(member.external_attr >> 16):
                    raise SourceLoadError(
                        f"symlink XLSX archive member rejected: {path.name}"
                    )
                expanded_bytes += member.file_size
                if expanded_bytes > MAX_XLSX_EXPANDED_BYTES:
                    raise SourceLoadError(
                        f"XLSX expands beyond {MAX_XLSX_EXPANDED_BYTES} bytes: "
                        f"{path.name}"
                    )
                if (
                    member.file_size > 0
                    and (
                        member.compress_size == 0
                        or member.file_size / member.compress_size
                        > MAX_XLSX_MEMBER_COMPRESSION_RATIO
                    )
                ):
                    raise SourceLoadError(
                        f"suspicious XLSX compression ratio in "
                        f"{member.filename!r}: {path.name}"
                    )

            content_types_info = archive.getinfo("[Content_Types].xml")
            if content_types_info.file_size > MAX_XLSX_METADATA_BYTES:
                raise SourceLoadError(
                    f"XLSX content-type metadata is too large: {path.name}"
                )
            content_types = archive.read(content_types_info)
            if b"macroEnabled" in content_types or b"vbaProject" in content_types:
                raise SourceLoadError(
                    f"macro-enabled XLSX content rejected: {path.name}"
                )

            prohibited_prefixes = (
                "xl/externalLinks/",
                "xl/embeddings/",
            )
            prohibited_names = {
                "xl/vbaProject.bin",
                "xl/connections.xml",
            }
            unsafe = sorted(
                name
                for name in names
                if name in prohibited_names
                or any(name.startswith(prefix) for prefix in prohibited_prefixes)
            )
            if unsafe:
                raise SourceLoadError(
                    f"XLSX contains prohibited active or external content "
                    f"{unsafe!r}: {path.name}"
                )
    except zipfile.BadZipFile as exc:
        raise SourceLoadError(
            f"file is not a readable, unencrypted XLSX container: {path.name}"
        ) from exc


def validate_source_file(path: str | Path) -> None:
    """Validate an intake file before it is accepted into a project.

    Dataset-specific sheet, header, and type checks still happen during source
    inspection.  This public boundary performs format and container checks
    without requiring a mapping profile.
    """

    source_path = Path(path)
    extension = source_path.suffix.casefold()
    if extension == ".xlsx":
        _validate_xlsx_container(source_path)
        return
    if extension == ".csv":
        with source_path.open("rb") as stream:
            sample = stream.read(64 * 1024)
        if b"\x00" in sample:
            raise SourceLoadError(
                f"CSV contains binary null bytes: {source_path.name}"
            )
        return
    raise SourceLoadError(
        f"only .csv and .xlsx source files are accepted: {source_path.name}"
    )


def prepare_sources(
    plan: CompiledMigrationPlan,
    input_directory: str | Path,
) -> PreparedBundle:
    """Convert validated source tables into engine-ready records.

    Each row is parsed according to its :class:`DatasetSpec`.  Scalar fields
    become canonical Python values, while target identities and relations that
    require lookup remain symbolic ``LogicalReference`` objects.  Cross-row
    duplicate source identities are marked after all rows have been prepared.
    """

    tables = load_source_tables(plan, input_directory)
    root = Path(input_directory).resolve()
    return prepare_source_tables(
        plan,
        tables,
        source_hashes={
            table.path.relative_to(root).as_posix(): table.content_hash
            for table in sorted(tables, key=lambda item: item.dataset)
        },
    )


