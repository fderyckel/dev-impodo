"""Write the portable Stage 3 matching review workbook.

The workbook projects one immutable mapping revision and its exact validation
result.  It uses captured source and Odoo schema labels, but it neither opens
source artifacts nor contacts Odoo.  Stage 4 prepared rows and the Stage 5
target comparison deliberately remain outside this artifact.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
import json
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from impodo.domain.mapping.artifacts import MappingRevision
from impodo.domain.mapping.contracts import (
    DatasetMapping,
    MappingTargetMode,
    RelationshipMapping,
    RelationshipValueSource,
    ScalarFieldMapping,
    ScalarValueSource,
    TargetFieldHandling,
)
from impodo.domain.mapping.validation.evidence import (
    MappingValidationIssue,
    MappingValidationResult,
    MappingValidationStatus,
)
from impodo.domain.recipe.value_rules import ScalarTransformPolicy
from impodo.domain.source_binding import SourceOriginKind
from impodo.domain.workspace.contracts import (
    OdooSchemaCatalog,
    SchemaField,
    SourceDataset,
    SourceSelection,
)


MAPPING_REVIEW_WORKBOOK_PREFIX = "impodo_mapping_review"

# These values intentionally match the Stage 5 workbook palette.  Keep the
# Stage 5 writer independent so adding this projection cannot alter it.
_COLORS = {
    "brand": "E8473F",
    "charcoal": "494D46",
    "charcoal_dark": "292C28",
    "gray": "868981",
    "paper": "F4F4F1",
    "surface": "FFFFFF",
    "soft": "EEEEEA",
    "line": "D6D7D2",
    "ready": "4D7C5B",
    "ready_text": "315B3B",
    "ready_bg": "EDF7EF",
    "warning": "7D4F00",
    "warning_bg": "FFF5DF",
    "danger": "9F2F2F",
    "danger_bg": "FCE8E7",
    "prepared": "2F628F",
    "prepared_bg": "EAF2FB",
    "white": "FFFFFF",
}
_STYLE_COLORS = {
    "ready": (_COLORS["ready_bg"], _COLORS["ready_text"]),
    "warning": (_COLORS["warning_bg"], _COLORS["warning"]),
    "danger": (_COLORS["danger_bg"], _COLORS["danger"]),
    "prepared": (_COLORS["prepared_bg"], _COLORS["prepared"]),
    "neutral": (_COLORS["soft"], _COLORS["charcoal"]),
}
_THIN_BORDER = Border(bottom=Side(style="thin", color=_COLORS["line"]))
_EXCEL_MAX_ROW = 1_048_576
_EXCEL_MAX_COLUMN = 16_384
_INVALID_SHEET_CHARACTERS = re.compile(r"[\\/*?:\[\]]")


class MappingReviewGenerationError(RuntimeError):
    """Raised when current Stage 3 evidence cannot produce a safe workbook."""


@dataclass(frozen=True, slots=True)
class _FieldReview:
    dataset: DatasetMapping
    source_dataset: SourceDataset
    field: SchemaField
    status: str
    style: str
    provider: str
    source_fields: str
    requirement: str
    next_action: str


def mapping_review_workbook_name(revision: MappingRevision) -> str:
    """Return the artifact name bound to one exact mapping revision."""

    digest = revision.definition.content_hash.removeprefix("sha256:")
    if not re.fullmatch(r"[0-9a-fA-F]{64}", digest):
        raise MappingReviewGenerationError("Mapping identity is invalid")
    return (
        f"{MAPPING_REVIEW_WORKBOOK_PREFIX}_v{revision.version}_"
        f"{digest[:16].casefold()}.xlsx"
    )


def write_mapping_review_workbook(
    revision: MappingRevision,
    validation: MappingValidationResult,
    selection: SourceSelection,
    schema: OdooSchemaCatalog,
    workbook_path: str | Path,
) -> Path:
    """Write one read-only matching review from exact checked evidence."""

    _require_bound_evidence(revision, validation, selection)
    target = Path(workbook_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Matching overview"
    attention = workbook.create_sheet("Needs attention")
    field_matches = workbook.create_sheet("Field matches")
    value_coverage = workbook.create_sheet("Value coverage")
    checked_later = workbook.create_sheet("Checked later")

    source_by_id = {item.dataset_id: item for item in selection.datasets}
    models_by_name = {item.name: item for item in schema.models}
    issues_by_field: dict[tuple[str, str], list[MappingValidationIssue]] = (
        defaultdict(list)
    )
    for issue in validation.issues:
        if issue.dataset_id and issue.target_field:
            issues_by_field[(issue.dataset_id, issue.target_field)].append(issue)

    field_reviews: list[_FieldReview] = []
    dataset_reviews: list[tuple[DatasetMapping, SourceDataset, tuple[_FieldReview, ...]]] = []
    for dataset in revision.definition.datasets:
        source_dataset = source_by_id.get(dataset.dataset_id)
        model = models_by_name.get(dataset.target_model)
        if source_dataset is None or model is None:
            continue
        reviews = _dataset_field_reviews(
            dataset,
            source_dataset,
            model.fields,
            issues_by_field,
        )
        field_reviews.extend(reviews)
        dataset_reviews.append((dataset, source_dataset, reviews))

    _write_overview(overview, revision, validation, selection, schema, field_reviews)
    _write_attention(attention, validation, source_by_id, models_by_name)
    _write_field_matches(field_matches, field_reviews)
    _write_value_coverage(value_coverage, validation, source_by_id, models_by_name)
    _write_checked_later(checked_later, validation, source_by_id)

    used_names = {sheet.title for sheet in workbook.worksheets}
    for index, (_dataset, source_dataset, reviews) in enumerate(dataset_reviews, start=1):
        sheet_name = _unique_sheet_name(
            f"{index} {source_dataset.name} fields",
            used_names,
        )
        used_names.add(sheet_name)
        _write_dataset_columns(workbook.create_sheet(sheet_name), source_dataset, reviews)

    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    temporary = target.with_name(f".{target.name}.partial")
    try:
        workbook.save(temporary)
        temporary.replace(target)
    except (OSError, ValueError) as error:
        raise MappingReviewGenerationError(
            "Matching review workbook generation failed"
        ) from error
    finally:
        temporary.unlink(missing_ok=True)
        workbook.close()
    return target


def _require_bound_evidence(
    revision: MappingRevision,
    validation: MappingValidationResult,
    selection: SourceSelection,
) -> None:
    definition = revision.definition
    if validation.mapping_content_hash != definition.content_hash:
        raise MappingReviewGenerationError(
            "Check matches again before creating the workbook"
        )
    if (
        validation.source_selection_hash != selection.content_hash
        or definition.source_selection_hash != selection.content_hash
    ):
        raise MappingReviewGenerationError(
            "The checked source data is no longer current"
        )
    if validation.schema_hash != definition.schema_hash:
        raise MappingReviewGenerationError(
            "The checked Odoo fields are no longer current"
        )


def _dataset_field_reviews(
    dataset: DatasetMapping,
    source_dataset: SourceDataset,
    schema_fields: tuple[SchemaField, ...],
    issues_by_field: dict[tuple[str, str], list[MappingValidationIssue]],
) -> tuple[_FieldReview, ...]:
    fields_by_name = {item.name: item for item in schema_fields}
    scalar_by_field = {item.target_field: item for item in dataset.fields}
    relation_by_field = {
        item.target_field: item for item in dataset.relationships
    }
    disposition_by_field = {
        item.target_field: item for item in dataset.target_field_dispositions
    }
    identity_fields = {
        target_field
        for component in (*dataset.target_identity, *dataset.target_scope)
        for target_field in component.target_fields
    }
    relevant = set(scalar_by_field) | set(relation_by_field) | set(
        disposition_by_field
    ) | identity_fields
    relevant.update(
        item.name
        for item in schema_fields
        if item.required
        and not item.readonly
        and dataset.mode is not MappingTargetMode.REFERENCE
    )
    relevant.update(
        target_field
        for issue_dataset, target_field in issues_by_field
        if issue_dataset == dataset.dataset_id
    )
    source_labels = {
        item.stable_key: item.source_name for item in source_dataset.columns
    }

    reviews: list[_FieldReview] = []
    for field_name in sorted(
        relevant,
        key=lambda name: (
            fields_by_name.get(name).label.casefold()
            if fields_by_name.get(name) is not None
            else name.casefold()
        ),
    ):
        field = fields_by_name.get(field_name)
        if field is None:
            continue
        issues = issues_by_field.get((dataset.dataset_id, field_name), ())
        scalar = scalar_by_field.get(field_name)
        relation = relation_by_field.get(field_name)
        disposition = disposition_by_field.get(field_name)
        status, style, next_action = _field_status(
            field,
            issues,
            scalar,
            relation,
            disposition.handling if disposition is not None else None,
            field_name in identity_fields,
        )
        provider, source_fields = _field_provider(
            scalar,
            relation,
            disposition.handling if disposition is not None else None,
            field_name in identity_fields,
            source_labels,
        )
        reviews.append(
            _FieldReview(
                dataset=dataset,
                source_dataset=source_dataset,
                field=field,
                status=status,
                style=style,
                provider=provider,
                source_fields=source_fields,
                requirement=(
                    "Required by Odoo"
                    if field.required and not field.readonly
                    else "Optional or managed by Odoo"
                ),
                next_action=next_action,
            )
        )
    return tuple(reviews)


def _field_status(
    field: SchemaField,
    issues: Iterable[MappingValidationIssue],
    scalar: ScalarFieldMapping | None,
    relation: RelationshipMapping | None,
    handling: TargetFieldHandling | None,
    identity: bool,
) -> tuple[str, str, str]:
    issues = tuple(issues)
    errors = tuple(item for item in issues if item.severity == "error")
    warnings = tuple(item for item in issues if item.severity == "warning")
    if errors:
        return "Must fix", "danger", errors[0].remediation
    if warnings:
        return "Review required", "warning", warnings[0].remediation
    if handling is TargetFieldHandling.ODOO_DEFAULT:
        return (
            "Odoo will choose",
            "warning",
            "Review the verified Odoo default before confirming the matches.",
        )
    if handling is TargetFieldHandling.ODOO_MANAGED:
        return (
            "Odoo manages this field",
            "warning",
            "Review that Odoo creates or maintains this field.",
        )
    if relation is not None:
        return (
            "Relationship configured",
            "ready",
            "Preparation will check every related key for one exact record.",
        )
    if scalar is not None:
        if scalar.value_source is ScalarValueSource.ODOO_DEFAULT:
            return (
                "Odoo will choose",
                "warning",
                "Review the verified Odoo default before confirming the matches.",
            )
        if (
            scalar.value_source is not ScalarValueSource.SOURCE
            or scalar.transform != ScalarTransformPolicy()
        ):
            return (
                "Impodo supplies or prepares",
                "prepared",
                "Review the configured value or rule effect.",
            )
        return "Mapped", "ready", "No mapping correction is currently required."
    if identity:
        return "Identity configured", "ready", "No mapping correction is currently required."
    if field.required and not field.readonly:
        return (
            "Must fix",
            "danger",
            "Map incoming data, provide a fixed value, or confirm an available Odoo default.",
        )
    return "Not used", "neutral", "No action is currently required."


def _field_provider(
    scalar: ScalarFieldMapping | None,
    relation: RelationshipMapping | None,
    handling: TargetFieldHandling | None,
    identity: bool,
    source_labels: dict[str, str],
) -> tuple[str, str]:
    if handling is TargetFieldHandling.ODOO_DEFAULT:
        return "Verified Odoo default", ""
    if handling is TargetFieldHandling.ODOO_MANAGED:
        return "Odoo-managed field", ""
    if relation is not None:
        if relation.value_source is RelationshipValueSource.CONSTANT_EXISTING:
            reference = relation.constant_reference
            display = " · ".join(
                f"{item.target_field}={item.value}"
                for item in (
                    *((reference.key_values if reference is not None else ())),
                    *((reference.scope_values if reference is not None else ())),
                )
            )
            return "Same existing Odoo record for every row", display
        source = ", ".join(
            source_labels.get(item, item) for item in relation.source_column_keys
        )
        return f"{relation.kind} relationship", source
    if scalar is not None:
        source = (
            " + ".join(
                source_labels.get(key, key)
                for key in scalar.concatenation.source_column_keys
            )
            if scalar.concatenation is not None
            else (
                source_labels.get(
                    scalar.source_column_key,
                    scalar.source_column_key,
                )
                if scalar.source_column_key
                else ""
            )
        )
        provider = {
            ScalarValueSource.SOURCE: "Incoming value",
            ScalarValueSource.CONSTANT: "Fixed value",
            ScalarValueSource.SOURCE_WITH_FALLBACK: "Incoming value with backup",
            ScalarValueSource.CONCATENATE: "Combined source columns",
            ScalarValueSource.CONDITIONAL_RULES: "Ordered choice rules",
            ScalarValueSource.ODOO_DEFAULT: "Odoo default",
        }[scalar.value_source]
        return provider, source
    if identity:
        return "Business identity", "See the confirmed identity mapping"
    return "No value provider", ""


def _write_overview(
    sheet,
    revision: MappingRevision,
    validation: MappingValidationResult,
    selection: SourceSelection,
    schema: OdooSchemaCatalog,
    field_reviews: list[_FieldReview],
) -> None:
    _title_band(
        sheet,
        "Impodo matching review",
        "Stage 3 checked matches — not final load readiness",
        8,
    )
    error_count = sum(item.severity == "error" for item in validation.issues)
    warning_count = sum(item.severity == "warning" for item in validation.issues)
    if validation.status is MappingValidationStatus.INVALID:
        status, style, action = (
            "Cannot confirm matches",
            "danger",
            "Open Needs attention and fix every red item in Impodo.",
        )
    elif warning_count:
        status, style, action = (
            "Review required",
            "warning",
            "Review every amber item before confirming the matches.",
        )
    else:
        status, style, action = (
            "Ready to confirm",
            "ready",
            "Review the field matches, then confirm them in Impodo.",
        )
    rows = [
        ("Check result", status),
        ("Must fix", error_count),
        ("Review items", warning_count),
        ("Tables checked", len(revision.definition.datasets)),
        ("Fields shown", len(field_reviews)),
        ("Next action", action),
        ("Odoo target", f"{schema.database} — Odoo {schema.odoo_version}"),
        ("Checked mapping", f"Version {revision.version}"),
    ]
    _write_key_values(sheet, rows, start_row=4)
    _style_status_cell(sheet["B4"], style)

    sheet["D4"] = "How to use this workbook"
    sheet["D4"].fill = PatternFill("solid", fgColor=_COLORS["brand"])
    sheet["D4"].font = Font(bold=True, color=_COLORS["white"])
    sheet.merge_cells("D4:H4")
    sheet["D5"] = (
        "Start with Needs attention. Each table also has a field-column view: "
        "red means Must fix, amber means Review required, green means mapped, "
        "and blue means Impodo supplies or prepares the value. Correct data or "
        "rules in Impodo, then run Check matches and recreate this workbook."
    )
    sheet["D5"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet["D5"].fill = PatternFill("solid", fgColor=_COLORS["paper"])
    sheet.merge_cells("D5:H10")
    sheet.row_dimensions[5].height = 88

    legend = (
        ("Must fix", "The checked mapping cannot be confirmed.", "danger"),
        ("Review required", "A deliberate decision needs review.", "warning"),
        ("Mapped", "The field mapping is currently valid.", "ready"),
        ("Impodo prepares", "Impodo supplies or transforms the value.", "prepared"),
        ("Not used", "No action is currently required.", "neutral"),
    )
    sheet["A14"] = "Colour and status meanings"
    sheet.merge_cells("A14:H14")
    _style_header(sheet, 14, 1, 8, _COLORS["charcoal_dark"])
    for row_index, (label, meaning, legend_style) in enumerate(legend, start=15):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, meaning)
        sheet.merge_cells(
            start_row=row_index,
            start_column=2,
            end_row=row_index,
            end_column=8,
        )
        _style_status_cell(sheet.cell(row_index, 1), legend_style)
    sheet["A22"] = "What this workbook does not contain"
    sheet.merge_cells("A22:H22")
    _style_header(sheet, 22, 1, 8, _COLORS["charcoal_dark"])
    sheet["A23"] = (
        "Stage 4 still prepares every row and resolves duplicates and related "
        "records. Stage 5 still compares the final prepared rows with fresh Odoo "
        "evidence. Use the separate Stage 5 review workbook for the proposed load."
    )
    sheet.merge_cells("A23:H25")
    sheet["A23"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet["A23"].fill = PatternFill("solid", fgColor=_COLORS["soft"])
    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 24
    for column in "CDEFGH":
        sheet.column_dimensions[column].width = 18


def _write_attention(sheet, validation, source_by_id, models_by_name) -> None:
    headers = (
        "Status",
        "Source table",
        "Odoo model",
        "Odoo field",
        "Source field",
        "Problem",
        "How to fix",
        "Support code",
    )
    rows = []
    for issue in sorted(
        validation.issues,
        key=lambda item: (
            0 if item.severity == "error" else 1,
            item.dataset_id or "",
            item.target_field or "",
            item.code,
        ),
    ):
        source = source_by_id.get(issue.dataset_id or "")
        model = models_by_name.get(issue.target_model or "")
        fields = {item.name: item for item in model.fields} if model else {}
        field = fields.get(issue.target_field or "")
        source_fields = (
            {item.stable_key: item.source_name for item in source.columns}
            if source
            else {}
        )
        rows.append(
            (
                "Must fix" if issue.severity == "error" else "Review",
                source.name if source else issue.dataset_id or "All tables",
                model.label if model else issue.target_model or "",
                field.label if field else issue.target_field or "",
                source_fields.get(
                    issue.source_column_key or "",
                    issue.source_column_key or "",
                ),
                issue.message,
                issue.remediation,
                issue.code,
            )
        )
    if not rows:
        rows.append(("Ready", "", "", "", "", "No current findings", "No action required", ""))
    _write_table_sheet(
        sheet,
        "Needs attention",
        "Fix red items and review amber items in Impodo",
        headers,
        rows,
    )
    for row_index in range(4, sheet.max_row + 1):
        status = str(sheet.cell(row_index, 1).value or "")
        style = "danger" if status == "Must fix" else "warning" if status == "Review" else "ready"
        _style_status_cell(sheet.cell(row_index, 1), style)
        if sheet.cell(row_index, 4).value:
            _style_status_cell(sheet.cell(row_index, 4), style)


def _write_field_matches(sheet, reviews: list[_FieldReview]) -> None:
    headers = (
        "Source table",
        "Odoo model",
        "Odoo field",
        "Technical field",
        "Type",
        "Requirement",
        "Value source",
        "Source field",
        "Status",
        "Next action or later check",
    )
    rows = [
        (
            item.source_dataset.name,
            item.dataset.target_model,
            item.field.label,
            item.field.name,
            item.field.type,
            item.requirement,
            item.provider,
            item.source_fields,
            item.status,
            item.next_action,
        )
        for item in reviews
    ]
    _write_table_sheet(
        sheet,
        "Field matches",
        "One row for each mapped, required, or checked Odoo field",
        headers,
        rows,
    )
    for row_index, review in enumerate(reviews, start=4):
        _style_status_cell(sheet.cell(row_index, 3), review.style)
        _style_status_cell(sheet.cell(row_index, 9), review.style)


def _write_value_coverage(sheet, validation, source_by_id, models_by_name) -> None:
    headers = (
        "Status",
        "Source table",
        "Odoo field",
        "Coverage rule",
        "Source columns",
        "Distinct values",
        "Source rows",
        "Uncovered source value",
    )
    rows: list[tuple[Any, ...]] = []
    styles: list[str] = []
    evidence = validation.categorical_coverage
    if evidence is not None:
        for result in evidence.field_results:
            source = source_by_id.get(result.dataset_id)
            definition_model = next(
                (
                    item.get("target_model")
                    for item in validation.coverage
                    if item.get("dataset_id") == result.dataset_id
                ),
                None,
            )
            model = models_by_name.get(str(definition_model or ""))
            fields = {item.name: item for item in model.fields} if model else {}
            field = fields.get(result.target_field)
            distinct_count = len(result.distinct_values)
            row_count = sum(item.count for item in result.distinct_values)
            protected_source = bool(
                source is not None and source.origin is SourceOriginKind.ODOO
            )
            uncovered = (
                ((),)
                if protected_source
                else result.uncovered_values or ((),)
            )
            for values in uncovered:
                if result.status == "UNCOVERED":
                    status, style = "Must fix", "danger"
                elif result.status == "UNSUPPORTED":
                    status, style = "Review required", "warning"
                else:
                    status, style = "Covered", "ready"
                rows.append(
                    (
                        status,
                        source.name if source else result.dataset_id,
                        field.label if field else result.target_field,
                        result.policy.replace("_", " ").title(),
                        ", ".join(result.source_column_keys),
                        distinct_count,
                        row_count,
                        (
                            "Protected values remain in Impodo"
                            if protected_source
                            else " | ".join(values)
                        ),
                    )
                )
                styles.append(style)
    if not rows:
        rows.append(("Not applicable", "", "", "", "", 0, 0, ""))
        styles.append("neutral")
    _write_table_sheet(
        sheet,
        "Value coverage",
        "Source choices checked against captured Odoo choices or business keys",
        headers,
        rows,
    )
    for row_index, style in enumerate(styles, start=4):
        _style_status_cell(sheet.cell(row_index, 1), style)
        if sheet.cell(row_index, 3).value:
            _style_status_cell(sheet.cell(row_index, 3), style)


def _write_checked_later(sheet, validation, source_by_id) -> None:
    headers = ("Status", "Source table", "Check", "Why it happens later")
    rows = [
        (
            "Not yet checked",
            (
                source_by_id[item.dataset_id].name
                if item.dataset_id in source_by_id
                else item.dataset_id
            ),
            item.message,
            "Stage 4 preparation or Stage 5 final comparison supplies the required evidence.",
        )
        for item in validation.deferred_runtime_checks
    ]
    if not rows:
        rows.append(("None listed", "", "No deferred checks are listed", ""))
    _write_table_sheet(
        sheet,
        "Checked later",
        "These results are not claimed by the Stage 3 workbook",
        headers,
        rows,
    )
    for row_index in range(4, sheet.max_row + 1):
        _style_status_cell(sheet.cell(row_index, 1), "neutral")


def _write_dataset_columns(sheet, source_dataset, reviews) -> None:
    if len(reviews) + 1 > _EXCEL_MAX_COLUMN:
        raise MappingReviewGenerationError(
            f"{source_dataset.name} has too many reviewed fields for Excel"
        )
    _title_band(
        sheet,
        source_dataset.name,
        "Colour-coded Odoo field matching decisions",
        max(len(reviews) + 1, 2),
    )
    labels = (
        "Odoo field",
        "Technical field",
        "Status",
        "Value source",
        "Source field",
        "Requirement",
        "Next action or later check",
    )
    for row_index, label in enumerate(labels, start=3):
        cell = sheet.cell(row_index, 1, label)
        cell.fill = PatternFill("solid", fgColor=_COLORS["charcoal_dark"])
        cell.font = Font(bold=True, color=_COLORS["white"])
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_index, review in enumerate(reviews, start=2):
        values = (
            review.field.label,
            review.field.name,
            review.status,
            review.provider,
            review.source_fields,
            review.requirement,
            review.next_action,
        )
        fill, font = _STYLE_COLORS[review.style]
        for row_index, value in enumerate(values, start=3):
            cell = sheet.cell(row_index, column_index, _safe_cell(value))
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(
                bold=row_index in {3, 5},
                color=font,
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
        sheet.column_dimensions[get_column_letter(column_index)].width = 28
    sheet.column_dimensions["A"].width = 24
    for row_index in range(3, 10):
        sheet.row_dimensions[row_index].height = 34 if row_index != 9 else 58
    sheet.freeze_panes = "B3"
    sheet.sheet_view.showGridLines = False


def _write_table_sheet(sheet, title, subtitle, headers, rows) -> None:
    if len(rows) > _EXCEL_MAX_ROW - 3:
        raise MappingReviewGenerationError(f"{title} exceeds the Excel row limit")
    if len(headers) > _EXCEL_MAX_COLUMN:
        raise MappingReviewGenerationError(f"{title} exceeds the Excel column limit")
    _title_band(sheet, title, subtitle, len(headers))
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(3, column_index, header)
    _style_header(sheet, 3, 1, len(headers), _COLORS["charcoal_dark"])
    for row_index, row in enumerate(rows, start=4):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row_index, column_index, _safe_cell(value))
            cell.fill = PatternFill(
                "solid",
                fgColor=(
                    _COLORS["surface"]
                    if row_index % 2 == 0
                    else _COLORS["paper"]
                ),
            )
            cell.font = Font(color=_COLORS["charcoal"])
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(sheet.max_row, 3)}"
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    for column_index, header in enumerate(headers, start=1):
        width = 18
        if header in {"Problem", "How to fix", "Next action or later check", "Check", "Why it happens later"}:
            width = 42
        elif header in {"Source table", "Odoo model", "Odoo field", "Source field", "Value source"}:
            width = 26
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _write_key_values(sheet, rows, *, start_row) -> None:
    for row_index, (label, value) in enumerate(rows, start=start_row):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, _safe_cell(value))
        sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=_COLORS["soft"])
        sheet.cell(row_index, 1).font = Font(bold=True, color=_COLORS["charcoal_dark"])
        sheet.cell(row_index, 2).font = Font(color=_COLORS["charcoal"])
        sheet.cell(row_index, 1).border = _THIN_BORDER
        sheet.cell(row_index, 2).border = _THIN_BORDER
        sheet.cell(row_index, 2).alignment = Alignment(wrap_text=True)


def _title_band(sheet, title: str, subtitle: str, width: int) -> None:
    width = max(width, 1)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=_COLORS["charcoal_dark"])
    sheet["A1"].font = Font(size=16, bold=True, color=_COLORS["white"])
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    sheet["A2"] = subtitle
    sheet["A2"].fill = PatternFill("solid", fgColor=_COLORS["paper"])
    sheet["A2"].font = Font(color=_COLORS["charcoal"])
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 28


def _style_header(sheet, row: int, start: int, end: int, color: str) -> None:
    for column in range(start, end + 1):
        cell = sheet.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=_COLORS["white"])
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 30


def _style_status_cell(cell, style: str) -> None:
    fill, font = _STYLE_COLORS[style]
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=font)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _unique_sheet_name(value: str, used: set[str]) -> str:
    base = _INVALID_SHEET_CHARACTERS.sub(" ", value).strip() or "Table fields"
    base = re.sub(r"\s+", " ", base)[:31]
    candidate = base
    suffix = 2
    while candidate in used:
        marker = f" {suffix}"
        candidate = f"{base[: 31 - len(marker)]}{marker}"
        suffix += 1
    return candidate


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(
            value,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        )
    if isinstance(value, str) and value.lstrip("\t\r\n").startswith(
        ("=", "+", "-", "@")
    ):
        return f"'{value}"
    return value
