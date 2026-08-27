"""Write portable Stage-H evidence and its data-manager review workbook.

``models.PreflightResult`` is the canonical decision source.  The JSON
manifest is written first and the workbook is then generated locally with the
same controlled Python runtime that Impodo already uses for XLSX intake.  The
workbook takes classifications and issues only from that manifest and never
feeds conclusions back into the engine.

The manifest contains business identities and portable differences only. For
file sources, an exact frozen-input projection may add prepared values so the
data manager can see what Impodo will load. Protected target snapshots—with
environment-local numeric Odoo IDs—and all Odoo-source business values remain
outside the portable workbook. The CLI writes no repository state and builds
both files from the same result.
"""

from __future__ import annotations

import csv
import json
from collections import Counter
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from impodo.domain.preflight.reports import (
    ReviewWorkbookCellEffect,
    ReviewWorkbookCellFeedback,
    ReviewWorkbookCellStatus,
    ReviewWorkbookEvidence,
    plain_readiness_guidance,
    review_workbook_cell_feedback,
)
from impodo.domain.shared.models import (
    BusinessReference,
    Classification,
    LogicalReference,
    PreflightResult,
    PreparedRecord,
    canonical_json_bytes,
)

MANIFEST_NAME = "impodo_preflight_manifest.json"
WORKBOOK_NAME = "impodo_preflight_report.xlsx"

_COLORS = {
    "brand": "E8473F",
    "brand_dark": "C93C35",
    "charcoal": "494D46",
    "charcoal_dark": "292C28",
    "gray": "868981",
    "paper": "F4F4F1",
    "surface": "FFFFFF",
    "soft": "EEEEEA",
    "line": "D6D7D2",
    "ready": "4D7C5B",
    "ready_text": "315B3B",
    "ready_bg": "EDF7EF",
    "warning": "7D4F00",
    "warning_bg": "FFF5DF",
    "danger": "9F2F2F",
    "danger_bg": "FCE8E7",
    "prepared": "2F628F",
    "prepared_bg": "EAF2FB",
    "white": "FFFFFF",
}
_THIN_BORDER = Border(
    bottom=Side(style="thin", color=_COLORS["line"]),
)
_BODY_FONT = Font(color=_COLORS["charcoal"])
_SURFACE_FILL = PatternFill("solid", fgColor=_COLORS["surface"])
_PAPER_FILL = PatternFill("solid", fgColor=_COLORS["paper"])
_EXCEL_MAX_ROW = 1_048_576
_EXCEL_MAX_COLUMN = 16_384


class ReportGenerationError(RuntimeError):
    """Raised when canonical report artifacts cannot be generated safely."""


@dataclass(frozen=True, slots=True)
class _RecordCellProjection:
    """Workbook coordinate and portable feedback for one prepared value."""

    row: int
    column: int
    feedback: ReviewWorkbookCellFeedback


@dataclass(frozen=True, slots=True)
class _RecordSheetProjection:
    """Rows, feedback coordinates, and reconciled preparation counts."""

    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]
    cells: tuple[_RecordCellProjection, ...]
    feedback_counts: Mapping[str, int]


def write_preflight_outputs(
    result: PreflightResult,
    output_directory: str | Path,
    *,
    preview_directory: str | Path | None = None,
    prepared_records: Sequence[PreparedRecord] = (),
) -> tuple[Path, Path]:
    """Write the canonical manifest and its Excel review workbook.

    The manifest and workbook use same-directory temporary files followed by
    atomic replacement.  No Node.js runtime, network request, or browser
    extension is involved.
    """

    output = Path(output_directory)
    output.mkdir(parents=True, exist_ok=True)
    manifest_path = output / MANIFEST_NAME
    workbook_path = output / WORKBOOK_NAME
    manifest = result.to_portable_dict()
    temporary_manifest = manifest_path.with_suffix(".json.partial")
    temporary_manifest.write_bytes(canonical_json_bytes(manifest) + b"\n")
    temporary_manifest.replace(manifest_path)

    _build_workbook(
        manifest_path,
        workbook_path,
        preview_directory=preview_directory,
        review_evidence=(
            ReviewWorkbookEvidence(
                frozen_input_hash="",
                records=tuple(prepared_records),
                dataset_labels={
                    item.dataset: item.dataset.replace("_", " ").title()
                    for item in prepared_records
                },
                target_model_labels={},
                target_field_labels={},
                normalization_content_hash="",
                cell_effects=(),
                target_field_required={},
            )
            if prepared_records
            else None
        ),
    )
    return manifest_path, workbook_path


def write_review_workbook(
    manifest_path: str | Path,
    workbook_path: str | Path,
    *,
    preview_directory: str | Path | None = None,
    review_evidence: ReviewWorkbookEvidence | None = None,
) -> Path:
    """Build the Excel review projection from an existing manifest."""

    manifest = Path(manifest_path)
    workbook = Path(workbook_path)
    if not manifest.is_file():
        raise ReportGenerationError("The readiness manifest does not exist")
    workbook.parent.mkdir(parents=True, exist_ok=True)
    _build_workbook(
        manifest,
        workbook,
        preview_directory=preview_directory,
        review_evidence=review_evidence,
    )
    return workbook


def _build_workbook(
    manifest_path: Path,
    workbook_path: Path,
    *,
    preview_directory: str | Path | None,
    review_evidence: ReviewWorkbookEvidence | None,
) -> None:
    """Create the business-facing workbook with the bundled Python runtime."""

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as error:
        raise ReportGenerationError("The readiness manifest is invalid") from error

    decisions = tuple(manifest.get("decisions", ()))
    prepared_by_trace = _prepared_records_by_trace(
        manifest,
        decisions,
        review_evidence,
    )
    cell_effects = _cell_effects_by_coordinate(
        manifest,
        prepared_by_trace,
        review_evidence,
    )
    record_projection = _record_sheet_projection(
        manifest,
        decisions,
        prepared_by_trace,
        review_evidence,
        cell_effects,
    )
    attention_rows = _attention_rows(
        manifest,
        decisions,
        prepared_by_trace,
        review_evidence,
    )

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Review overview"
    sheets = {
        name: workbook.create_sheet(name)
        for name in (
            "Needs attention",
            "Records to load",
            "Changes to Odoo",
            "Evidence",
        )
    }

    _write_review_overview(
        overview,
        manifest,
        attention_rows,
        record_projection.feedback_counts,
    )
    _write_data_sheet(
        sheets["Needs attention"],
        [
            "Priority",
            "Dataset",
            "Record",
            "Field",
            "What needs attention",
            "Details",
            "Next action",
            "Source row",
            "Affected records",
            "Technical code",
        ],
        attention_rows
        or [
            [
                "Ready",
                "",
                "",
                "",
                "No items require attention.",
                "Every compared record has a safe outcome.",
                "Review the prepared records before continuing to Load into Odoo.",
                "",
                0,
                "",
            ]
        ],
        _COLORS["warning"],
    )
    _style_status_cells(sheets["Needs attention"], 1)

    _write_data_sheet(
        sheets["Records to load"],
        record_projection.headers,
        record_projection.rows,
        _COLORS["ready"],
    )
    _style_status_cells(sheets["Records to load"], 1)
    _style_record_value_cells(
        sheets["Records to load"],
        record_projection.cells,
    )

    change_rows = _change_rows(
        decisions,
        prepared_by_trace,
        review_evidence,
    )
    _write_data_sheet(
        sheets["Changes to Odoo"],
        [
            "Status",
            "Dataset",
            "Record",
            "Field",
            "Value that will be loaded",
            "Current Odoo value",
            "Comparison rule",
            "Source row",
        ],
        change_rows
        or [
            [
                "No change",
                "",
                "",
                "",
                "",
                "",
                "No existing Odoo values will be changed by this review.",
                "",
            ]
        ],
        _COLORS["prepared"],
    )
    _style_status_cells(sheets["Changes to Odoo"], 1)
    _style_value_columns(sheets["Changes to Odoo"], prepared_column=5, current_column=6)

    evidence_rows = _evidence_rows(manifest, review_evidence)
    _write_data_sheet(
        sheets["Evidence"],
        [
            "Evidence",
            "Dataset",
            "Field or attribute",
            "Status",
            "Value",
            "Affected records",
            "Support details",
        ],
        evidence_rows,
        _COLORS["brand"],
    )

    temporary = workbook_path.with_name(f".{workbook_path.name}.partial")
    try:
        workbook.save(temporary)
        temporary.replace(workbook_path)
    except (OSError, ValueError) as error:
        raise ReportGenerationError(
            "Excel review workbook generation failed"
        ) from error
    finally:
        temporary.unlink(missing_ok=True)

    if preview_directory is not None:
        _write_csv_previews(workbook, Path(preview_directory))


def _write_review_overview(
    sheet,
    manifest: dict[str, Any],
    attention_rows: Sequence[Sequence[Any]],
    feedback_counts: Mapping[str, int],
) -> None:
    _title_band(
        sheet,
        "Impodo load review",
        "Read-only comparison evidence - this workbook cannot write to Odoo",
        8,
    )
    decisions = tuple(manifest.get("decisions", ()))
    counts = {
        classification: _classification_count(decisions, classification)
        for classification in (
            "CREATE",
            "UPDATE",
            "UNCHANGED",
            "AMBIGUOUS",
            "BLOCKED",
        )
    }
    warning_count = sum(row[0] == "Review recommended" for row in attention_rows)
    if counts["BLOCKED"]:
        review_status = "Cannot proceed"
        next_action = (
            "Return to Impodo and resolve every red item before checking again."
        )
    elif counts["AMBIGUOUS"]:
        review_status = "Needs attention"
        next_action = (
            "Review the amber identity matches in Impodo before checking again."
        )
    elif warning_count:
        review_status = "Review recommended"
        next_action = (
            "Review the amber item. Compare again if you change its field match."
        )
    else:
        review_status = "Ready for review"
        next_action = "Review the prepared records, then continue to Load into Odoo when approved."

    summary_rows = [
        ["Outcome", "Records"],
        ["Will create", counts["CREATE"]],
        ["Will update", counts["UPDATE"]],
        ["No change", counts["UNCHANGED"]],
        ["Needs attention", counts["AMBIGUOUS"]],
        ["Cannot proceed", counts["BLOCKED"]],
    ]
    for row_index, values in enumerate(summary_rows, start=4):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    _style_header(sheet, 4, 1, 2, _COLORS["brand"])
    classification_styles = (
        (_COLORS["ready_bg"], _COLORS["ready_text"]),
        (_COLORS["prepared_bg"], _COLORS["prepared"]),
        (_COLORS["soft"], _COLORS["charcoal"]),
        (_COLORS["warning_bg"], _COLORS["warning"]),
        (_COLORS["danger_bg"], _COLORS["danger"]),
    )
    for row_index in range(5, 10):
        fill_color, font_color = classification_styles[row_index - 5]
        for column_index in range(1, 3):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(
                bold=column_index == 1,
                color=font_color,
            )
        sheet.cell(row=row_index, column=2).number_format = "#,##0"

    target = dict(manifest.get("target", {}))
    decision_rows = [
        ["Review decision", ""],
        ["Status", review_status],
        ["Next action", next_action],
        ["Target database", target.get("database")],
        ["Odoo version", target.get("odoo_version")],
        ["Checked against Odoo", target.get("snapshot_timestamp")],
    ]
    for row_index, values in enumerate(decision_rows, start=4):
        sheet.cell(row=row_index, column=4, value=values[0])
        sheet.cell(row=row_index, column=5, value=_safe_cell(values[1]))
        sheet.merge_cells(
            start_row=row_index, start_column=5, end_row=row_index, end_column=8
        )
    _style_header(sheet, 4, 4, 8, _COLORS["charcoal_dark"])
    for row_index in range(5, 10):
        sheet.cell(row=row_index, column=4).fill = PatternFill(
            "solid", fgColor=_COLORS["soft"]
        )
        sheet.cell(row=row_index, column=4).font = Font(
            bold=True, color=_COLORS["charcoal_dark"]
        )

    preparation_rows = [
        ["Prepared value feedback", "Cells", "Meaning"],
        [
            ReviewWorkbookCellStatus.CHANGED.value,
            feedback_counts.get(ReviewWorkbookCellStatus.CHANGED.value, 0),
            "Impodo transformed the source value.",
        ],
        [
            ReviewWorkbookCellStatus.ADDED.value,
            feedback_counts.get(ReviewWorkbookCellStatus.ADDED.value, 0),
            "Impodo supplied a value through a prepared rule.",
        ],
        [
            ReviewWorkbookCellStatus.REVIEW_RECOMMENDED.value,
            feedback_counts.get(
                ReviewWorkbookCellStatus.REVIEW_RECOMMENDED.value,
                0,
            ),
            "Review the field warning in Impodo.",
        ],
        [
            ReviewWorkbookCellStatus.NEEDS_ATTENTION.value,
            feedback_counts.get(ReviewWorkbookCellStatus.NEEDS_ATTENTION.value, 0),
            "Resolve the field problem in Impodo.",
        ],
        [
            ReviewWorkbookCellStatus.EMPTY_ALLOWED.value,
            feedback_counts.get(ReviewWorkbookCellStatus.EMPTY_ALLOWED.value, 0),
            "The blank has no current blocker.",
        ],
        [
            ReviewWorkbookCellStatus.AS_PROVIDED.value,
            feedback_counts.get(ReviewWorkbookCellStatus.AS_PROVIDED.value, 0),
            "Impodo retained the prepared value.",
        ],
    ]
    for row_index, values in enumerate(preparation_rows, start=11):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    _style_header(sheet, 11, 1, 3, _COLORS["prepared"])
    preparation_styles = {
        ReviewWorkbookCellStatus.CHANGED.value: "prepared",
        ReviewWorkbookCellStatus.ADDED.value: "ready",
        ReviewWorkbookCellStatus.REVIEW_RECOMMENDED.value: "warning",
        ReviewWorkbookCellStatus.NEEDS_ATTENTION.value: "danger",
        ReviewWorkbookCellStatus.EMPTY_ALLOWED.value: "neutral",
        ReviewWorkbookCellStatus.AS_PROVIDED.value: "neutral",
    }
    for row_index in range(12, 18):
        _style_status_cell(
            sheet.cell(row=row_index, column=1),
            preparation_styles[str(sheet.cell(row=row_index, column=1).value)],
        )
        sheet.cell(row=row_index, column=2).number_format = "#,##0"
        sheet.cell(row=row_index, column=3).alignment = Alignment(wrap_text=True)

    sheet.merge_cells("D11:H11")
    sheet["D11"] = "How to use this workbook"
    for column in range(4, 9):
        sheet.cell(row=11, column=column).fill = PatternFill(
            "solid", fgColor=_COLORS["brand"]
        )
        sheet.cell(row=11, column=column).font = Font(bold=True, color=_COLORS["white"])
    sheet.merge_cells("D12:H17")
    sheet["D12"] = (
        "Start with Needs attention. In Records to load, coloured prepared "
        "cells contain only the final values. Open a changed or added cell's "
        "note for its original value and preparation rule. Correct data or "
        "rules in Impodo, not in this workbook."
    )
    sheet["D12"].fill = PatternFill("solid", fgColor=_COLORS["paper"])
    sheet["D12"].font = Font(color=_COLORS["charcoal"])
    sheet["D12"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet.row_dimensions[11].height = 28
    for row_index in range(12, 18):
        sheet.row_dimensions[row_index].height = 30

    legend = [
        ("Ready", "The record has a safe comparison outcome.", "ready"),
        ("Review recommended", "Review this item before approval.", "warning"),
        ("Cannot proceed", "Resolve this item in Impodo.", "danger"),
        ("Value to load", "This is the prepared value Impodo will use.", "prepared"),
        ("No change", "The value already matches or will not be sent.", "neutral"),
    ]
    sheet.merge_cells("A19:H19")
    sheet["A19"] = "Record and comparison status"
    sheet["A19"].fill = PatternFill("solid", fgColor=_COLORS["charcoal_dark"])
    sheet["A19"].font = Font(bold=True, color=_COLORS["white"])
    for row_index, (label, meaning, style) in enumerate(legend, start=20):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.merge_cells(
            start_row=row_index,
            start_column=2,
            end_row=row_index,
            end_column=8,
        )
        sheet.cell(row=row_index, column=2, value=meaning)
        _style_status_cell(sheet.cell(row=row_index, column=1), style)
    sheet.freeze_panes = "A3"
    for column in range(1, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 18
    sheet.column_dimensions["C"].width = 34
    sheet.column_dimensions["D"].width = 24
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_area = "A1:H24"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = _COLORS["brand"]


def _prepared_records_by_trace(
    manifest: dict[str, Any],
    decisions: Sequence[dict[str, Any]],
    evidence: ReviewWorkbookEvidence | None,
) -> dict[str, PreparedRecord]:
    if evidence is None:
        return {}
    manifest_input_hash = dict(manifest.get("preflight_evidence", {})).get(
        "frozen_input_hash"
    )
    if evidence.frozen_input_hash and manifest_input_hash != evidence.frozen_input_hash:
        raise ReportGenerationError(
            "Prepared review values do not match the readiness manifest"
        )
    records = {item.source_trace_id: item for item in evidence.records}
    if len(records) != len(evidence.records) or "" in records:
        raise ReportGenerationError("Prepared review rows have invalid trace evidence")
    decision_ids = {str(item.get("source_trace_id", "")) for item in decisions}
    if decision_ids != set(records):
        raise ReportGenerationError(
            "Prepared review rows do not match the readiness decisions"
        )
    for decision in decisions:
        record = records[str(decision.get("source_trace_id", ""))]
        if record.dataset != str(
            decision.get("dataset", "")
        ) or record.source_row != int(decision.get("source_row", 0)):
            raise ReportGenerationError(
                "Prepared review row lineage does not match the readiness decision"
            )
    return records


def _cell_effects_by_coordinate(
    manifest: dict[str, Any],
    prepared_by_trace: Mapping[str, PreparedRecord],
    evidence: ReviewWorkbookEvidence | None,
) -> dict[tuple[str, str], tuple[ReviewWorkbookCellEffect, ...]]:
    if evidence is None:
        return {}
    manifest_normalization_hash = dict(manifest.get("preflight_evidence", {})).get(
        "normalization_content_hash"
    )
    if evidence.normalization_content_hash and (
        manifest_normalization_hash != evidence.normalization_content_hash
    ):
        raise ReportGenerationError(
            "Prepared-value feedback does not match the readiness manifest"
        )
    if evidence.cell_effects and not evidence.normalization_content_hash:
        raise ReportGenerationError(
            "Prepared-value feedback has no normalization evidence"
        )
    grouped: dict[tuple[str, str], list[ReviewWorkbookCellEffect]] = {}
    for effect in evidence.cell_effects:
        record = prepared_by_trace.get(effect.source_trace_id)
        if (
            record is None
            or record.dataset != effect.dataset
            or record.source_row != effect.source_row
        ):
            raise ReportGenerationError(
                "Prepared-value feedback does not match its prepared row"
            )
        grouped.setdefault(
            (effect.source_trace_id, effect.target_field),
            [],
        ).append(effect)
    return {
        key: tuple(
            sorted(
                items,
                key=lambda item: (
                    item.rule_name.casefold(),
                    item.explanation.casefold(),
                    item.before,
                    item.after,
                ),
            )
        )
        for key, items in sorted(grouped.items())
    }


def _attention_rows(
    manifest: dict[str, Any],
    decisions: Sequence[dict[str, Any]],
    prepared_by_trace: dict[str, PreparedRecord],
    evidence: ReviewWorkbookEvidence | None,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    source_issues = tuple(manifest.get("source_issues", ()))
    for issue in source_issues:
        dataset = str(issue.get("dataset") or "")
        source_row = issue.get("row")
        decision = next(
            (
                item
                for item in decisions
                if item.get("dataset") == dataset
                and item.get("source_row") == source_row
            ),
            None,
        )
        classification = (
            _decision_classification(decision)
            if decision is not None
            else (
                Classification.BLOCKED
                if issue.get("severity") == "error"
                else Classification.CREATE
            )
        )
        reason, action = plain_readiness_guidance(
            str(issue.get("code") or ""),
            classification,
        )
        record = (
            _display_value(decision.get("business_identity", ()))
            if decision is not None
            else "All relevant records"
        )
        rows.append(
            [
                (
                    "Cannot proceed"
                    if issue.get("severity") == "error"
                    else "Review recommended"
                ),
                _dataset_label(dataset, evidence),
                record,
                _field_label(
                    dataset,
                    str(issue.get("field") or ""),
                    prepared_by_trace,
                    evidence,
                ),
                reason,
                issue.get("message"),
                action,
                source_row,
                issue.get("affected_count", 1),
                issue.get("code"),
            ]
        )

    def represented_by_source_issue(
        issue: dict[str, Any],
        decision: dict[str, Any],
    ) -> bool:
        return any(
            item.get("code") == issue.get("code")
            and (item.get("dataset") or decision.get("dataset"))
            == decision.get("dataset")
            and item.get("field") == issue.get("field")
            and item.get("row") in (None, issue.get("row"), decision.get("source_row"))
            for item in source_issues
        )

    for decision in decisions:
        issues = tuple(decision.get("issues", ()))
        added = False
        for issue in issues:
            if represented_by_source_issue(issue, decision):
                continue
            rows.append(
                _decision_attention_row(
                    decision,
                    issue,
                    prepared_by_trace,
                    evidence,
                )
            )
            added = True
        if _decision_classification(decision) is Classification.AMBIGUOUS and not added:
            rows.append(
                _decision_attention_row(
                    decision,
                    {
                        "code": "TARGET_IDENTITY_AMBIGUOUS",
                        "message": "More than one Odoo record matches this business key.",
                        "severity": "error",
                        "field": "",
                        "affected_count": 1,
                    },
                    prepared_by_trace,
                    evidence,
                )
            )
    priority = {"Cannot proceed": 0, "Needs attention": 1, "Review recommended": 2}
    return sorted(
        rows,
        key=lambda item: (
            priority.get(str(item[0]), 9),
            str(item[1]),
            int(item[7]) if isinstance(item[7], int) else 0,
            str(item[9]),
        ),
    )


def _decision_attention_row(
    decision: dict[str, Any],
    issue: dict[str, Any],
    prepared_by_trace: dict[str, PreparedRecord],
    evidence: ReviewWorkbookEvidence | None,
) -> list[Any]:
    classification = _decision_classification(decision)
    reason, action = plain_readiness_guidance(
        str(issue.get("code") or ""),
        classification,
    )
    dataset = str(decision.get("dataset") or "")
    return [
        (
            "Cannot proceed"
            if classification is Classification.BLOCKED
            or issue.get("severity") == "error"
            else (
                "Needs attention"
                if classification is Classification.AMBIGUOUS
                else "Review recommended"
            )
        ),
        _dataset_label(dataset, evidence),
        _display_value(decision.get("business_identity", ())),
        _field_label(
            dataset,
            str(issue.get("field") or ""),
            prepared_by_trace,
            evidence,
        ),
        reason,
        issue.get("message"),
        action,
        decision.get("source_row"),
        issue.get("affected_count", 1),
        issue.get("code"),
    ]


def _record_sheet_projection(
    manifest: dict[str, Any],
    decisions: Sequence[dict[str, Any]],
    prepared_by_trace: dict[str, PreparedRecord],
    evidence: ReviewWorkbookEvidence | None,
    cell_effects: Mapping[
        tuple[str, str],
        tuple[ReviewWorkbookCellEffect, ...],
    ],
) -> _RecordSheetProjection:
    value_columns = _record_value_columns(
        tuple(prepared_by_trace.values()),
        decisions,
        cell_effects,
        evidence,
    )
    global_issue_keys = {
        (
            str(item.get("code") or ""),
            str(item.get("dataset") or ""),
            str(item.get("field") or ""),
        )
        for item in manifest.get("source_issues", ())
        if item.get("row") is None
    }
    headers = [
        "Status",
        "What will happen",
        "Dataset",
        "Record type",
        "Business identity",
        "Source row",
        *(item[2] for item in value_columns),
        "Prepared value feedback",
        "Why / next action",
    ]
    dataset_models = {
        str(item.get("dataset") or ""): str(item.get("model") or "")
        for item in manifest.get("metadata_coverage", ())
    }
    rows: list[tuple[Any, ...]] = []
    cells: list[_RecordCellProjection] = []
    feedback_counts: Counter[str] = Counter()
    for decision in decisions:
        trace_id = str(decision.get("source_trace_id") or "")
        record = prepared_by_trace.get(trace_id)
        target_model = (
            record.target_model
            if record is not None
            else dataset_models.get(str(decision.get("dataset") or ""), "")
        )
        values = (
            {**record.scalar_values, **record.references} if record is not None else {}
        )
        reason, action = _decision_guidance(
            decision,
            ignored_issue_keys=global_issue_keys,
        )
        displayed_values: list[Any] = []
        row_feedback: list[tuple[str, ReviewWorkbookCellFeedback]] = []
        workbook_row = len(rows) + 4
        for value_index, (model, field, label) in enumerate(value_columns):
            if target_model != model:
                displayed_values.append("")
                continue
            value = values.get(field)
            issue = _decision_field_issue(
                decision,
                field,
                ignored_issue_keys=global_issue_keys,
            )
            feedback = review_workbook_cell_feedback(
                value,
                cell_effects.get((trace_id, field), ()),
                issue_severity=(str(issue.get("severity") or "") if issue else ""),
                issue_message=(str(issue.get("message") or "") if issue else ""),
                required=(
                    evidence is not None
                    and evidence.target_field_required.get((model, field), False)
                ),
            )
            displayed_values.append(_display_value(value))
            row_feedback.append((label, feedback))
            feedback_counts[feedback.status.value] += 1
            cells.append(
                _RecordCellProjection(
                    row=workbook_row,
                    column=7 + value_index,
                    feedback=feedback,
                )
            )
        represented_fields = {
            field for model, field, _label in value_columns if target_model == model
        }
        identity_effects = tuple(
            effect
            for (effect_trace_id, field), effects in cell_effects.items()
            if effect_trace_id == trace_id and field not in represented_fields
            for effect in effects
        )
        if identity_effects:
            identity_issue = next(
                (
                    issue
                    for field in sorted(
                        {item.target_field for item in identity_effects}
                    )
                    if (
                        issue := _decision_field_issue(
                            decision,
                            field,
                            ignored_issue_keys=global_issue_keys,
                        )
                    )
                    is not None
                ),
                None,
            )
            identity_feedback = review_workbook_cell_feedback(
                decision.get("business_identity", ()),
                identity_effects,
                issue_severity=(
                    str(identity_issue.get("severity") or "") if identity_issue else ""
                ),
                issue_message=(
                    str(identity_issue.get("message") or "") if identity_issue else ""
                ),
            )
            row_feedback.append(("Business identity", identity_feedback))
            feedback_counts[identity_feedback.status.value] += 1
            cells.append(
                _RecordCellProjection(
                    row=workbook_row,
                    column=5,
                    feedback=identity_feedback,
                )
            )
        rows.append(
            (
                _decision_status(
                    decision,
                    ignored_issue_keys=global_issue_keys,
                ),
                _classification_action(_decision_classification(decision)),
                _dataset_label(str(decision.get("dataset") or ""), evidence),
                _model_label(target_model, evidence),
                _display_value(decision.get("business_identity", ())),
                decision.get("source_row"),
                *displayed_values,
                _prepared_feedback_summary(row_feedback),
                f"{reason} {action}",
            )
        )
    if not rows:
        rows.append(
            (
                "Ready",
                "Nothing to load",
                "",
                "",
                "",
                "",
                *("" for _item in value_columns),
                "No prepared values are shown.",
                "No records were selected for this comparison.",
            )
        )
    return _RecordSheetProjection(
        headers=tuple(headers),
        rows=tuple(rows),
        cells=tuple(cells),
        feedback_counts=dict(sorted(feedback_counts.items())),
    )


def _record_value_columns(
    records: Sequence[PreparedRecord],
    decisions: Sequence[dict[str, Any]],
    cell_effects: Mapping[
        tuple[str, str],
        tuple[ReviewWorkbookCellEffect, ...],
    ],
    evidence: ReviewWorkbookEvidence | None,
) -> tuple[tuple[str, str, str], ...]:
    records_by_trace = {item.source_trace_id: item for item in records}
    pairs = {
        (record.target_model, field)
        for record in records
        for field in (*record.scalar_values, *record.references)
    }
    for decision in decisions:
        record = records_by_trace.get(str(decision.get("source_trace_id") or ""))
        if record is None:
            continue
        values = {**record.scalar_values, **record.references}
        trace_id = record.source_trace_id
        pairs.update(
            (record.target_model, str(issue.get("field") or ""))
            for issue in decision.get("issues", ())
            if issue.get("field")
            and (
                str(issue.get("field") or "") in values
                or (trace_id, str(issue.get("field") or "")) not in cell_effects
            )
        )
    pairs = sorted(pairs)
    base_labels = [
        _target_field_label(model, field, evidence) for model, field in pairs
    ]
    duplicate_labels = {
        label.casefold()
        for label in base_labels
        if sum(item.casefold() == label.casefold() for item in base_labels) > 1
    }
    columns = []
    for (model, field), label in zip(pairs, base_labels, strict=True):
        if label.casefold() in duplicate_labels:
            label = f"{label} ({_model_label(model, evidence)})"
        columns.append((model, field, label))
    return tuple(columns)


def _decision_field_issue(
    decision: dict[str, Any],
    field: str,
    *,
    ignored_issue_keys: set[tuple[str, str, str]],
) -> dict[str, Any] | None:
    issues = tuple(
        item
        for item in decision.get("issues", ())
        if str(item.get("field") or "") == field
        and (
            str(item.get("code") or ""),
            str(item.get("dataset") or decision.get("dataset") or ""),
            field,
        )
        not in ignored_issue_keys
    )
    return next(
        (item for item in issues if item.get("severity") == "error"),
        issues[0] if issues else None,
    )


def _prepared_feedback_summary(
    feedback: Sequence[tuple[str, ReviewWorkbookCellFeedback]],
) -> str:
    if not feedback:
        return "No prepared values are shown."
    notable = [
        f"{label}: {item.status.value}"
        for label, item in feedback
        if item.status is not ReviewWorkbookCellStatus.AS_PROVIDED
    ]
    as_provided = sum(
        item.status is ReviewWorkbookCellStatus.AS_PROVIDED for _label, item in feedback
    )
    if not notable:
        return "All displayed values are as provided."
    if as_provided:
        notable.append(f"{as_provided} as provided")
    return "; ".join(notable)


def _change_rows(
    decisions: Sequence[dict[str, Any]],
    prepared_by_trace: dict[str, PreparedRecord],
    evidence: ReviewWorkbookEvidence | None,
) -> list[list[Any]]:
    rows = []
    for decision in decisions:
        record = prepared_by_trace.get(str(decision.get("source_trace_id") or ""))
        target_model = record.target_model if record is not None else ""
        for difference in decision.get("differences", ()):
            rows.append(
                [
                    "Will update",
                    _dataset_label(str(decision.get("dataset") or ""), evidence),
                    _display_value(decision.get("business_identity", ())),
                    _target_field_label(
                        target_model,
                        str(difference.get("field") or ""),
                        evidence,
                    ),
                    _display_value(difference.get("proposed")),
                    _display_value(difference.get("existing")),
                    difference.get("comparison_rule"),
                    decision.get("source_row"),
                ]
            )
    return rows


def _evidence_rows(
    manifest: dict[str, Any],
    evidence: ReviewWorkbookEvidence | None,
) -> list[list[Any]]:
    target = dict(manifest.get("target", {}))
    rows = [
        [
            "Target",
            "",
            "Connection mode",
            "Checked",
            target.get("connection_mode"),
            "",
            "",
        ],
        ["Target", "", "Database", "Checked", target.get("database"), "", ""],
        ["Target", "", "Odoo version", "Checked", target.get("odoo_version"), "", ""],
        [
            "Target",
            "",
            "Snapshot timestamp",
            "Checked",
            target.get("snapshot_timestamp"),
            "",
            "",
        ],
        [
            "Run",
            "",
            "Profile",
            "Recorded",
            manifest.get("profile", {}).get("id"),
            "",
            "",
        ],
        ["Run", "", "Semantic hash", "Recorded", manifest.get("semantic_hash"), "", ""],
        [
            "Run",
            "",
            "Metadata snapshot hash",
            "Recorded",
            manifest.get("snapshot_hashes", {}).get("metadata"),
            "",
            "",
        ],
        [
            "Run",
            "",
            "Record snapshot hash",
            "Recorded",
            manifest.get("snapshot_hashes", {}).get("records"),
            "",
            "",
        ],
        [
            "Run",
            "",
            "Module versions",
            "Recorded",
            target.get("module_versions"),
            "",
            "",
        ],
        [
            "Source",
            "",
            "Source hashes",
            "Recorded",
            manifest.get("source_hashes"),
            "",
            "",
        ],
    ]
    if evidence is not None and evidence.frozen_input_hash:
        rows.append(
            [
                "Prepared values",
                "",
                "Frozen input hash",
                "Verified",
                evidence.frozen_input_hash,
                len(evidence.records),
                "The values match the prepared input used for this comparison.",
            ]
        )
    for item in manifest.get("reference_resolutions", ()):
        rows.append(
            [
                "Relationship",
                _dataset_label(str(item.get("dataset") or ""), evidence),
                str(item.get("field") or "").replace("_", " ").title(),
                str(item.get("status") or "").replace("_", " ").title(),
                _display_value(item.get("reference")),
                item.get("affected_count"),
                f"{item.get('match_count', 0)} matching record(s)",
            ]
        )
    for item in manifest.get("source_issues", ()):
        rows.append(
            [
                "Source check",
                _dataset_label(str(item.get("dataset") or ""), evidence),
                str(item.get("field") or "").replace("_", " ").title(),
                str(item.get("severity") or "").title(),
                item.get("message"),
                item.get("affected_count"),
                item.get("code"),
            ]
        )
    for item in manifest.get("metadata_coverage", ()):
        rows.append(
            [
                "Odoo fields",
                _dataset_label(str(item.get("dataset") or ""), evidence),
                item.get("model"),
                str(item.get("status") or "").title(),
                f"{item.get('available_fields', 0)} of {item.get('requested_fields', 0)} fields available",
                "",
                "",
            ]
        )
    return rows


def _decision_classification(decision: dict[str, Any]) -> Classification:
    try:
        return Classification(str(decision.get("classification") or ""))
    except ValueError as error:
        raise ReportGenerationError("The readiness decision is invalid") from error


def _decision_issue(
    decision: dict[str, Any],
    *,
    ignored_issue_keys: set[tuple[str, str, str]] | None = None,
) -> dict[str, Any] | None:
    ignored = ignored_issue_keys or set()
    issues = tuple(
        item
        for item in decision.get("issues", ())
        if (
            str(item.get("code") or ""),
            str(item.get("dataset") or decision.get("dataset") or ""),
            str(item.get("field") or ""),
        )
        not in ignored
    )
    return next(
        (item for item in issues if item.get("severity") == "error"),
        issues[0] if issues else None,
    )


def _decision_guidance(
    decision: dict[str, Any],
    *,
    ignored_issue_keys: set[tuple[str, str, str]] | None = None,
) -> tuple[str, str]:
    classification = _decision_classification(decision)
    issue = _decision_issue(decision, ignored_issue_keys=ignored_issue_keys)
    code = str(issue.get("code") or "") if issue is not None else ""
    if not code and classification is Classification.AMBIGUOUS:
        code = "TARGET_IDENTITY_AMBIGUOUS"
    return plain_readiness_guidance(code, classification)


def _decision_status(
    decision: dict[str, Any],
    *,
    ignored_issue_keys: set[tuple[str, str, str]] | None = None,
) -> str:
    classification = _decision_classification(decision)
    issue = _decision_issue(decision, ignored_issue_keys=ignored_issue_keys)
    if classification is Classification.BLOCKED or (
        issue is not None and issue.get("severity") == "error"
    ):
        return "Cannot proceed"
    if classification is Classification.AMBIGUOUS:
        return "Needs attention"
    if issue is not None and issue.get("severity") == "warning":
        return "Review recommended"
    return "Ready"


def _classification_action(classification: Classification) -> str:
    return {
        Classification.CREATE: "Create record",
        Classification.UPDATE: "Update record",
        Classification.UNCHANGED: "No change",
        Classification.AMBIGUOUS: "Held back",
        Classification.BLOCKED: "Held back",
    }[classification]


def _dataset_label(dataset: str, evidence: ReviewWorkbookEvidence | None) -> str:
    if evidence is not None and dataset in evidence.dataset_labels:
        return evidence.dataset_labels[dataset]
    return dataset.replace("_", " ").title()


def _model_label(model: str, evidence: ReviewWorkbookEvidence | None) -> str:
    if evidence is not None and model in evidence.target_model_labels:
        return evidence.target_model_labels[model]
    return model.replace(".", " ").replace("_", " ").title()


def _target_field_label(
    model: str,
    field: str,
    evidence: ReviewWorkbookEvidence | None,
) -> str:
    if evidence is not None and (model, field) in evidence.target_field_labels:
        return evidence.target_field_labels[(model, field)]
    return field.rsplit(":", 1)[-1].replace("_", " ").title()


def _field_label(
    dataset: str,
    field: str,
    prepared_by_trace: dict[str, PreparedRecord],
    evidence: ReviewWorkbookEvidence | None,
) -> str:
    model = next(
        (
            item.target_model
            for item in prepared_by_trace.values()
            if item.dataset == dataset
        ),
        "",
    )
    return _target_field_label(model, field, evidence) if field else ""


def _display_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (BusinessReference, LogicalReference)):
        return _display_value(value.key)
    if isinstance(value, dict):
        if value.get("type") in {"date", "datetime", "decimal"}:
            return value.get("value", "")
        if "key" in value:
            return _display_value(value.get("key"))
        return _json_cell(value)
    if isinstance(value, set):
        value = tuple(sorted(value, key=str))
    if isinstance(value, (list, tuple)):
        return " Â· ".join(str(_display_value(item)) for item in value)
    return value


def _style_status_cells(sheet, column: int) -> None:
    for row_index in range(4, sheet.max_row + 1):
        cell = sheet.cell(row=row_index, column=column)
        status = str(cell.value or "").casefold()
        if status in {"cannot proceed", "held back"}:
            style = "danger"
        elif status in {"needs attention", "review recommended"}:
            style = "warning"
        elif status in {"will update", "value to load"}:
            style = "prepared"
        elif status in {"ready", "will create"}:
            style = "ready"
        else:
            style = "neutral"
        _style_status_cell(cell, style)


def _style_status_cell(cell, style: str) -> None:
    colors = {
        "ready": (_COLORS["ready_bg"], _COLORS["ready_text"]),
        "warning": (_COLORS["warning_bg"], _COLORS["warning"]),
        "danger": (_COLORS["danger_bg"], _COLORS["danger"]),
        "prepared": (_COLORS["prepared_bg"], _COLORS["prepared"]),
        "neutral": (_COLORS["soft"], _COLORS["charcoal"]),
    }
    fill, font = colors[style]
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=font)


def _style_value_columns(sheet, *, prepared_column: int, current_column: int) -> None:
    for row_index in range(4, sheet.max_row + 1):
        prepared = sheet.cell(row=row_index, column=prepared_column)
        prepared.fill = PatternFill("solid", fgColor=_COLORS["prepared_bg"])
        prepared.font = Font(color=_COLORS["prepared"])
        current = sheet.cell(row=row_index, column=current_column)
        current.fill = PatternFill("solid", fgColor=_COLORS["soft"])
        current.font = Font(color=_COLORS["charcoal"])


def _style_record_value_cells(
    sheet,
    cells: Sequence[_RecordCellProjection],
) -> None:
    styles = {
        ReviewWorkbookCellStatus.CHANGED: (
            _COLORS["prepared_bg"],
            _COLORS["prepared"],
        ),
        ReviewWorkbookCellStatus.ADDED: (
            _COLORS["ready_bg"],
            _COLORS["ready_text"],
        ),
        ReviewWorkbookCellStatus.REVIEW_RECOMMENDED: (
            _COLORS["warning_bg"],
            _COLORS["warning"],
        ),
        ReviewWorkbookCellStatus.NEEDS_ATTENTION: (
            _COLORS["danger_bg"],
            _COLORS["danger"],
        ),
        ReviewWorkbookCellStatus.EMPTY_ALLOWED: (
            _COLORS["soft"],
            _COLORS["gray"],
        ),
    }
    comment_statuses = {
        ReviewWorkbookCellStatus.CHANGED,
        ReviewWorkbookCellStatus.ADDED,
        ReviewWorkbookCellStatus.REVIEW_RECOMMENDED,
        ReviewWorkbookCellStatus.NEEDS_ATTENTION,
    }
    for projection in cells:
        feedback = projection.feedback
        cell = sheet.cell(row=projection.row, column=projection.column)
        if feedback.status in styles:
            fill, font = styles[feedback.status]
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=font)
        if feedback.status in comment_statuses:
            comment = f"Status: {feedback.status.value}\n{feedback.note}"
            if feedback.original_value:
                comment = f"{comment}\nOriginal value: {feedback.original_value}"
            cell.comment = Comment(comment, "Impodo")


def _write_data_sheet(
    sheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    accent: str,
) -> None:
    if len(rows) > _EXCEL_MAX_ROW - 3:
        raise ReportGenerationError(
            "Excel review workbook exceeds the supported rows on "
            f"the {sheet.title} sheet"
        )
    if len(headers) > _EXCEL_MAX_COLUMN or any(len(row) > len(headers) for row in rows):
        raise ReportGenerationError(
            "Excel review workbook exceeds the supported columns on "
            f"the {sheet.title} sheet"
        )
    _title_band(
        sheet,
        sheet.title,
        f"{len(rows):,} review row{'s' if len(rows) != 1 else ''}",
        len(headers),
    )
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column_index, value=header)
    _style_header(sheet, 3, 1, len(headers), accent)

    for row_index, values in enumerate(rows, start=4):
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(
                row=row_index,
                column=column_index,
                value=_safe_cell(value),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
            cell.font = _BODY_FONT
            cell.fill = _SURFACE_FILL if row_index % 2 == 0 else _PAPER_FILL

    if rows:
        sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(rows) + 3}"

    for column_index, header in enumerate(headers, start=1):
        lengths = [len(header)]
        lengths.extend(
            len(str(_safe_cell(row[column_index - 1])))
            for row in rows[:100]
            if column_index <= len(row)
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            48,
            max(13, max(lengths) + 3),
        )
    sheet.freeze_panes = "A4"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape" if len(headers) > 5 else "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:3"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = accent


def _title_band(sheet, title: str, subtitle: str, columns: int) -> None:
    last_column = max(1, columns)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    sheet.cell(row=1, column=1, value=title)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sheet.cell(row=2, column=1, value=subtitle)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=_COLORS["charcoal_dark"])
        cell.font = Font(bold=True, color=_COLORS["white"], size=16)
    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor=_COLORS["soft"])
        cell.font = Font(italic=True, color=_COLORS["charcoal"])
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 24


def _style_header(
    sheet, row: int, start_column: int, end_column: int, color: str
) -> None:
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=_COLORS["white"])
        cell.border = Border(
            left=Side(style="thin", color=_COLORS["line"]),
            right=Side(style="thin", color=_COLORS["line"]),
            top=Side(style="thin", color=_COLORS["line"]),
            bottom=Side(style="thin", color=_COLORS["line"]),
        )
    sheet.row_dimensions[row].height = 30


def _classification_count(rows: Iterable[dict[str, Any]], classification: str) -> int:
    return sum(item.get("classification") == classification for item in rows)


def _decision_rows(
    decisions: Sequence[dict[str, Any]],
    classification: str,
) -> list[list[Any]]:
    return [
        [
            decision.get("dataset"),
            decision.get("source_row"),
            _json_cell(decision.get("business_identity", ())),
            _json_cell(decision.get("business_scope", ())),
            decision.get("target_match_count"),
            "; ".join(
                str(issue.get("code", "")) for issue in decision.get("issues", ())
            ),
        ]
        for decision in decisions
        if decision.get("classification") == classification
    ]


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        value = _json_cell(value)
    if isinstance(value, str) and (
        value.lstrip("\t\r\n").startswith(("=", "+", "-", "@"))
        or _looks_like_structured_identifier(value)
    ):
        return f"'{value}"
    return value


def _json_cell(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)


def _looks_like_structured_identifier(value: str) -> bool:
    parts = value.split(".")
    return len(parts) > 1 and all(part.isdigit() for part in parts)


def _write_csv_previews(workbook: Workbook, directory: Path) -> None:
    """Write dependency-free CSV previews for developer verification.

    PNG rendering was the only part of the former report path that required a
    JavaScript runtime.  CSV previews keep this optional verification output
    inspectable without adding a second executable to the workstation.
    """

    directory.mkdir(parents=True, exist_ok=True)
    for sheet in workbook.worksheets:
        safe_name = "-".join(
            token for token in sheet.title.casefold().replace("_", "-").split() if token
        )
        target = directory / f"{safe_name}.csv"
        with target.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(stream)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)


def read_manifest(path: str | Path) -> dict[str, Any]:
    """Load a previously generated canonical JSON manifest."""

    return json.loads(Path(path).read_text(encoding="utf-8"))
