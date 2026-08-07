from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
from pathlib import Path
import shutil
import tempfile
import unittest
import zipfile

from openpyxl import Workbook
from pydantic import ValidationError

from impodo.domain.compiler import compile_profile_document
from impodo.models import LogicalReference
from impodo.planner import (
    plan_metadata_requests,
    plan_preflight_requirements,
    plan_record_requests,
)
from impodo.profile import SourceSpec, load_profile
from impodo.source import (
    SourceLoadError,
    load_selected_source_table,
    open_selected_source_batches,
    prepare_sources,
    validated_xlsx_table_bounds,
)


ROOT = Path(__file__).resolve().parents[1]


class PreparedRecordTests(unittest.TestCase):
    def test_named_table_bounds_share_the_source_row_limit(self) -> None:
        self.assertEqual(
            validated_xlsx_table_bounds("A1:F16101"),
            (1, 1, 6, 16101),
        )
        with self.assertRaisesRegex(
            SourceLoadError,
            "1,048,575 possible data rows",
        ):
            validated_xlsx_table_bounds("A1:F1048576")

    def test_selected_csv_batches_match_materialized_reader(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "products.csv"
            path.write_text(
                "code,name,quantity\n"
                + "".join(
                    f"P-{index:03d},Product {index},{index}.5\n"
                    for index in range(35)
                ),
                encoding="utf-8",
            )
            materialized = load_selected_source_table(
                path,
                dataset="products",
                table_key="csv",
                encoding="utf-8",
                delimiter=",",
                header_row=1,
            )

            for batch_size in (1, 17, 1_000):
                with self.subTest(batch_size=batch_size):
                    with open_selected_source_batches(
                        path,
                        dataset="products",
                        table_key="csv",
                        encoding="utf-8",
                        delimiter=",",
                        header_row=1,
                        batch_size=batch_size,
                    ) as source:
                        batches = tuple(source.iter_batches())
                        self.assertEqual(source.headers, materialized.headers)
                        self.assertEqual(
                            source.content_hash,
                            materialized.content_hash,
                        )
                    self.assertTrue(
                        all(len(batch) <= batch_size for batch in batches)
                    )
                    self.assertEqual(
                        tuple(row for batch in batches for row in batch),
                        materialized.rows,
                    )

    def test_selected_xlsx_batches_match_materialized_reader(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "products.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Products"
            worksheet.append(["Governed export"])
            worksheet.append([])
            worksheet.append(["code", "active", "quantity"])
            for index in range(35):
                worksheet.append([f"P-{index:03d}", index % 2 == 0, index + 0.5])
            workbook.save(path)
            workbook.close()

            materialized = load_selected_source_table(
                path,
                dataset="products",
                table_key="sheet:Products",
                encoding=None,
                delimiter=None,
                header_row=3,
            )

            for batch_size in (1, 17, 1_000):
                with self.subTest(batch_size=batch_size):
                    with open_selected_source_batches(
                        path,
                        dataset="products",
                        table_key="sheet:Products",
                        encoding=None,
                        delimiter=None,
                        header_row=3,
                        batch_size=batch_size,
                    ) as source:
                        batches = tuple(source.iter_batches())
                    self.assertTrue(
                        all(len(batch) <= batch_size for batch in batches)
                    )
                    self.assertEqual(
                        tuple(row for batch in batches for row in batch),
                        materialized.rows,
                    )

    def test_bom_values_and_references_are_preserved(self) -> None:
        profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/bom.yaml")
        )
        bundle = prepare_sources(profile, ROOT / "examples/bom")
        self.assertEqual(len(bundle.records), 3)
        line = next(record for record in bundle.records if record.dataset == "bom_lines")
        self.assertEqual(line.target_identity[1], 10)
        self.assertEqual(line.scalar_values["product_qty"], Decimal("0.4500"))
        self.assertIsInstance(line.target_identity[0], LogicalReference)
        self.assertIsInstance(line.references["bom_id"], LogicalReference)
        self.assertNotIn("odoo", repr(line).casefold())

    def test_duplicate_source_identity_blocks_all_duplicates(self) -> None:
        profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/bom.yaml")
        )
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory)
            shutil.copytree(ROOT / "examples/bom", target / "input")
            lines = target / "input/bom_lines.csv"
            lines.write_text(lines.read_text() + "BOM-001-L10,BOM-001,30,CAP-X,1.0\n")
            bundle = prepare_sources(profile, target / "input")
            duplicates = [
                record
                for record in bundle.records
                if record.dataset == "bom_lines"
                and record.source_identity == ("BOM-001-L10",)
            ]
            self.assertEqual(len(duplicates), 2)
            self.assertTrue(all(record.blocked for record in duplicates))
            self.assertTrue(
                all(
                    any(
                        issue.code == "SOURCE_IDENTITY_DUPLICATE"
                        for issue in record.issues
                    )
                    for record in duplicates
                )
            )

    def test_xlsx_sheet_is_prepared_with_native_cell_types(self) -> None:
        profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/products.yaml")
        )
        products = profile.dataset("products")
        source = products.source.model_copy(
            update={
                "file": "products.xlsx",
                "sheet": "Products",
                "header_row": 3,
            }
        )
        xlsx_profile = profile.model_copy(
            update={
                "datasets": (
                    products.model_copy(update={"source": source}),
                )
            }
        )

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Products"
            worksheet.append(["Export generated for a governed test"])
            worksheet.append([])
            worksheet.append(
                [
                    "article_code",
                    "company_code",
                    "description",
                    "active",
                    "uom_code",
                ]
            )
            worksheet.append(["P-001", "BE", "Product one", True, "UNIT"])
            workbook.save(target / "products.xlsx")

            bundle = prepare_sources(xlsx_profile, target)

        self.assertEqual(len(bundle.records), 1)
        self.assertEqual(bundle.records[0].source_row, 4)
        self.assertEqual(bundle.records[0].source_identity, ("P-001", "BE"))
        self.assertIs(bundle.records[0].scalar_values["active"], True)
        self.assertEqual(set(bundle.source_hashes), {"products.xlsx"})

    def test_xlsx_formula_cells_are_rejected(self) -> None:
        profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/products.yaml")
        )
        products = profile.dataset("products")
        source = products.source.model_copy(
            update={"file": "products.xlsx", "sheet": "Products"}
        )
        xlsx_profile = profile.model_copy(
            update={
                "datasets": (
                    products.model_copy(update={"source": source}),
                )
            }
        )

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Products"
            worksheet.append(
                [
                    "article_code",
                    "company_code",
                    "description",
                    "active",
                    "uom_code",
                ]
            )
            worksheet.append(["P-001", "BE", "=1+1", True, "UNIT"])
            workbook.save(target / "products.xlsx")

            with self.assertRaisesRegex(
                SourceLoadError,
                'Excel formula found in "description" at Products!C2 in products.xlsx',
            ):
                prepare_sources(xlsx_profile, target)

    def test_xlsx_missing_sheet_and_arbitrary_zip_are_rejected(self) -> None:
        profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/products.yaml")
        )
        products = profile.dataset("products")
        source = products.source.model_copy(
            update={"file": "products.xlsx", "sheet": "Products"}
        )
        xlsx_profile = profile.model_copy(
            update={
                "datasets": (
                    products.model_copy(update={"source": source}),
                )
            }
        )

        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory)
            workbook = Workbook()
            workbook.active.title = "Other sheet"
            workbook.save(target / "products.xlsx")
            with self.assertRaisesRegex(SourceLoadError, "does not exist"):
                prepare_sources(xlsx_profile, target)

            with zipfile.ZipFile(target / "products.xlsx", "w") as archive:
                archive.writestr("not-an-office-file.txt", "data")
            with self.assertRaisesRegex(SourceLoadError, "valid XLSX container"):
                prepare_sources(xlsx_profile, target)

    def test_duplicate_csv_headers_are_rejected(self) -> None:
        profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/products.yaml")
        )
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory)
            (target / "products.csv").write_text(
                "article_code,company_code,description,active,uom_code,uom_code\n"
                "P-001,BE,Product one,true,UNIT,UNIT\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(SourceLoadError, "duplicate headers"):
                prepare_sources(profile, target)

    def test_source_profile_rejects_legacy_and_escaping_paths(self) -> None:
        with self.assertRaisesRegex(ValidationError, ".csv or .xlsx"):
            SourceSpec(file="legacy.xls")
        with self.assertRaisesRegex(ValidationError, "contained relative path"):
            SourceSpec(file="../outside.csv")
        with self.assertRaisesRegex(ValidationError, "source.sheet is required"):
            SourceSpec(file="products.xlsx")


class PlannerTests(unittest.TestCase):
    def setUp(self) -> None:
        self.profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/golden_slice.yaml")
        )
        self.bundle = prepare_sources(self.profile, ROOT / "examples/golden")
        self.plannable_records = tuple(
            record
            for record in self.bundle.records
            if record.source_identity != ("MISSING-L1",)
        )

    def test_metadata_requests_only_profile_models_and_fields(self) -> None:
        requests = plan_metadata_requests(self.profile)
        by_model = {request.model: set(request.fields) for request in requests}
        self.assertIn("product.template", by_model)
        self.assertEqual(
            by_model["uom.uom"],
            {"x_external_code"},
        )
        self.assertNotIn("message_ids", by_model["product.template"])

    def test_record_requests_are_batched_by_model(self) -> None:
        requests = plan_record_requests(self.profile, self.plannable_records)
        models = [request.model for request in requests]
        self.assertLess(len(models), len(self.bundle.records))
        product_request = next(
            request for request in requests if request.model == "product.template"
        )
        self.assertIn("default_code", product_request.fields)
        self.assertNotIn("description_sale", product_request.fields)

    def test_target_domain_restriction_is_preserved(self) -> None:
        products = self.profile.dataset("products")
        changed_products = products.model_copy(
            update={"target_domain": (["active", "=", True],)}
        )
        changed_profile = self.profile.model_copy(
            update={
                "datasets": tuple(
                    changed_products if item.name == "products" else item
                    for item in self.profile.datasets
                )
            }
        )
        request = next(
            item
            for item in plan_record_requests(changed_profile, self.plannable_records)
            if item.model == "product.template"
        )
        self.assertIn(["active", "=", True], request.domain)

    def test_datasets_for_same_model_use_only_bounded_domains(self) -> None:
        assets = self.profile.dataset("assets")
        changed_assets = assets.model_copy(
            update={
                "target": assets.target.model_copy(
                    update={"model": "res.partner"}
                )
            }
        )
        changed_profile = self.profile.model_copy(
            update={
                "datasets": tuple(
                    changed_assets if item.name == "assets" else item
                    for item in self.profile.datasets
                )
            }
        )

        requests = plan_record_requests(changed_profile, self.plannable_records)
        partner_requests = [
            item for item in requests if item.model == "res.partner"
        ]

        self.assertTrue(partner_requests)
        self.assertTrue(all(request.domain for request in partner_requests))
        domain_fields = {
            item[0]
            for request in partner_requests
            for item in request.domain
            if isinstance(item, list) and len(item) == 3
        }
        self.assertEqual(domain_fields, {"ref", "code"})

    def test_no_eligible_records_produce_no_record_reads(self) -> None:
        requirements = plan_preflight_requirements(self.profile, ())

        self.assertEqual(requirements.record_requests, ())

    def test_record_keys_are_split_into_bounded_chunks(self) -> None:
        product = next(
            record for record in self.bundle.records if record.dataset == "products"
        )
        records = tuple(
            replace(
                product,
                source_identity=(f"P-{index}", "BE"),
                target_identity=(f"P-{index}",),
            )
            for index in range(5)
        )

        requirements = plan_preflight_requirements(
            self.profile,
            records,
            maximum_keys_per_request=2,
        )
        product_requests = tuple(
            request
            for request in requirements.record_requests
            if request.model == "product.template"
        )

        self.assertEqual(len(product_requests), 3)
        self.assertTrue(all(request.domain for request in product_requests))

    def test_incoming_identity_is_narrowed_by_parent_business_key(self) -> None:
        profile = compile_profile_document(
            load_profile(ROOT / "profiles/examples/bom.yaml")
        )
        bundle = prepare_sources(profile, ROOT / "examples/bom")

        line_request = next(
            request
            for request in plan_record_requests(profile, bundle.records)
            if request.model == "mrp.bom.line"
        )

        self.assertTrue(line_request.domain)
        self.assertIn(
            "bom_id.code",
            {
                item[0]
                for item in line_request.domain
                if isinstance(item, list) and len(item) == 3
            },
        )

    def test_missing_incoming_parent_never_creates_an_unbounded_read(self) -> None:
        requests = plan_record_requests(self.profile, self.bundle.records)
        line_requests = tuple(
            request
            for request in requests
            if request.model == "x_uc.asset.line"
        )

        self.assertTrue(line_requests)
        self.assertTrue(all(request.domain for request in line_requests))
        self.assertNotIn("ASSET-MISSING", repr(line_requests))


if __name__ == "__main__":
    unittest.main()
