from __future__ import annotations

from tests.support.paths import REPOSITORY_ROOT

from datetime import datetime, timezone
from hashlib import sha256
import json
from pathlib import Path
import re
import tempfile
import unittest
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table

from impodo.application.data_version.inspection import (
    SourceFileCatalog,
    SourceInspectionError,
    SourceInspectionOptions,
    inspect_source_file,
)
from impodo.application.data_version.source_files import load_selected_source_table
from impodo.domain.preparation.source import SourceLoadError
from impodo.domain.workspace.workbench import SourceFile


ROOT = REPOSITORY_ROOT


class SourceInspectionTests(unittest.TestCase):
    def setUp(self) -> None:
        (ROOT / ".tmp").mkdir(exist_ok=True)
        self.temporary = tempfile.TemporaryDirectory(dir=ROOT / ".tmp")
        self.directory = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_csv_inventory_profiles_values_without_mapping(self) -> None:
        path = self.directory / "products.csv"
        path.write_bytes(
            (
                "code;active;amount;available_on;description\n"
                "001;true;12.50;2026-01-02;First\n"
                "002;false;7.25;;Second\n"
                "002;;7.25;2026-01-03;\n"
            ).encode()
        )

        catalog = inspect_source_file(path, source_file=_source_evidence(path))

        self.assertEqual(catalog.format, "CSV")
        self.assertEqual(catalog.encoding, "utf-8")
        self.assertEqual(catalog.delimiter, ";")
        table = catalog.tables[0]
        self.assertEqual(table.header_row, 1)
        self.assertEqual(table.row_count, 3)
        self.assertEqual(table.column_count, 5)
        profiles = {column.name: column for column in table.columns}
        self.assertEqual(profiles["code"].candidate_type, "string")
        self.assertEqual(profiles["code"].duplicate_count, 1)
        self.assertEqual(profiles["active"].candidate_type, "boolean")
        self.assertEqual(profiles["active"].null_count, 1)
        self.assertEqual(profiles["amount"].candidate_type, "decimal")
        self.assertEqual(profiles["amount"].minimum, "7.25")
        self.assertEqual(profiles["available_on"].candidate_type, "date")
        self.assertEqual(table.preview_rows[0][0], "001")

        restored = SourceFileCatalog.from_json(catalog.to_json())
        self.assertEqual(restored, catalog)

    def test_noncurrent_catalog_json_is_rejected(self) -> None:
        path = self.directory / "noncurrent.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["code", "name"])
        worksheet.append(["C1", "First"])
        worksheet.add_table(Table(displayName="Customers", ref="A1:B2"))
        workbook.save(path)
        workbook.close()
        catalog = inspect_source_file(path, source_file=_source_evidence(path))
        payload = json.loads(catalog.to_json())
        payload["contract_version"] = 1
        for table in payload["tables"]:
            for named_table in table["named_tables"]:
                named_table.pop("disposition")
                named_table.pop("message")
        noncurrent_json = json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        )

        with self.assertRaisesRegex(SourceInspectionError, "current contract"):
            SourceFileCatalog.from_json(noncurrent_json)

    def test_xlsx_inventory_finds_sheets_named_tables_and_warnings(self) -> None:
        path = self.directory / "products.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Product export"
        worksheet.append([])
        worksheet.append(["Code", "Name", "Amount", "Calculated"])
        worksheet.append(["001", "First", 12.5, "=C4*2"])
        worksheet.append(["002", "Second", 7.25, "=C5*2"])
        worksheet.add_table(Table(displayName="ProductTable", ref="A3:D5"))
        hidden = workbook.create_sheet("Reference")
        hidden.sheet_state = "hidden"
        hidden.append(["Code", "Label"])
        hidden.append(["A", "Alpha"])
        workbook.save(path)

        catalog = inspect_source_file(path, source_file=_source_evidence(path))

        self.assertEqual(catalog.format, "XLSX")
        self.assertEqual(
            [table.name for table in catalog.tables],
            ["Products", "Reference"],
        )
        products = catalog.tables[0]
        self.assertEqual(products.header_row, 3)
        self.assertEqual(products.row_count, 2)
        self.assertEqual(products.named_tables[0].display_name, "ProductTable")
        self.assertEqual(products.named_tables[0].cell_range, "A3:D5")
        self.assertEqual(products.named_tables[0].disposition, "EQUIVALENT")
        self.assertIn("combined", products.named_tables[0].message or "")
        self.assertEqual(products.merged_range_count, 1)
        self.assertEqual(products.formula_cell_count, 2)
        self.assertEqual(products.first_formula_cell, "D4")
        self.assertEqual(products.first_formula_column, "Calculated")
        self.assertTrue(any("merged range" in warning for warning in products.warnings))
        self.assertFalse(any("formula cell" in warning for warning in products.warnings))
        self.assertTrue(catalog.tables[1].hidden)
        self.assertTrue(any("is hidden" in warning for warning in catalog.warnings))

        overridden = inspect_source_file(
            path,
            source_file=_source_evidence(path),
            options=SourceInspectionOptions(
                worksheet_header_rows=(("sheet:Products", 3),),
            ),
        )
        self.assertEqual(overridden.tables[0].header_row, 3)

    def test_xlsx_inventory_streams_a_sheet_without_declared_dimensions(self) -> None:
        source = self.directory / "customers-with-dimensions.xlsx"
        path = self.directory / "customers-without-dimensions.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Customers"
        worksheet.append(["Code", "Name"])
        worksheet.append(["C001", "First customer"])
        worksheet.append(["C002", "Second customer"])
        workbook.save(source)
        workbook.close()

        removed_dimension = False
        with ZipFile(source) as reader, ZipFile(
            path,
            mode="w",
            compression=ZIP_DEFLATED,
        ) as writer:
            for member in reader.infolist():
                payload = reader.read(member.filename)
                if member.filename == "xl/worksheets/sheet1.xml":
                    payload, count = re.subn(
                        rb"<dimension[^>]*/>",
                        b"",
                        payload,
                        count=1,
                    )
                    removed_dimension = count == 1
                writer.writestr(member, payload)
        self.assertTrue(removed_dimension)

        catalog = inspect_source_file(path, source_file=_source_evidence(path))

        self.assertEqual(catalog.tables[0].header_row, 1)
        self.assertEqual(catalog.tables[0].row_count, 2)
        self.assertEqual(catalog.tables[0].column_count, 2)
        self.assertEqual(catalog.tables[0].preview_rows[0], ("C001", "First customer"))

    def test_xlsx_strict_reader_ignores_a_trailing_formatted_blank_column(self) -> None:
        path = self.directory / "stored-source-id.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PLW"
        worksheet.append(["Code", "Name"])
        worksheet.append(["A001", "First article"])
        worksheet["C1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        workbook.save(path)
        workbook.close()

        catalog = inspect_source_file(
            path,
            source_file=_source_evidence(path),
            options=SourceInspectionOptions(
                worksheet_header_rows=(("sheet:PLW", 1),),
            ),
        )
        loaded = load_selected_source_table(
            path,
            dataset="articles",
            table_key="sheet:PLW",
            encoding=None,
            delimiter=None,
            header_row=1,
            source_display_name="PLW-Article.xlsx",
        )

        self.assertEqual(catalog.tables[0].column_count, 2)
        self.assertEqual(loaded.headers, ("Code", "Name"))
        self.assertEqual(len(loaded.rows), 1)

    def test_xlsx_strict_reader_names_source_and_cell_for_empty_header(self) -> None:
        path = self.directory / "stored-source-id.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PLW"
        worksheet.append(["Code", None, "Name"])
        worksheet.append(["A001", None, "First article"])
        workbook.save(path)
        workbook.close()

        with self.assertRaises(SourceLoadError) as raised:
            load_selected_source_table(
                path,
                dataset="articles",
                table_key="sheet:PLW",
                encoding=None,
                delimiter=None,
                header_row=1,
                source_display_name="PLW-Article.xlsx",
            )

        self.assertEqual(
            str(raised.exception),
            "Source file 'PLW-Article.xlsx', sheet 'PLW', has an empty column "
            "header at B1. Add a name in B1, or remove the column if it is "
            "unused, then replace the file in Source review.",
        )

    def test_xlsx_strict_reader_names_source_when_data_exceeds_headers(self) -> None:
        path = self.directory / "stored-source-id.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PLW"
        worksheet.append(["Code", "Name"])
        worksheet.append(["A001", "First article", "Unexpected"])
        workbook.save(path)
        workbook.close()

        catalog = inspect_source_file(
            path,
            source_file=_source_evidence(path),
            options=SourceInspectionOptions(
                worksheet_header_rows=(("sheet:PLW", 1),),
            ),
        )
        with self.assertRaises(SourceLoadError) as raised:
            load_selected_source_table(
                path,
                dataset="articles",
                table_key="sheet:PLW",
                encoding=None,
                delimiter=None,
                header_row=1,
                source_display_name="PLW-Article.xlsx",
            )

        self.assertIn(
            "1 data row(s) contain cells beyond the candidate header; "
            "the first value is at C2",
            catalog.tables[0].warnings,
        )
        self.assertEqual(
            str(raised.exception),
            "Source file 'PLW-Article.xlsx', sheet 'PLW', has data in column "
            "C, but header cell C1 is empty. Add a name in C1, or remove the "
            "unexpected data from column C, then replace the file in Source "
            "review.",
        )

    def test_xlsx_inventory_keeps_a_distinct_named_table_as_an_option(self) -> None:
        path = self.directory / "distinct-table.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["Code", "Name", "Comment"])
        worksheet.append(["001", "First", "Outside the Excel table"])
        worksheet.add_table(Table(displayName="CoreData", ref="A1:B2"))
        workbook.save(path)

        catalog = inspect_source_file(path, source_file=_source_evidence(path))

        self.assertEqual([table.name for table in catalog.tables], ["Data", "CoreData"])
        self.assertEqual(catalog.tables[0].named_tables[0].disposition, "DISTINCT")
        self.assertEqual(catalog.tables[1].kind, "NAMED_TABLE")
        self.assertEqual(catalog.tables[1].column_count, 2)

    def test_xlsx_inventory_ignores_an_oversized_named_table_before_scan(self) -> None:
        path = self.directory / "oversized-table.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "PLW"
        worksheet.append(["Code", "Name", "Type"])
        worksheet.append(["001", "First", "Bottle"])
        worksheet.add_table(Table(displayName="Table34", ref="A1:B1048576"))
        workbook.save(path)

        catalog = inspect_source_file(path, source_file=_source_evidence(path))

        self.assertEqual([table.name for table in catalog.tables], ["PLW"])
        ignored = catalog.tables[0].named_tables[0]
        self.assertEqual(ignored.disposition, "INVALID")
        self.assertIn("1,048,575 possible data rows", ignored.message or "")
        self.assertEqual(catalog.tables[0].column_count, 3)

    def test_inspection_rejects_changed_registered_bytes(self) -> None:
        path = self.directory / "customers.csv"
        path.write_bytes(b"code,name\nC1,First\n")
        evidence = _source_evidence(path)
        path.write_bytes(b"code,name\nC1,Changed\n")

        with self.assertRaisesRegex(
            SourceInspectionError,
            "no longer matches",
        ):
            inspect_source_file(path, source_file=evidence)


def _source_evidence(path: Path) -> SourceFile:
    data = path.read_bytes()
    return SourceFile(
        file_id="5df764bb-25df-4a64-95ec-50eafd9635bd",
        display_name=path.name,
        stored_name=path.name,
        size_bytes=len(data),
        sha256=sha256(data).hexdigest(),
        received_at=datetime.now(timezone.utc),
    )

