from __future__ import annotations

from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from impodo.inspection import (
    SourceFileCatalog,
    SourceInspectionError,
    SourceInspectionOptions,
    inspect_source_file,
)
from impodo.projects import SourceFile


ROOT = Path(__file__).resolve().parents[1]


class SourceInspectionTests(unittest.TestCase):
    def setUp(self) -> None:
        (ROOT / ".tmp").mkdir(exist_ok=True)
        self.temporary = tempfile.TemporaryDirectory(dir=ROOT / ".tmp")
        self.directory = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_csv_inventory_profiles_values_without_mapping(self) -> None:
        path = self.directory / "products.csv"
        path.write_bytes(
            (
                "code;active;amount;available_on;description\n"
                "001;true;12.50;2026-01-02;First\n"
                "002;false;7.25;;Second\n"
                "002;;7.25;2026-01-03;\n"
            ).encode()
        )

        catalog = inspect_source_file(path, source_file=_source_evidence(path))

        self.assertEqual(catalog.format, "CSV")
        self.assertEqual(catalog.encoding, "utf-8")
        self.assertEqual(catalog.delimiter, ";")
        table = catalog.tables[0]
        self.assertEqual(table.header_row, 1)
        self.assertEqual(table.row_count, 3)
        self.assertEqual(table.column_count, 5)
        profiles = {column.name: column for column in table.columns}
        self.assertEqual(profiles["code"].candidate_type, "string")
        self.assertEqual(profiles["code"].duplicate_count, 1)
        self.assertEqual(profiles["active"].candidate_type, "boolean")
        self.assertEqual(profiles["active"].null_count, 1)
        self.assertEqual(profiles["amount"].candidate_type, "decimal")
        self.assertEqual(profiles["amount"].minimum, "7.25")
        self.assertEqual(profiles["available_on"].candidate_type, "date")
        self.assertEqual(table.preview_rows[0][0], "001")

        restored = SourceFileCatalog.from_json(catalog.to_json())
        self.assertEqual(restored, catalog)

    def test_xlsx_inventory_finds_sheets_named_tables_and_warnings(self) -> None:
        path = self.directory / "products.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Products"
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = "Product export"
        worksheet.append([])
        worksheet.append(["Code", "Name", "Amount", "Calculated"])
        worksheet.append(["001", "First", 12.5, "=C4*2"])
        worksheet.append(["002", "Second", 7.25, "=C5*2"])
        worksheet.add_table(Table(displayName="ProductTable", ref="A3:D5"))
        hidden = workbook.create_sheet("Reference")
        hidden.sheet_state = "hidden"
        hidden.append(["Code", "Label"])
        hidden.append(["A", "Alpha"])
        workbook.save(path)

        catalog = inspect_source_file(path, source_file=_source_evidence(path))

        self.assertEqual(catalog.format, "XLSX")
        self.assertEqual(
            [table.name for table in catalog.tables],
            ["Products", "ProductTable", "Reference"],
        )
        products = catalog.tables[0]
        self.assertEqual(products.header_row, 3)
        self.assertEqual(products.row_count, 2)
        self.assertEqual(products.named_tables[0].display_name, "ProductTable")
        self.assertEqual(products.named_tables[0].cell_range, "A3:D5")
        self.assertEqual(products.merged_range_count, 1)
        self.assertEqual(products.formula_cell_count, 2)
        self.assertEqual(products.first_formula_cell, "D4")
        self.assertEqual(products.first_formula_column, "Calculated")
        self.assertTrue(any("merged range" in warning for warning in products.warnings))
        self.assertFalse(any("formula cell" in warning for warning in products.warnings))
        named_table = catalog.tables[1]
        self.assertEqual(named_table.kind, "NAMED_TABLE")
        self.assertEqual(named_table.header_row, 3)
        self.assertEqual(named_table.row_count, 2)
        self.assertEqual(named_table.preview_rows[0][0], "001")
        self.assertEqual(named_table.formula_cell_count, 2)
        self.assertEqual(named_table.first_formula_cell, "D4")
        self.assertEqual(named_table.first_formula_column, "Calculated")
        self.assertTrue(catalog.tables[2].hidden)
        self.assertTrue(any("is hidden" in warning for warning in catalog.warnings))

        overridden = inspect_source_file(
            path,
            source_file=_source_evidence(path),
            options=SourceInspectionOptions(
                worksheet_header_rows=(("sheet:Products", 3),),
            ),
        )
        self.assertEqual(overridden.tables[0].header_row, 3)

    def test_inspection_rejects_changed_registered_bytes(self) -> None:
        path = self.directory / "customers.csv"
        path.write_bytes(b"code,name\nC1,First\n")
        evidence = _source_evidence(path)
        path.write_bytes(b"code,name\nC1,Changed\n")

        with self.assertRaisesRegex(
            SourceInspectionError,
            "no longer matches",
        ):
            inspect_source_file(path, source_file=evidence)


def _source_evidence(path: Path) -> SourceFile:
    data = path.read_bytes()
    return SourceFile(
        file_id="5df764bb-25df-4a64-95ec-50eafd9635bd",
        display_name=path.name,
        stored_name=path.name,
        size_bytes=len(data),
        sha256=sha256(data).hexdigest(),
        received_at=datetime.now(timezone.utc),
    )
