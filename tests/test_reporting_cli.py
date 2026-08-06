from __future__ import annotations

import json
from pathlib import Path
import tempfile
import unittest
import zipfile

from openpyxl import load_workbook

from impodo.cli import build_parser, main
from impodo.reporting import (
    MANIFEST_NAME,
    WORKBOOK_NAME,
    write_preflight_outputs,
)
from test_engine import ROOT, golden_result


class CliTests(unittest.TestCase):
    def test_help_lists_read_only_commands(self) -> None:
        help_text = build_parser().format_help()
        self.assertIn("profile", help_text)
        self.assertIn("snapshot-metadata", help_text)
        self.assertIn("snapshot-records", help_text)
        self.assertIn("preflight", help_text)
        self.assertNotIn("import", help_text.casefold())

    def test_profile_command_remains_available(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "prepared.json"
            exit_code = main(
                [
                    "profile",
                    "--profile",
                    str(ROOT / "profiles/examples/bom.yaml"),
                    "--input",
                    str(ROOT / "examples/bom"),
                    "--output",
                    str(output),
                ]
            )
            self.assertEqual(exit_code, 0)
            payload = json.loads(output.read_text())
            self.assertEqual(payload["profile"], {"id": "bom_example"})
            self.assertEqual(len(payload["records"]), 3)
            self.assertNotIn("odoo_id", output.read_text())


class WorkbookIntegrationTests(unittest.TestCase):
    def test_review_workbook_and_manifest_are_generated(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "report"
            manifest_path, workbook_path = write_preflight_outputs(
                golden_result(), output
            )
            self.assertEqual(manifest_path.name, MANIFEST_NAME)
            self.assertEqual(workbook_path.name, WORKBOOK_NAME)
            manifest_text = manifest_path.read_text()
            self.assertNotIn("odoo_id", manifest_text)
            with zipfile.ZipFile(workbook_path) as archive:
                archive_names = archive.namelist()
                workbook_xml = archive.read("xl/workbook.xml").decode()
                worksheet_xml = [
                    archive.read(name).decode()
                    for name in archive_names
                    if name.startswith("xl/worksheets/sheet")
                    and name.endswith(".xml")
                ]
            self.assertFalse(
                any(name.startswith("xl/tables/") for name in archive_names)
            )
            self.assertFalse(any("<tableParts" in xml for xml in worksheet_xml))
            self.assertTrue(
                all(xml.count("<autoFilter") <= 1 for xml in worksheet_xml)
            )
            for sheet_name in (
                "Dashboard",
                "Target",
                "Dataset Summary",
                "Proposed Creates",
                "Proposed Updates",
                "Field Differences",
                "Unchanged",
                "Ambiguous Matches",
                "Blocked Records",
                "Reference Resolution",
                "Source Issues",
                "Metadata Coverage",
            ):
                self.assertIn(sheet_name, workbook_xml)
            workbook = load_workbook(workbook_path, data_only=False)
            self.assertEqual(
                workbook["Field Differences"]["F3"].value,
                "Existing Target",
            )
            self.assertEqual(
                workbook["Field Differences"]["G3"].value,
                "Proposed Source",
            )
            self.assertIsInstance(workbook["Dashboard"]["B5"].value, int)
            self.assertEqual(workbook["Dashboard"]["A11"].value, "Start here")
            self.assertTrue(
                workbook["Dashboard"]["A1"].fill.fgColor.rgb.endswith("292C28")
            )
            for sheet in workbook.worksheets[1:]:
                self.assertEqual(len(sheet.tables), 0)
                if sheet.max_row > 3:
                    last_column = sheet.cell(
                        3,
                        sheet.max_column,
                    ).column_letter
                    self.assertEqual(
                        sheet.auto_filter.ref,
                        f"A3:{last_column}{sheet.max_row}",
                    )
                else:
                    self.assertIsNone(sheet.auto_filter.ref)
            workbook.close()

    def test_workbook_csv_previews_need_no_external_runtime(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "report"
            previews = Path(directory) / "previews"
            write_preflight_outputs(
                golden_result(),
                output,
                preview_directory=previews,
            )
            self.assertTrue((previews / "dashboard.csv").is_file())
            self.assertTrue((previews / "field-differences.csv").is_file())


if __name__ == "__main__":
    unittest.main()
