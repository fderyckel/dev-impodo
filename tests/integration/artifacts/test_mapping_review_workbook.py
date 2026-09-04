from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

from impodo.adapters.artifacts.mapping_review import (
    _field_provider,
    mapping_review_workbook_name,
    write_mapping_review_workbook,
)
from impodo.adapters.artifacts.reporting import WORKBOOK_NAME
from impodo.domain.mapping.artifacts import MappingRevision
from impodo.domain.mapping.contracts import (
    ConstantBusinessReference,
    ConstantReferenceComponent,
    DatasetMapping,
    IdentityComponentMapping,
    MappingDefinition,
    MappingTargetMode,
    RelationshipMapping,
    RelationshipResolver,
    RelationshipValueSource,
    ResolverOrigin,
    ScalarFieldMapping,
    ScalarValueSource,
    TargetFieldDisposition,
    TargetFieldHandling,
)
from impodo.domain.mapping.validation.evidence import (
    CategoricalCoverageEvidence,
    CategoricalFieldResult,
    CategoricalValueCount,
    DeferredRuntimeCheck,
    MappingValidationIssue,
    MappingValidationResult,
    MappingValidationStatus,
)
from impodo.domain.source_binding import FileSourceBinding, OdooSourceBinding
from impodo.domain.workspace.contracts import (
    OdooSchemaCatalog,
    SchemaField,
    SchemaModel,
    SchemaOrigin,
    SourceDataset,
    SourceDatasetColumn,
    SourceSelection,
)


HASH_A = "sha256:" + "a" * 64
HASH_B = "sha256:" + "b" * 64
HASH_C = "sha256:" + "c" * 64


class MappingReviewWorkbookTests(unittest.TestCase):
    def test_constant_relationship_is_not_attributed_to_a_source_field(self) -> None:
        provider, source = _field_provider(
            None,
            RelationshipMapping(
                target_field="uom_id",
                kind="many2one",
                source_column_keys=(),
                resolver=RelationshipResolver(
                    origin=ResolverOrigin.TARGET_CATALOG,
                    model="uom.uom",
                ),
                value_source=RelationshipValueSource.CONSTANT_EXISTING,
                constant_reference=ConstantBusinessReference(
                    key_values=(ConstantReferenceComponent("name", "PCE"),),
                ),
            ),
            None,
            False,
            {},
        )

        self.assertEqual(provider, "Same existing Odoo record for every row")
        self.assertEqual(source, "name=PCE")

    def test_odoo_source_business_values_stay_out_of_portable_workbook(self) -> None:
        revision, validation, selection, schema = self._evidence()
        protected_source = OdooSourceBinding(
            capture_selection_hash=HASH_A,
            model="sale.order",
            policy_hash=HASH_A,
            connection_target_hash=HASH_A,
            schema_scope_hash=HASH_A,
            read_principal_hash=HASH_A,
            read_permission_hash=HASH_A,
            context_hash=HASH_A,
        )
        selection = replace(
            selection,
            datasets=(replace(selection.datasets[0], source=protected_source),),
        )
        coverage = CategoricalCoverageEvidence(
            mapping_content_hash=revision.definition.content_hash,
            effective_source_selection_hash=selection.content_hash,
            source_snapshot_hashes=(),
            scan_contract_hash=HASH_A,
            provider_and_normalization_semantics_hash=HASH_A,
            target_schema_dependency_hash=HASH_A,
            target_reference_evidence=None,
            field_results=(
                CategoricalFieldResult(
                    path="datasets/0/fields/0",
                    dataset_id="orders",
                    target_field="state",
                    policy="EXACT_TARGET_VALUE",
                    source_column_keys=("name",),
                    distinct_values=(
                        CategoricalValueCount(
                            values=("secret-customer-value",),
                            count=3,
                        ),
                    ),
                    uncovered_values=(("secret-customer-value",),),
                    status="UNCOVERED",
                ),
            ),
        )
        validation = replace(validation, categorical_coverage=coverage)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / mapping_review_workbook_name(revision)
            write_mapping_review_workbook(
                revision,
                validation,
                selection,
                schema,
                path,
            )
            workbook = load_workbook(path, data_only=True)

        coverage_sheet = workbook["Value coverage"]
        self.assertEqual(
            coverage_sheet["H4"].value,
            "Protected values remain in Impodo",
        )
        all_text = "\n".join(
            str(cell.value or "")
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        )
        self.assertNotIn("secret-customer-value", all_text)
        workbook.close()

    def test_failed_check_marks_missing_fields_and_keeps_stage_five_separate(
        self,
    ) -> None:
        revision, validation, selection, schema = self._evidence()

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / mapping_review_workbook_name(revision)
            write_mapping_review_workbook(
                revision,
                validation,
                selection,
                schema,
                path,
            )
            workbook = load_workbook(path, data_only=True)

        self.assertEqual(WORKBOOK_NAME, "impodo_preflight_report.xlsx")
        self.assertIn("Matching overview", workbook.sheetnames)
        self.assertIn("Needs attention", workbook.sheetnames)
        self.assertIn("Field matches", workbook.sheetnames)
        self.assertIn("Checked later", workbook.sheetnames)
        self.assertIn("1 Orders fields", workbook.sheetnames)
        self.assertNotIn("Records to load", workbook.sheetnames)
        self.assertEqual(
            workbook["Matching overview"]["B4"].value,
            "Cannot confirm matches",
        )

        columns = workbook["1 Orders fields"]
        field_columns = {
            columns.cell(3, column).value: column
            for column in range(2, columns.max_column + 1)
        }
        company_column = field_columns["Company"]
        state_column = field_columns["Status"]
        name_column = field_columns["Order Reference"]
        note_column = field_columns["Notes"]

        self.assertEqual(columns.cell(5, company_column).value, "Must fix")
        self.assertTrue(
            columns.cell(3, company_column).fill.fgColor.rgb.endswith("FCE8E7")
        )
        self.assertEqual(columns.cell(5, state_column).value, "Odoo will choose")
        self.assertTrue(
            columns.cell(3, state_column).fill.fgColor.rgb.endswith("FFF5DF")
        )
        self.assertEqual(columns.cell(5, name_column).value, "Mapped")
        self.assertTrue(
            columns.cell(3, name_column).fill.fgColor.rgb.endswith("EDF7EF")
        )
        self.assertEqual(
            columns.cell(5, note_column).value,
            "Impodo supplies or prepares",
        )
        self.assertTrue(
            columns.cell(3, note_column).fill.fgColor.rgb.endswith("EAF2FB")
        )
        workbook.close()

    def _evidence(self):
        now = datetime.now(timezone.utc)
        source = SourceDataset(
            dataset_id="orders",
            name="Orders",
            source=FileSourceBinding(
                file_id="orders-file",
                table_key="Orders",
                source_sha256=HASH_A,
                catalog_hash=HASH_B,
                encoding=None,
                delimiter=None,
                header_row=1,
            ),
            row_count=3,
            columns=(
                SourceDatasetColumn(0, "External ID", "external_id", "string"),
                SourceDatasetColumn(1, "Order Reference", "name", "string"),
            ),
        )
        selection = SourceSelection(
            selection_id="selection",
            version=1,
            data_version_id="data-version",
            created_at=now,
            created_by="Data Manager",
            datasets=(source,),
            content_hash=HASH_A,
        )
        dataset = DatasetMapping(
            dataset_id="orders",
            target_model="sale.order",
            mode=MappingTargetMode.UPSERT,
            source_identity_column_keys=("external_id",),
            target_identity=(
                IdentityComponentMapping(
                    source_column_keys=("external_id",),
                    target_fields=("client_order_ref",),
                ),
            ),
            fields=(
                ScalarFieldMapping(
                    target_field="name",
                    source_column_key="name",
                ),
                ScalarFieldMapping(
                    target_field="note",
                    value_source=ScalarValueSource.CONSTANT,
                    literal_value="Imported by Impodo",
                ),
            ),
            target_field_dispositions=(
                TargetFieldDisposition(
                    target_field="state",
                    handling=TargetFieldHandling.ODOO_DEFAULT,
                ),
            ),
        )
        definition = MappingDefinition(
            mapping_id="mapping",
            source_selection_hash=selection.content_hash,
            schema_hash=HASH_C,
            datasets=(dataset,),
        )
        revision = MappingRevision(
            mapping_id="mapping",
            version=2,
            parent_version=1,
            definition=definition,
            created_at=now,
            created_by="Data Manager",
        )
        validation = MappingValidationResult(
            mapping_content_hash=definition.content_hash,
            source_selection_hash=selection.content_hash,
            schema_hash=definition.schema_hash,
            status=MappingValidationStatus.INVALID,
            issues=(
                MappingValidationIssue(
                    code="MAPPING_REQUIRED_FIELD_UNMAPPED",
                    severity="error",
                    path="datasets/0/target_model",
                    message="Required target field sale.order.company_id has no value provider.",
                    remediation="Choose incoming data or one fixed value.",
                    dataset_id="orders",
                    target_model="sale.order",
                    target_field="company_id",
                ),
            ),
            coverage=({"dataset_id": "orders", "target_model": "sale.order"},),
            deferred_runtime_checks=(
                DeferredRuntimeCheck(
                    code="TARGET_IDENTITY_UNIQUENESS",
                    dataset_id="orders",
                    message="Verify every target identity against fresh Odoo data.",
                ),
            ),
        )
        schema = OdooSchemaCatalog(
            workspace_id="workspace",
            policy_hash=HASH_A,
            captured_at=now,
            captured_by="Data Manager",
            connection_mode="REMOTE",
            database="example",
            odoo_version="19.0",
            models=(
                SchemaModel(
                    name="sale.order",
                    label="Sales Order",
                    fields=(
                        SchemaField("client_order_ref", "Customer Reference", "char", False, False, None, None, ()),
                        SchemaField("name", "Order Reference", "char", True, False, None, None, ()),
                        SchemaField("note", "Notes", "text", False, False, None, None, ()),
                        SchemaField("company_id", "Company", "many2one", True, False, "res.company", None, ()),
                        SchemaField(
                            "state",
                            "Status",
                            "selection",
                            True,
                            False,
                            None,
                            None,
                            (("draft", "Quotation"),),
                            create_default_present=True,
                            create_default_value="draft",
                        ),
                    ),
                ),
            ),
            content_hash=HASH_C,
            origin=SchemaOrigin.LIVE_API,
            read_credential_binding_hash=HASH_A,
            read_principal_hash=HASH_A,
            read_permission_hash=HASH_A,
            read_context_hash=HASH_A,
            connection_target_hash=HASH_A,
        )
        return revision, validation, selection, schema


if __name__ == "__main__":
    unittest.main()
