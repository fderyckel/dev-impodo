from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook
from tests.integration.artifacts.test_preflight_outputs import ROOT, golden_result

from impodo.web.composition.cli import build_parser, main
from impodo.domain.shared.models import Issue, PreparedRecord, Severity
from impodo.adapters.artifacts.reporting import (
    MANIFEST_NAME,
    WORKBOOK_NAME,
    write_preflight_outputs,
)


class CliTests(unittest.TestCase):
    def test_help_lists_read_only_commands(self) -> None:
        help_text = build_parser().format_help()
        self.assertIn("profile", help_text)
        self.assertIn("snapshot-metadata", help_text)
        self.assertIn("snapshot-records", help_text)
        self.assertIn("preflight", help_text)
        self.assertNotIn("import", help_text.casefold())

    def test_profile_command_remains_available(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "prepared.json"
            exit_code = main(
                [
                    "profile",
                    "--profile",
                    str(ROOT / "profiles/examples/bom.yaml"),
                    "--input",
                    str(ROOT / "examples/bom"),
                    "--output",
                    str(output),
                ]
            )
            self.assertEqual(exit_code, 0)
            payload = json.loads(output.read_text())
            self.assertEqual(payload["profile"], {"id": "bom_example"})
            self.assertEqual(len(payload["records"]), 3)
            self.assertNotIn("odoo_id", output.read_text())


class WorkbookIntegrationTests(unittest.TestCase):
    def test_review_workbook_and_manifest_are_generated(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "report"
            result = golden_result()
            manifest_path, workbook_path = write_preflight_outputs(
                result,
                output,
                prepared_records=_prepared_records(result),
            )
            self.assertEqual(manifest_path.name, MANIFEST_NAME)
            self.assertEqual(workbook_path.name, WORKBOOK_NAME)
            manifest_text = manifest_path.read_text()
            self.assertNotIn("odoo_id", manifest_text)
            with zipfile.ZipFile(workbook_path) as archive:
                archive_names = archive.namelist()
                workbook_xml = archive.read("xl/workbook.xml").decode()
                worksheet_xml = [
                    archive.read(name).decode()
                    for name in archive_names
                    if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
                ]
            self.assertFalse(
                any(name.startswith("xl/tables/") for name in archive_names)
            )
            self.assertFalse(any("<tableParts" in xml for xml in worksheet_xml))
            self.assertTrue(all(xml.count("<autoFilter") <= 1 for xml in worksheet_xml))
            for sheet_name in (
                "Review overview",
                "Needs attention",
                "Records to load",
                "Changes to Odoo",
                "Evidence",
            ):
                self.assertIn(sheet_name, workbook_xml)
            workbook = load_workbook(workbook_path, data_only=False)
            self.assertEqual(
                workbook["Changes to Odoo"]["E3"].value,
                "Value that will be loaded",
            )
            self.assertEqual(
                workbook["Changes to Odoo"]["F3"].value,
                "Current Odoo value",
            )
            self.assertIsInstance(workbook["Review overview"]["B5"].value, int)
            self.assertEqual(
                workbook["Review overview"]["A11"].value,
                "How to use this workbook",
            )
            self.assertTrue(
                workbook["Review overview"]["A1"].fill.fgColor.rgb.endswith("292C28")
            )
            records = workbook["Records to load"]
            headers = [cell.value for cell in records[3]]
            name_column = headers.index("Name") + 1
            self.assertEqual(records.cell(4, name_column).value, "Prepared PARTNER-NEW")
            self.assertEqual(records["A4"].value, "Ready")
            attention_text = " ".join(
                str(cell.value or "")
                for row in workbook["Needs attention"].iter_rows(min_row=4)
                for cell in row
            )
            self.assertIn("A related record cannot be found.", attention_text)
            self.assertIn("REFERENCE_NOT_FOUND", attention_text)
            self.assertFalse(
                any(
                    cell.data_type == "f"
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows()
                    for cell in row
                )
            )
            for sheet in workbook.worksheets[1:]:
                self.assertEqual(len(sheet.tables), 0)
                if sheet.max_row > 3:
                    last_column = sheet.cell(
                        3,
                        sheet.max_column,
                    ).column_letter
                    self.assertEqual(
                        sheet.auto_filter.ref,
                        f"A3:{last_column}{sheet.max_row}",
                    )
                else:
                    self.assertIsNone(sheet.auto_filter.ref)
            workbook.close()

    def test_workbook_csv_previews_need_no_external_runtime(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "report"
            previews = Path(directory) / "previews"
            write_preflight_outputs(
                golden_result(),
                output,
                preview_directory=previews,
            )
            self.assertTrue((previews / "review-overview.csv").is_file())
            self.assertTrue((previews / "changes-to-odoo.csv").is_file())

    def test_field_level_warning_is_not_repeated_as_a_record_defect(self) -> None:
        warning = Issue(
            code="TARGET_SELECTION_CHOICES_CHANGED",
            message=(
                "Current Odoo choices for product.template.tracking changed since mapping "
                "(0 removed, 1 added)"
            ),
            severity=Severity.WARNING,
            dataset="products",
            field="tracking",
        )
        original = golden_result()
        result = replace(
            original,
            decisions=tuple(
                (
                    replace(decision, issues=(*decision.issues, warning))
                    if decision.dataset == "products"
                    else decision
                )
                for decision in original.decisions
            ),
            issues=(*original.issues, warning),
        )
        with tempfile.TemporaryDirectory() as directory:
            _manifest_path, workbook_path = write_preflight_outputs(
                result,
                Path(directory) / "report",
                prepared_records=_prepared_records(result),
            )

            workbook = load_workbook(workbook_path, data_only=True)
            attention = workbook["Needs attention"]
            technical_codes = [
                cell.value for cell in attention["J"] if cell.row >= 4
            ]
            self.assertEqual(
                technical_codes.count("TARGET_SELECTION_CHOICES_CHANGED"),
                1,
            )
            records = workbook["Records to load"]
            product_create = next(
                row
                for row in records.iter_rows(min_row=4, values_only=True)
                if row[4] == "P-CREATE"
            )
            self.assertEqual(product_create[0], "Ready")
            self.assertEqual(product_create[-1], "Ready to create. No action needed.")
            workbook.close()


def _prepared_records(result):
    return tuple(
        PreparedRecord(
            dataset=decision.dataset,
            source_row=decision.source_row,
            target_model="res.partner",
            source_identity=decision.business_identity,
            target_identity=decision.business_identity,
            target_scope=decision.business_scope,
            scalar_values={
                "name": f"Prepared {decision.business_identity[0]}",
                "email": "reviewer@example.invalid",
            },
            references={},
            source_trace_id=decision.source_trace_id,
            issues=(),
        )
        for decision in result.decisions
    )


if __name__ == "__main__":
    unittest.main()
